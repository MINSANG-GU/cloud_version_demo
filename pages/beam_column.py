##################################################################################
##################################################################################
#Code for paper
##################################################################################
from PIL import Image as PILImage
import xlsxwriter
import os
import torch
os.environ['QT_QPA_PLATFORM'] = 'offscreen'
import cv2
import numpy as np
import streamlit as st
import subprocess
import shutil  # 결과 파일 이동을 위해 필요
from tqdm import tqdm
from huggingface_hub import hf_hub_download
from doclayout_yolo import YOLOv10
from doclayout_yolo.nn.tasks import YOLOv10DetectionModel
import json
import re
from PIL import Image
import ast  # 필요 시 literal_eval 대신 사용 가능
import csv
import requests
import asyncio
import sys
import ast
import pandas as pd
import concurrent.futures
from pdf2image import convert_from_bytes
import math
from io import BytesIO
import streamlit as st
from streamlit_drawable_canvas import st_canvas
import itertools
import glob
import unicodedata
from PIL import Image, ImageDraw, ImageEnhance
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl import Workbook
from collections import defaultdict
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.pagebreak import Break
import sys, os
import tempfile
import psutil
import gc


df_gt_scd_global = None
df_gt_sd_global = None

# sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
# from config import get_base_dir, get_ocr_results_folder

# # ✅ 저장 폴더 설정 (D:\TEST_Streamlit)
# BASE_DIR = get_base_dir()
# SURYA_RESULTS_FOLDER =get_ocr_results_folder()


if 'temp_dir' not in st.session_state:
    st.session_state.temp_dir = tempfile.mkdtemp()
    
# ✅ 저장 폴더 설정 (임시 디렉토리 기반)
BASE_DIR = st.session_state.temp_dir
SURYA_RESULTS_FOLDER = os.path.join(BASE_DIR, "surya_results")




raw_data_folder = os.path.join(BASE_DIR, "raw_data")            # 원본 이미지 저장
plain_text_folder = os.path.join(BASE_DIR, "Plain_Texts")         # plain text 저장 폴더
figure_folder = os.path.join(BASE_DIR, "Figures")                 # figure 저장 폴더
table_folder = os.path.join(BASE_DIR, "Table")
surya_output_folder = os.path.join(BASE_DIR, "Surya_output")        # OCR 결과 저장 폴더
filtered_json_folder = os.path.join(BASE_DIR, "Filtered_Output")    # 필터링된 JSON 저장 폴더
figure_ocr_output_folder = os.path.join(BASE_DIR, "Figure_OCR")      # Figure OCR 결과 저장 폴더
figure_filtered_folder = os.path.join(BASE_DIR, "Figure_Filtered")    # Figure OCR 필터링된 결과
figure_element_folder = os.path.join(BASE_DIR, "Figure_Element")     # RC Beam/Column 추출 데이터 저장
visualized_results_folder = os.path.join(figure_ocr_output_folder, "visualized_results")  # 시각화 결과 저장
OCR_Member_Extraction_folder = os.path.join(BASE_DIR, "MEMBER_EXTRACTION")
Vertical_recog_folder = os.path.join(BASE_DIR, "Vertical_recog")
Figure_Rectangles = os.path.join(BASE_DIR, "Figure_Rectangles")
Figure_Rectangles_Binarized = os.path.join(BASE_DIR, "Figure_Rectangles_Binarized")
YOLO_Rebar_Detection = os.path.join(BASE_DIR, "YOLO_Rebar_Detection")
YOLO_Rebar_Detection_csv = os.path.join(BASE_DIR, "YOLO_Rebar_Detection_csv")
Vertical_image = os.path.join(BASE_DIR, "vertical_iamge")
Vertical_txt = os.path.join(BASE_DIR, "vertical_text")
rotated_crop_folder =os.path.join(BASE_DIR,"rotated_img")
Element_extraction =os.path.join(BASE_DIR,"Element_extraction")
Table_OCR_Img = os.path.join(BASE_DIR,"Table_OCR_Img")
Table_OCR_surya =os.path.join(BASE_DIR,"Table_OCR_surya")
Table_OCR_surya_filter =os.path.join(BASE_DIR,"Table_OCR_surya_filter")
image_output_SCD =os.path.join(BASE_DIR,"image_output_SCD")
image_output_SD =os.path.join(BASE_DIR,"image_output_SD")





raw_data_folder = os.path.join(BASE_DIR, "raw_data")            # 원본 이미지 저장
plain_text_folder = os.path.join(BASE_DIR, "Plain_Texts")         # plain text 저장 폴더
figure_folder = os.path.join(BASE_DIR, "Figures")                 # figure 저장 폴더
table_folder = os.path.join(BASE_DIR, "Table")
surya_output_folder = os.path.join(BASE_DIR, "Surya_output")        # OCR 결과 저장 폴더
filtered_json_folder = os.path.join(BASE_DIR, "Filtered_Output")    # 필터링된 JSON 저장 폴더
figure_ocr_output_folder = os.path.join(BASE_DIR, "Figure_OCR")      # Figure OCR 결과 저장 폴더
figure_filtered_folder = os.path.join(BASE_DIR, "Figure_Filtered")    # Figure OCR 필터링된 결과
figure_element_folder = os.path.join(BASE_DIR, "Figure_Element")     # RC Beam/Column 추출 데이터 저장
visualized_results_folder = os.path.join(figure_ocr_output_folder, "visualized_results")  # 시각화 결과 저장
OCR_Member_Extraction_folder = os.path.join(BASE_DIR, "MEMBER_EXTRACTION")
Vertical_recog_folder = os.path.join(BASE_DIR, "Vertical_recog")
Figure_Rectangles = os.path.join(BASE_DIR, "Figure_Rectangles")
Figure_Rectangles_Binarized = os.path.join(BASE_DIR, "Figure_Rectangles_Binarized")
YOLO_Rebar_Detection = os.path.join(BASE_DIR, "YOLO_Rebar_Detection")
YOLO_Rebar_Detection_csv = os.path.join(BASE_DIR, "YOLO_Rebar_Detection_csv")
Vertical_image = os.path.join(BASE_DIR, "vertical_iamge")
Vertical_txt = os.path.join(BASE_DIR, "vertical_text")
rotated_crop_folder = os.path.join(BASE_DIR, "rotated_img")
Element_extraction = os.path.join(BASE_DIR, "Element_extraction")
Table_OCR_Img = os.path.join(BASE_DIR, "Table_OCR_Img")
Table_OCR_surya = os.path.join(BASE_DIR, "Table_OCR_surya")
Table_OCR_surya_filter = os.path.join(BASE_DIR, "Table_OCR_surya_filter")
image_output_SCD = os.path.join(BASE_DIR, "image_output_SCD")
image_output_SD = os.path.join(BASE_DIR, "image_output_SD")

# SD 관련 폴더들
raw_data_folder_SD = os.path.join(BASE_DIR, "raw_data_SD")  
table_extraction_SD = os.path.join(BASE_DIR, "table_extraction_SD")  
surya_output_SD_folder = os.path.join(BASE_DIR, "surya_output_SD_folder")  
MemberCode_extraction_ALL = os.path.join(BASE_DIR, "MemberCode_extraction_ALL")  
MemberCode_surya_ocr_SD = os.path.join(BASE_DIR, "MemberCode_surya_ocr_SD")  
SD_Element_hir = os.path.join(BASE_DIR, "SD_Element_hir")  
SD_Text = os.path.join(BASE_DIR, "SD_Text")  
SD_Drawing_Extraction = os.path.join(BASE_DIR, "SD_Drawing_Extraction")  
SD_Drawing_Extraction_Binary = os.path.join(BASE_DIR, "SD_Drawing_Extraction_Binary")  
SD_Drawing_Rebar_extraction = os.path.join(BASE_DIR, "SD_Drawing_Rebar_extraction")  
SD_Drawing_image = os.path.join(BASE_DIR, "SD_Drawing_image")
raw_data_folder_SD_processing_folder = os.path.join(raw_data_folder_SD, "processing_folder")

# ✅ 폴더 생성
folders_to_create = [
    BASE_DIR, raw_data_folder, plain_text_folder, figure_folder, table_folder,
    surya_output_folder, filtered_json_folder, figure_ocr_output_folder,
    visualized_results_folder, figure_filtered_folder, figure_element_folder,
    OCR_Member_Extraction_folder, Figure_Rectangles, Figure_Rectangles_Binarized,
    YOLO_Rebar_Detection, YOLO_Rebar_Detection_csv, raw_data_folder_SD,
    table_extraction_SD, surya_output_SD_folder, MemberCode_extraction_ALL,
    MemberCode_surya_ocr_SD, SD_Element_hir, SD_Text, SD_Drawing_Extraction,
    SD_Drawing_Extraction_Binary, SD_Drawing_Rebar_extraction, Vertical_image,
    Vertical_txt, rotated_crop_folder, Element_extraction, Table_OCR_Img,
    Table_OCR_surya, Table_OCR_surya_filter, SD_Drawing_image,
    Vertical_recog_folder, image_output_SCD, image_output_SD,
    raw_data_folder_SD_processing_folder
]

for folder in folders_to_create:
    os.makedirs(folder, exist_ok=True)

# 임시 폴더 정리 함수 (선택적으로 사용)
def cleanup_temp_dir():
    """임시 폴더를 정리합니다."""
    if 'temp_dir' in st.session_state:
        try:
            shutil.rmtree(st.session_state.temp_dir, ignore_errors=True)
            del st.session_state.temp_dir
            return True
        except Exception as e:
            st.error(f"임시 폴더 정리 중 오류: {e}")
            return False
    return False

# 디버깅용: 현재 임시 폴더 경로 확인
def show_temp_dir_info():
    """디버깅용 - 현재 임시 폴더 정보를 표시합니다."""
    if 'temp_dir' in st.session_state:
        st.info(f"현재 임시 폴더: {st.session_state.temp_dir}")
        
        # 폴더 존재 여부 확인
        if os.path.exists(st.session_state.temp_dir):
            folder_count = len([f for f in os.listdir(st.session_state.temp_dir) 
                              if os.path.isdir(os.path.join(st.session_state.temp_dir, f))])
            st.info(f"생성된 하위 폴더 수: {folder_count}")
        else:
            st.warning("임시 폴더가 존재하지 않습니다!")

# 사이드바에 관리 버튼 추가 (선택사항)
def add_temp_dir_management():
    """사이드바에 임시 폴더 관리 기능을 추가합니다."""
    with st.sidebar:
        st.subheader("🗂️ 임시 폴더 관리")
        
        if st.button("📁 폴더 정보 확인"):
            show_temp_dir_info()
        
        if st.button("🗑️ 임시 폴더 정리"):
            if cleanup_temp_dir():
                st.success("임시 폴더가 정리되었습니다!")
                st.rerun()  # 페이지 새로고침으로 새 임시 폴더 생성

# ==========================
# 저장 폴더 설정
# ==========================
raw_data_folder     = os.path.join(BASE_DIR, "raw_data")
figure_folder       = os.path.join(BASE_DIR, "Figures")

os.makedirs(raw_data_folder, exist_ok=True)
os.makedirs(figure_folder,   exist_ok=True)
# ✅ Figure 크기 기준
MIN_FIGURE_WIDTH = 1500  # figure의 최소 너비
MIN_FIGURE_HEIGHT = 1100 # figure의 최소 높이

# ✅ Plain text 최소 높이
MIN_PLAIN_TEXT_HEIGHT = 150
# ==========================
# DOCLAYNET-YOLO 모델 로드
# ==========================
torch.serialization.add_safe_globals([YOLOv10DetectionModel])
model_file = hf_hub_download(
    repo_id="juliozhao/DocLayout-YOLO-DocStructBench",
    filename="doclayout_yolo_docstructbench_imgsz1024.pt"
)
model = YOLOv10(model_file)

# ==========================
# 설정값
# ==========================
CONFIDENCE_THRESHOLDS    = {"figure":0.3, "plain text":0.5, "table":0.1}
TOP_PADDING    = 100
BOTTOM_PADDING = 30
LEFT_PADDING   = 10
RIGHT_PADDING  = 40

# ==========================
# 1) 키워드 정의 함수 (사이드바에서 바로 입력)
# ==========================
def keyword_inputs():
    code_kw  = st.text_input("Member Code Keyword")
    scope_kw = st.text_input("Section Property Keyword")
    rebar_kw = st.text_input("Rebar Pattern Keyword")
    material_kw = st.text_input("Material Data Keyword")
    return code_kw, scope_kw, rebar_kw, material_kw

# ==========================
# 2) PDF→PNG 변환 함수
# ==========================
def correct_image_simple(image_path):
    """간단하고 빠른 이미지 보정"""
    img = cv2.imread(image_path)
    if img is None:
        return False
    
    
    # 2. 간단한 품질 향상
    # 노이즈 제거
    img = cv2.bilateralFilter(img, 9, 75, 75)
    
    # 대비 향상
    img = cv2.convertScaleAbs(img, alpha=1.1, beta=5)
    
    # 보정된 이미지로 덮어쓰기
    cv2.imwrite(image_path, img)
    return True

def process_uploaded_images_with_correction(uploaded_pdfs):
    """PDF 처리 + 자동 이미지 보정 통합"""
    image_paths = []
    progress_bar = st.progress(0)
    status = st.empty()
    correction_status = st.empty()

    for idx, pdf in enumerate(uploaded_pdfs):
        expected_img_path = os.path.join(raw_data_folder, f"SCD_{idx+1}_page_1.png")
        
        # 기존 이미지가 있는지 확인
        if os.path.exists(expected_img_path):
            st.info(f"🖼️ {os.path.basename(expected_img_path)} 이미 존재 → 변환 생략")
            existing_imgs = sorted(
                [os.path.join(raw_data_folder, f) for f in os.listdir(raw_data_folder)
                 if f.startswith(f"SCD_{idx+1}_page_") and f.endswith(".png")]
            )
            
            # 기존 이미지들도 보정 적용 (한 번만)
            correction_status.text(f"🔧 기존 이미지 보정 중...")
            corrected_count = 0
            for img_path in existing_imgs:
                # 보정 여부 체크 (파일명에 _corrected 추가하거나 별도 플래그 파일 사용)
                correction_flag = img_path.replace('.png', '_corrected.flag')
                if not os.path.exists(correction_flag):
                    if correct_image_simple(img_path):
                        # 보정 완료 플래그 생성
                        with open(correction_flag, 'w') as f:
                            f.write('corrected')
                        corrected_count += 1
            
            if corrected_count > 0:
                correction_status.text(f"✅ {corrected_count}개 이미지 보정 완료")
            else:
                correction_status.text(f"ℹ️ 이미지 보정 이미 완료됨")
                
            image_paths.extend(existing_imgs)
            progress_bar.progress((idx + 1) / len(uploaded_pdfs))
            continue

        # PDF → PNG 변환
        status.text(f"📄 {pdf.name} 변환 중...")
        pages = convert_from_bytes(pdf.getvalue(), dpi=300)
        
        current_pdf_images = []
        for pg, img in enumerate(pages):
            out_path = os.path.join(raw_data_folder, f"SCD_{idx+1}_page_{pg+1}.png")
            img.save(out_path, "PNG")
            current_pdf_images.append(out_path)
            image_paths.append(out_path)

        # 변환 완료 후 즉시 이미지 보정
        correction_status.text(f"🔧 {pdf.name} 이미지 보정 중...")
        corrected_count = 0
        failed_count = 0
        
        for img_path in current_pdf_images:
            if correct_image_simple(img_path):
                corrected_count += 1
                # 보정 완료 플래그 생성
                correction_flag = img_path.replace('.png', '_corrected.flag')
                with open(correction_flag, 'w') as f:
                    f.write('corrected')
            else:
                failed_count += 1

        correction_status.text(f"✅ {pdf.name}: 보정 성공 {corrected_count}개, 실패 {failed_count}개")
        progress_bar.progress((idx + 1) / len(uploaded_pdfs))
        status.text(f"📄 {pdf.name} 완료 ({len(pages)}페이지)")

    st.success(f"✅ 총 {len(image_paths)}개 이미지 처리 완료 (변환 + 보정)")
    st.markdown("**🔍 처리 결과 요약**")
    st.code("\n".join(image_paths[:5]), language="text")
    
    # 보정 통계 표시
    corrected_files = [f for f in os.listdir(raw_data_folder) if f.endswith('_corrected.flag')]
    if corrected_files:
        st.info(f"🔧 총 {len(corrected_files)}개 이미지가 자동 보정되었습니다.")
    
    return image_paths

# 기존 함수 이름을 바꾸고 새 함수를 기본으로 사용
def process_uploaded_images(uploaded_pdfs):
    """기본 함수 - 보정 기능 포함"""
    return process_uploaded_images_with_correction(uploaded_pdfs)

# 보정 없이 원본 기능만 원할 때 사용
def process_uploaded_images_original(uploaded_pdfs):
    """원본 기능 (보정 없음)"""
    image_paths = []
    progress_bar = st.progress(0)
    status = st.empty()

    for idx, pdf in enumerate(uploaded_pdfs):
        expected_img_path = os.path.join(raw_data_folder, f"SCD_{idx+1}_page_1.png")
        if os.path.exists(expected_img_path):
            st.info(f"🖼️ {os.path.basename(expected_img_path)} 이미 존재 → 변환 생략")
            existing_imgs = sorted(
                [os.path.join(raw_data_folder, f) for f in os.listdir(raw_data_folder)
                 if f.startswith(f"SCD_{idx+1}_page_") and f.endswith(".png")]
            )
            image_paths.extend(existing_imgs)
            progress_bar.progress((idx + 1) / len(uploaded_pdfs))
            continue

        pages = convert_from_bytes(pdf.getvalue(), dpi=300)
        for pg, img in enumerate(pages):
            out_path = os.path.join(raw_data_folder, f"SCD_{idx+1}_page_{pg+1}.png")
            img.save(out_path, "PNG")
            image_paths.append(out_path)

        progress_bar.progress((idx + 1) / len(uploaded_pdfs))
        status.text(f"📄 {pdf.name} 변환 완료 ({len(pages)}페이지)")

    st.success(f"✅ total {len(image_paths)} Complete")
    st.markdown("**🔍 results summary")
    st.code("\n".join(image_paths[:5]), language="text")
    return image_paths

# ==========================
# 3) DOCLAYNET-YOLO 적용 함수
# ==========================
def apply_yolo_on_images(image_paths=None):
    raw_data_folder = os.path.join(BASE_DIR, "raw_data")
    if image_paths is None:
        image_paths = [
            os.path.join(raw_data_folder, f)
            for f in os.listdir(raw_data_folder)
            if f.lower().endswith('.png')
        ]

    os.makedirs(plain_text_folder, exist_ok=True)
    os.makedirs(figure_folder, exist_ok=True)
    os.makedirs(table_folder, exist_ok=True)

    results = {"plain_text": {}, "figure": {}, "table": {}}
    device = "cuda" if torch.cuda.is_available() else "cpu"
    progress_bar = st.progress(0)
    status_text = st.empty()

    for idx, img_path in enumerate(image_paths):
        det = model.predict(img_path, imgsz=1024, conf=0.1, iou=0.6, device=device)[0]
        boxes, names = det.boxes, det.names
        img = cv2.imread(img_path)
        H, W = img.shape[:2]

        plain_boxes, fig_boxes, tbl_boxes = [], [], []
        for b in boxes:
            cls_id, conf = int(b.cls.item()), b.conf.item()
            name = names[cls_id]
            x1, y1, x2, y2 = map(int, b.xyxy[0].tolist())

            if (name == "plain text" and conf >= CONFIDENCE_THRESHOLDS["plain text"] and (y2 - y1) >= MIN_PLAIN_TEXT_HEIGHT):
                plain_boxes.append((x1, y1, x2, y2))
            elif name == "figure" and conf >= CONFIDENCE_THRESHOLDS["figure"]:
                fig_boxes.append((x1, y1, x2, y2))
            elif name == "table" and conf >= CONFIDENCE_THRESHOLDS["table"]:
                tbl_boxes.append((x1, y1, x2, y2))

        filtered = [b for i, b in enumerate(fig_boxes)
                    if all(not (ox1 <= b[0] <= b[2] <= ox2 and oy1 <= b[1] <= b[3] <= oy2)
                           for j, (ox1, oy1, ox2, oy2) in enumerate(fig_boxes) if i != j)]

        if filtered:
            x1, y1, x2, y2 = sorted(filtered, key=lambda b: b[1])[0] if len(filtered) > 1 else filtered[0]
            x1p, y1p = max(x1 - LEFT_PADDING, 0), max(y1 - TOP_PADDING, 0)
            x2p, y2p = min(x2 + RIGHT_PADDING, W), min(y2 + BOTTOM_PADDING, H)
            crop = img[y1p:y2p, x1p:x2p]
            out_fig = os.path.join(figure_folder, f"SCD_{idx+1}_figure.png")
            cv2.imwrite(out_fig, crop)
            results["figure"][img_path] = out_fig

        for i, (x1, y1, x2, y2) in enumerate(plain_boxes):
            # 마진 추가 (5-10픽셀 정도)
            PLAIN_TEXT_MARGIN = 50
            x1_margin = max(x1 - PLAIN_TEXT_MARGIN, 0)
            y1_margin = max(y1 - PLAIN_TEXT_MARGIN, 0) 
            x2_margin = min(x2 + PLAIN_TEXT_MARGIN, W)
            y2_margin = min(y2 + PLAIN_TEXT_MARGIN, H)
            
            crop = img[y1_margin:y2_margin, x1_margin:x2_margin]
            out_txt = os.path.join(plain_text_folder, f"SCD_{idx+1}_plain_{i}.png")
            cv2.imwrite(out_txt, crop)
            results["plain_text"].setdefault(img_path, []).append(out_txt)

        for i, (x1, y1, x2, y2) in enumerate(tbl_boxes):
            crop = img[y1:y2, x1:x2]
            out_tbl = os.path.join(table_folder, f"SCD_{idx+1}_table_{i}.png")
            cv2.imwrite(out_tbl, crop)
            results["table"].setdefault(img_path, []).append(out_tbl)

        progress_bar.progress((idx + 1) / len(image_paths))
        status_text.text(f"📌 YOLO 분석 중: {os.path.basename(img_path)}")

    progress_bar.empty()
    st.success("✅ Complete!")
    st.markdown("**results summary")
    st.json({k: list(v.values())[:3] if isinstance(v, dict) else v for k, v in results.items()})
    return results





# ✅ 3. Surya OCR 적용 (진행상황 + 캐시 확인 + 결과 이동)
def apply_surya_ocr():
    import os
    import shutil
    import subprocess
    

    os.makedirs(surya_output_folder, exist_ok=True)
    # 기존 결과 확인
    existing_jsons = [f for f in os.listdir(surya_output_folder) if f.endswith(".json")]
    if existing_jsons:
        st.info(f"이미 {len(existing_jsons)}개의 OCR 결과가 존재합니다. Surya OCR 생략.")
        return

    image_files = [f for f in os.listdir(plain_text_folder) if f.endswith(('.jpg', '.png'))]
    if not image_files:
        st.error("No files to apply OCR")
        return

    # surya_ocr 결과 저장 위치 확인
    st.info(f"SURYA_RESULTS_FOLDER 경로: {SURYA_RESULTS_FOLDER}")
    st.info(f"SURYA_RESULTS_FOLDER 존재 여부: {os.path.exists(SURYA_RESULTS_FOLDER)}")
    
    # OCR 실행 전 현재 디렉토리 상태 확인
    current_dir_before = os.listdir('.')
    st.info(f"OCR 실행 전 현재 디렉토리 파일 수: {len(current_dir_before)}")

    progress_bar = st.progress(0)
    status_text = st.empty()
    st.write("Running OCR...")

    for idx, image_file in enumerate(image_files):
        input_path = os.path.join(plain_text_folder, image_file)
        st.info(f"OCR 실행: {input_path}")
        
        command = ["surya_ocr", input_path]
        result = subprocess.run(command, capture_output=True, text=True, cwd='.')
        
        st.info(f"OCR 명령어 결과 - 반환코드: {result.returncode}")
        if result.stdout:
            st.info(f"STDOUT: {result.stdout[:200]}")
        
        stderr = result.stderr.strip()
        if stderr:
            if all(x not in stderr for x in ["Detecting bboxes", "Recognizing Text"]):
                st.warning(f"STDERR: {stderr}")
            else:
                st.info(f"OCR 진행 상황: {stderr}")

        # OCR 실행 후 파일 시스템 변화 확인
        current_dir_after = os.listdir('.')
        new_files = set(current_dir_after) - set(current_dir_before)
        if new_files:
            st.info(f"새로 생성된 파일/폴더: {list(new_files)}")
        
        # 가능한 결과 저장 위치들 확인
        possible_locations = [
            SURYA_RESULTS_FOLDER,
            './results',
            './output',
            '.',
            '/tmp',
            f'{BASE_DIR}/results'
        ]
        
        for location in possible_locations:
            if os.path.exists(location):
                files_in_location = os.listdir(location)
                if files_in_location:
                    st.info(f"{location}에 있는 파일들: {files_in_location}")

        progress = int((idx + 1) / len(image_files) * 100)
        progress_bar.progress(progress)
        status_text.write(f"running ocr: {image_file} ({progress}%)")

    progress_bar.empty()
    status_text.write("OCR complete! 결과 파일 위치 확인 중...")

    # OCR 완료 후 결과 파일 찾기
    st.info("결과 파일 검색 중...")
    
    # SURYA_RESULTS_FOLDER 확인
    if os.path.exists(SURYA_RESULTS_FOLDER):
        result_contents = os.listdir(SURYA_RESULTS_FOLDER)
        st.info(f"SURYA_RESULTS_FOLDER 내용: {result_contents}")
        
        moved, skipped = 0, 0
        for folder_name in result_contents:
            folder_path = os.path.join(SURYA_RESULTS_FOLDER, folder_name)
            if os.path.isdir(folder_path):
                folder_contents = os.listdir(folder_path)
                st.info(f"{folder_name} 폴더 내용: {folder_contents}")
                
                json_file = os.path.join(folder_path, "results.json")
                if os.path.exists(json_file):
                    dst_file = os.path.join(surya_output_folder, f"{folder_name}.json")
                    try:
                        shutil.move(json_file, dst_file)
                        moved += 1
                        st.info(f"파일 이동 성공: {folder_name}.json")
                    except Exception as e:
                        st.error(f"Move Error: {folder_name} → {e}")
                else:
                    skipped += 1
                    st.warning(f"results.json 없음: {folder_name}")
        
        st.success(f"OCR 결과 {moved}개 이동 완료 ({skipped}개는 누락됨)")
    else:
        st.error(f"SURYA_RESULTS_FOLDER가 존재하지 않음: {SURYA_RESULTS_FOLDER}")



        














# ✅ 4. JSON 필터링 (진행상황 + 조건 미충족 표시 + 결과 파일 개수)
def apply_filtering():
    at_number_pattern = re.compile(r'@\s*\(\s*\d+\s*\)')
    json_files = [(root, f) for root, _, files in os.walk(surya_output_folder) for f in files if f.endswith('.json')]
    if not json_files:
        st.error("❌ No JSON file exists to filter.")
        return

    os.makedirs(filtered_json_folder, exist_ok=True)
    progress_bar = st.progress(0)
    status_text = st.empty()
    saved, skipped = 0, 0

    for idx, (folder_path, file_name) in enumerate(json_files):
        input_file = os.path.join(folder_path, file_name)
        try:
            if os.path.getsize(input_file) == 0:
                skipped += 1
                continue

            with open(input_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            filtered_data, contains_target = [], False
            for values in data.values():
                if isinstance(values, list):
                    for item in values:
                        for line in item.get("text_lines", []):
                            text, bbox = line.get("text"), line.get("bbox")
                            if text and bbox:
                                filtered_data.append({"text": text, "bbox": bbox})
                                if ("Member" in text or "Section Property" in text or at_number_pattern.search(text)):
                                    contains_target = True

            if contains_target:
                base = os.path.splitext(file_name)[0]
                out_path = os.path.join(filtered_json_folder, f"{base}_results.json")
                with open(out_path, "w", encoding="utf-8") as out_f:
                    json.dump(filtered_data, out_f, indent=4, ensure_ascii=False)
                saved += 1
            else:
                skipped += 1

        except Exception as e:
            st.error(f"❌ error ({file_name}): {e}")

        progress = int((idx + 1) / len(json_files) * 100)
        progress_bar.progress(progress)
        status_text.text(f"📝 filtering: {file_name} ({progress}%)")

    progress_bar.empty()
    st.success(f"✅ fitering complete: {saved} are saved, {skipped} are excluded ")


def apply_keyword_extraction():
    global code_kw, scope_kw, rebar_kw, material_kw

    os.makedirs(MemberCode_extraction_ALL, exist_ok=True)
    y_tol = 10  # y축 정렬 허용 오차

    def normalize(s):
        return re.sub(r"\s+", "", s).lower()
    
    key_code = normalize(code_kw)
    key_scope = normalize(scope_kw)
    key_rebar = normalize(rebar_kw)
    key_material = normalize(material_kw)

    files = [f for f in os.listdir(filtered_json_folder) if f.endswith('_results.json')]
    for file_name in tqdm(files, desc="키워드 기반 추출"):
        data = json.load(open(os.path.join(filtered_json_folder, file_name), encoding='utf-8'))

        # 중심 좌표 계산
        entries = []
        for item in data:
            x1, y1, x2, y2 = item['bbox']
            cx, cy = (x1 + x2) / 2, (y1 + y2) / 2
            entries.append({**item, 'cx': cx, 'cy': cy})

        def extract_for_keyword(raw_kw, norm_kw):
            hits = []
            for e in entries:
                norm_txt = normalize(e['text'])
                if norm_kw and norm_kw in norm_txt:
                    # 키워드 제거를 더 유연하게 처리
                    original_text = e['text']
                    
                    # 방법 1: 원본 키워드로 제거 (대소문자 무관)
                    rem1 = re.sub(re.escape(raw_kw), '', original_text, flags=re.IGNORECASE).strip()
                    
                    # 방법 2: 정규화된 키워드 위치 찾아서 원본에서 제거
                    norm_original = normalize(original_text)
                    if norm_kw in norm_original:
                        # 정규화된 텍스트에서 키워드 위치 찾기
                        start_idx = norm_original.find(norm_kw)
                        if start_idx != -1:
                            # 원본 텍스트에서 해당 부분의 실제 길이 계산
                            # 정규화 과정에서 제거된 공백들을 고려해야 함
                            before_kw = original_text[:len(original_text)]  # 임시
                            
                            # 더 간단한 방법: 키워드 앞뒤 패턴으로 분할
                            # 키워드를 공백이나 구분자를 포함하여 유연하게 매칭
                            flexible_pattern = re.escape(raw_kw).replace(r'\ ', r'\s*')
                            rem2 = re.sub(flexible_pattern, '', original_text, flags=re.IGNORECASE).strip()
                        else:
                            rem2 = rem1
                    else:
                        rem2 = rem1
                    
                    # 더 깔끔하게 제거된 결과 선택
                    candidates = [rem1, rem2]
                    best_rem = min([r for r in candidates if r], key=len, default='')
                    
                    # 구분자 및 불필요한 문자 정리
                    best_rem = re.sub(r'^[\s:,\-=]+|[\s:,\-=]+$', '', best_rem).strip()
                    
                    # 의미있는 텍스트가 남아있는지 확인
                    if best_rem and len(best_rem) > 1 and best_rem != raw_kw.strip():
                        # 2번 케이스: 키워드와 함께 정보가 있는 경우
                        hits.append({'text': best_rem, 'bbox': e['bbox'], 'source': file_name})
                    else:
                        # 1번 케이스: 키워드만 있는 경우, 인접 텍스트 찾기
                        cands = [o for o in entries if abs(o['cy'] - e['cy']) < y_tol and o is not e]
                        if cands:
                            nearest = min(cands, key=lambda o: abs(o['cx'] - e['cx']))
                            hits.append({'text': nearest['text'], 'bbox': nearest['bbox'], 'source': file_name})
            return hits

        result = {
            'Member Code':     extract_for_keyword(code_kw,    key_code),
            'Section Property':extract_for_keyword(scope_kw,   key_scope),
            'Rebar Pattern':   extract_for_keyword(rebar_kw,   key_rebar),
            'Material':        extract_for_keyword(material_kw,key_material)
        }

        output_path = os.path.join(MemberCode_extraction_ALL, file_name)
        with open(output_path, 'w', encoding='utf-8') as fout:
            json.dump(result, fout, ensure_ascii=False, indent=4)

    st.success(f"✅ 키워드 기반 추출 완료! 결과가 {MemberCode_extraction_ALL}에 저장되었습니다!")


# ✅ 4. Figure OCR 적용 (진행상황 + 결과 로그)
def apply_surya_ocr_to_figures():
    st.write("🔍 Run OCR to figures")
    image_files = [f for f in os.listdir(figure_folder) if f.endswith(('.jpg', '.png'))]
    if not image_files:
        st.error("❌ No files.")
        return

    os.makedirs(figure_ocr_output_folder, exist_ok=True)
    progress_bar = st.progress(0)
    status_text = st.empty()

    for idx, image_file in enumerate(image_files):
        input_path = os.path.join(figure_folder, image_file)
        command = ["surya_ocr", input_path]
        result = subprocess.run(command, capture_output=True, text=True)

        stderr = result.stderr.strip()
        if stderr and all(x not in stderr for x in ["Detecting bboxes", "Recognizing Text"]):
            st.warning(f"⚠️ 오류: {image_file} → {stderr}")

        progress = int((idx + 1) / len(image_files) * 100)
        progress_bar.progress(progress)
        status_text.write(f"🖼️ processing: {image_file} ({progress}%)")

    progress_bar.empty()
    status_text.write("✅ All Figure OCR Processing Done!")


# ✅ 5. OCR 결과 이동 (Figure용)
def rename_and_move_ocr_results():

    surya_output_folder = os.path.join(BASE_DIR, "Figure_OCR")
    os.makedirs(surya_output_folder, exist_ok=True)

    if not os.path.exists(SURYA_RESULTS_FOLDER):
        st.error(f"❌ No OCR folder: {SURYA_RESULTS_FOLDER}")
        return

    moved, skipped = 0, 0
    for folder_name in os.listdir(SURYA_RESULTS_FOLDER):
        folder_path = os.path.join(SURYA_RESULTS_FOLDER, folder_name)
        json_file = os.path.join(folder_path, "results.json")
        if not os.path.exists(json_file):
            skipped += 1
            continue
        dst_file = os.path.join(surya_output_folder, f"{folder_name}.json")
        try:
            shutil.move(json_file, dst_file)
            moved += 1
        except Exception as e:
            st.error(f"❌ Error: {e} (folder: {folder_name})")

    st.success(f"✅ Result move complete: {moved}moved, {skipped}missing")


# ✅ 6. Figure OCR 필터링 및 병합

def filter_and_merge_figure_ocr():
    st.write("🔍 Filtering and merging Figure OCR results...")
    json_files = [os.path.join(figure_ocr_output_folder, f)
                  for f in os.listdir(figure_ocr_output_folder) if f.endswith(".json")]

    if not json_files:
        st.error("❌ No JSON file exists as a result of the Figure OCR to filter.")
        return

    os.makedirs(figure_filtered_folder, exist_ok=True)
    progress_bar = st.progress(0)
    status_text = st.empty()

    x_tolerance = 200
    y_tolerance = 30

    for idx, input_file in enumerate(json_files):
        folder_name = os.path.splitext(os.path.basename(input_file))[0]
        output_file = os.path.join(figure_filtered_folder, f"{folder_name}_filtered.txt")

        try:
            with open(input_file, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            st.error(f"❌ {input_file} error to load: {e}")
            continue

        all_boxes = []
        for key, values in data.items():
            for item in values:
                for line in item.get("text_lines", []):
                    if "text" in line and "bbox" in line:
                        all_boxes.append({"text": line["text"], "bbox": line["bbox"]})

        all_boxes.sort(key=lambda x: (x["bbox"][1], x["bbox"][0]))
        merged_boxes = merge_boxes(all_boxes, x_tolerance, y_tolerance)

        with open(output_file, "w", encoding="utf-8") as out_f:
            for box in merged_boxes:
                out_f.write(f"Text: {box['text']}\n")
                out_f.write(f"BBox: {box['bbox']}\n")

        progress_bar.progress((idx + 1) / len(json_files))
        status_text.write(f"📄 processing: {folder_name} ({(idx+1)*100//len(json_files)}%)")

    progress_bar.empty()
    status_text.write("✅ filtering complete!")
    st.success(f"✅ results route : {figure_filtered_folder}")


# ✅ RC 관련 유틸리티 함수 모음 (진행상황 + 결과 파일명 개선 포함)

def merge_boxes(boxes, x_tolerance, y_tolerance):
    if not boxes:
        return []
    merged_boxes = []
    used = [False] * len(boxes)
    def calculate_center(bbox):
        x_min, y_min, x_max, y_max = bbox
        return (x_min + x_max) / 2, (y_min + y_max) / 2
    for i, current_box in enumerate(boxes):
        if used[i]:
            continue
        current_cx, current_cy = calculate_center(current_box["bbox"])
        combined_text = current_box["text"]
        combined_bbox = current_box["bbox"]
        for j, next_box in enumerate(boxes[i + 1:], start=i + 1):
            if used[j]:
                continue
            next_cx, next_cy = calculate_center(next_box["bbox"])
            if abs(current_cx - next_cx) <= x_tolerance and abs(current_cy - next_cy) <= y_tolerance:
                combined_text += f" {next_box['text']}"
                combined_bbox = [
                    min(combined_bbox[0], next_box["bbox"][0]),
                    min(combined_bbox[1], next_box["bbox"][1]),
                    max(combined_bbox[2], next_box["bbox"][2]),
                    max(combined_bbox[3], next_box["bbox"][3]),
                ]
                used[j] = True
        merged_boxes.append({"text": combined_text, "bbox": combined_bbox})
        used[i] = True
    return merged_boxes

def clean_coordinates(coord_str):
    coord_str = coord_str.strip("[] \n")
    try:
        return list(map(float, coord_str.split(",")))
    except ValueError:
        return None

def calculate_area(coords):
    x_min, y_min, x_max, y_max = coords
    return (x_max - x_min) * (y_max - y_min)

def is_valid_number(text):
    try:
        float(text)
        return True
    except ValueError:
        return False

def clean_text_value(text):
    cleaned_text = ''.join([char for char in text if char.isdigit() or char == '.'])
    if cleaned_text.count('.') > 1:
        first_dot_index = cleaned_text.find('.')
        cleaned_text = cleaned_text[:first_dot_index+1] + cleaned_text[first_dot_index+1:].replace('.', '')
    if cleaned_text.endswith('.'):
        cleaned_text = cleaned_text[:-1]
    return cleaned_text if cleaned_text else "1"

def assign_height_and_width(filtered_boxes, image_width, image_height):
    width, height = 1, 1
    bottom_boxes = [box for box in filtered_boxes if box["coords"][3] <= 0.95 * image_height]
    if bottom_boxes:
        bottommost_box = max(bottom_boxes, key=lambda box: box["coords"][3])
        width = clean_text_value(bottommost_box["text"])
        filtered_boxes.remove(bottommost_box)

    leftmost_boxes = [box for box in filtered_boxes]
    if leftmost_boxes:
        leftmost_box = min(leftmost_boxes, key=lambda box: box["coords"][0])
        height = clean_text_value(leftmost_box["text"])
        height_value = float(height)
        if height_value < 0.1 or height_value > 10:
            height = "1"
        filtered_boxes.remove(leftmost_box)

    return width, height

def process_rc_beam(txt_path, output_folder, base_name):
    headers, boxes = [], []
    with open(txt_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    for line in lines:
        line = line.strip()
        if line.startswith("Text:"):
            text = line.split("Text:", 1)[1].strip()
        elif line.startswith("BBox:"):
            coords = clean_coordinates(line.split("BBox:", 1)[1])
            if coords:
                if text.startswith("[") and text.endswith("]"):
                    headers.append({"text": text.strip("[]"), "coords": coords})
                else:
                    boxes.append({"text": text, "coords": coords})

    if not headers:
        return

    grouped_data = {header['text']: [] for header in headers}
    for box in boxes:
        distances = [(abs(header['coords'][0] - (box['coords'][0] + box['coords'][2]) / 2), header['text']) for header in headers]
        if distances:
            _, closest_header = min(distances)
            grouped_data[closest_header].append(box)

    os.makedirs(output_folder, exist_ok=True)
    for header, group in grouped_data.items():
        results = {"HEIGHT": None, "WIDTH": None, "TOP_REBAR": None, "BOT_REBAR": None, "STIRRUPS": None}

        if group:
            leftmost_box = min(group, key=lambda box: box["coords"][0])
            results["HEIGHT"] = clean_text_value(leftmost_box["text"])

        for i, box in enumerate(group):
            if "TOP" in box["text"] and i > 0:
                results["WIDTH"] = clean_text_value(group[i - 1]["text"])
                break

        for box in group:
            text = box["text"]
            if "TOP" in text:
                results["TOP_REBAR"] = text.split("TOP", 1)[-1].strip()
            if "BOT" in text:
                results["BOT_REBAR"] = text.split("BOT", 1)[-1].strip()
            if "STIRRUPS" in text:
                results["STIRRUPS"] = text.split("STIRRUPS", 1)[-1].strip()

        header_name = header.replace(' ', '_')  # ✅ 여기 수정
        output_file = os.path.join(output_folder, f"{base_name}_{header_name}.txt")
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(f"=== {header_name} ===\n")
            for key, value in results.items():
                f.write(f"{key}: {value if value else 'N/A'}\n")

def process_rc_column(txt_path, output_folder, base_name, image_width, image_height):
    boxes = []
    with open(txt_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    for line in lines:
        line = line.strip()
        if line.startswith("Text:"):
            text = line.split("Text:", 1)[1].strip()
        elif line.startswith("BBox:"):
            coords = clean_coordinates(line.split("BBox:", 1)[1])
            if coords:
                boxes.append({"text": text, "coords": coords})

    filtered_boxes = [box for box in boxes if calculate_area(box["coords"]) >= 600]
    width, height = assign_height_and_width(filtered_boxes, image_width, image_height)

    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, f"{base_name}_RC_COLUMN.txt")
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("CODE: RC_COLUMN\n")
        f.write(f"HEIGHT: {height}\n")
        f.write(f"WIDTH: {width}\n")



# ✅ Streamlit과 연동된 RC Beam / Column 추출 메인 함수
def extract_rc_elements():
    st.write("🔍 Extracting RC Beam and RC Column data...")

    txt_files = [file for file in os.listdir(figure_filtered_folder) if file.endswith('_filtered.txt')]
    if not txt_files:
        st.error("❌ No txt files")
        return

    progress_bar = st.progress(0)
    status_text = st.empty()

    for idx, txt_file in enumerate(txt_files):
        txt_path = os.path.join(figure_filtered_folder, txt_file)
        base_name = txt_file.replace('_filtered.txt', '')
        img_path = os.path.join(figure_folder, f"{base_name}.png")

        if not os.path.exists(img_path):
            st.warning(f"⚠️ {base_name}.Png file does not exist. Skip")
            continue

        img = cv2.imread(img_path)
        if img is None:
            st.error(f"❌ {img_path} Unable to load image. Skip.")
            continue

        image_height, image_width, _ = img.shape

        if image_width >= 1500:
            process_rc_beam(txt_path, figure_element_folder, base_name)
        else:
            process_rc_column(txt_path, figure_element_folder, base_name, image_width, image_height)

        progress = int((idx + 1) / len(txt_files) * 100)
        progress_bar.progress(progress)
        status_text.write(f"📄 Extracting RC Element: {base_name} ({progress}%)")

    progress_bar.empty()
    status_text.write("✅ All RC Element data extraction completed!")
    st.success(f"✅ RC Beam/Column data extraction completed! results have been saved to {figure_element_folder}")





def rotate_and_crop_images():
    """
    Figures 폴더의 모든 이미지에 대해:
    1) 전체 이미지를 90° 시계방향으로 회전
    2) 회전된 이미지의 가로 중심부(양옆 25%)만 크롭
    3) 결과 이미지를 Vertical_Temp 폴더에 저장
    """
    if not os.path.isdir(figure_folder):
        st.error(f"Figures 폴더가 없습니다: {figure_folder}")
        return

    os.makedirs(rotated_crop_folder, exist_ok=True)

    st.write("🔄 runing rotation and crop")
    for img_name in tqdm(os.listdir(figure_folder), desc="이미지 처리"):
        if not img_name.lower().endswith(('.png', '.jpg', '.jpeg')):
            continue

        orig_path = os.path.join(figure_folder, img_name)
        img = cv2.imread(orig_path)
        if img is None:
            st.warning(f"⚠️ fail to load {img_name}")
            continue

        rot = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
        _, Wf = rot.shape[:2]

        x1 = int(Wf * 0.25)
        x2 = int(Wf * 0.75)
        crop = rot[:, x1:x2]

        base, _ = os.path.splitext(img_name)
        out_name = f"{base}_rot_crop.png"
        out_path = os.path.join(rotated_crop_folder, out_name)
        cv2.imwrite(out_path, crop)

    st.success(f"✅ Rotated+cropped image saved to {rotated_crop_folder}")





 

def OCR_figure_rotated_cropped_figure():
    st.write("🔍 Running OCR")
    image_files = [f for f in os.listdir(rotated_crop_folder) if f.endswith(('.jpg', '.png'))]
    if not image_files:
        st.error("❌ no files.")
        return

    progress_bar = st.progress(0)
    status_text = st.empty()
    for idx, image_file in enumerate(tqdm(image_files, desc="Applying OCR", unit="file")):
        input_path = os.path.join(rotated_crop_folder, image_file)
        command = ["surya_ocr", input_path]
        subprocess.run(command, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        progress = int((idx + 1) / len(image_files) * 100)
        progress_bar.progress(progress)
        status_text.write(f"🖼️ processing: {image_file} ({progress}%)")
    progress_bar.empty()
    status_text.write("✅ OCR processing completed for all files!")

    for folder_name in os.listdir(SURYA_RESULTS_FOLDER):
        folder_path = os.path.join(SURYA_RESULTS_FOLDER, folder_name)
        if os.path.isdir(folder_path):
            json_file = os.path.join(folder_path, "results.json")
            if os.path.exists(json_file):
                dst_file = os.path.join(Vertical_recog_folder, f"{folder_name}.json")
                try:
                    shutil.move(json_file, dst_file)
                except Exception as e:
                    st.error(f"❌ {folder_name} Error moving: {e}")
            
    st.success(f"✅ Surya OCR result saved in {Vertical_recog_folder}")



def extract_vertical_texts():
    """

    """
    files = [f for f in os.listdir(Vertical_recog_folder)
             if f.endswith('_figure_rot_crop.json')]
    if not files:
        st.error("❌ No vertical OCR result files found.")
        return

    for fn in tqdm(files, desc="Extracting vertical text"):
        path = os.path.join(Vertical_recog_folder, fn)
        try:
            with open(path, encoding='utf-8') as f:
                raw = json.load(f)
        except Exception as e:
            st.warning(f"⚠️ Failed to load {fn}: {e}")
            continue

        key = next(iter(raw))
        blocks = raw[key]

        all_lines = [line for blk in blocks for line in blk.get("text_lines", [])
                     if line.get("text", "").strip()]
        if not all_lines:
            st.warning(f"⚠️ No valid text_lines found in {fn}")
            continue

        top = min(all_lines, key=lambda e: e["bbox"][1])
        top_text = top["text"]
        same = [e for e in all_lines if e["text"] == top_text]
        same_sorted = sorted(same, key=lambda e: e["bbox"][1])

        n = len(same_sorted)
        if n == 1:
            directions = [(top_text, None)]
        elif n == 2:
            directions = [
                (same_sorted[0]["text"], "END-I"),
                (same_sorted[1]["text"], "END-J")
            ]
        else:
            directions = [
                (same_sorted[0]["text"], "END-I"),
                (same_sorted[n // 2]["text"], "MID"),
                (same_sorted[-1]["text"], "END-J")
            ]

        base = fn.replace("_figure_rot_crop.json", "")
        txt_name = f"{base}_Direction.txt"
        txt_path = os.path.join(Vertical_txt, txt_name)

        with open(txt_path, 'w', encoding='utf-8') as out:
            for text, label in directions:
                if label:
                    out.write(f"{text} [{label}]\n")
                else:
                    out.write(f"{text}\n")

    st.success(f"✅ Vertical text extraction complete. Results saved to {Vertical_txt}")



def update_height_from_direction_file():
    """
    Unified function to update HEIGHT values from Direction.txt files.
    Supports both:
        - "value [LABEL]" format (e.g., 1.2 [END-I])
        - "LABEL [value]" format (e.g., END-I [1.2])
        - single value with no label (e.g., 3.6) → applied to RC_COLUMN or general
    """
    direction_folder = os.path.join(BASE_DIR, "Vertical_Text")
    figure_folder = os.path.join(BASE_DIR, "Figure_Element")

    if not os.path.isdir(direction_folder) or not os.path.isdir(figure_folder):
        st.error("❌ Folder missing: Vertical_Text or Figure_Element")
        return

    direction_files = [f for f in os.listdir(direction_folder) if f.endswith('_Direction.txt')]
    if not direction_files:
        st.error("❌ No *_Direction.txt files found.")
        return

    for fn in tqdm(direction_files, desc="Updating HEIGHT"):
        base = fn.replace('_Direction.txt', '')  # e.g., SCD_106
        path = os.path.join(direction_folder, fn)

        mapping = {}
        with open(path, encoding='utf-8') as f:
            for line in f:
                line = line.strip()

                # Case 1: "value [LABEL]" → e.g., 1.2 [END-I]
                m1 = re.match(r'^(.+?)\s*\[(END-I|MID|END-J)\]$', line)
                if m1:
                    val = m1.group(1).strip()
                    label = m1.group(2).strip()
                    mapping[label] = val
                    continue

                # Case 2: "LABEL [value]" → e.g., END-I [1.2]
                m2 = re.match(r'^(END-I|MID|END-J)\s*\[(.+?)\]$', line)
                if m2:
                    label = m2.group(1).strip()
                    val = m2.group(2).strip()
                    mapping[label] = val
                    continue

                # Case 3: value only → for RC_COLUMN
                if re.match(r'^[0-9.]+$', line):
                    mapping[None] = line

        for label, val in mapping.items():
            if label is None:
                # Try RC_COLUMN file first, fallback to general
                rc_path = os.path.join(figure_folder, f"{base}_figure_RC_COLUMN.txt")
                general_path = os.path.join(figure_folder, f"{base}_figure.txt")
                target_file = rc_path if os.path.exists(rc_path) else general_path
            else:
                target_file = os.path.join(figure_folder, f"{base}_figure_{label}.txt")

            if not os.path.exists(target_file):
                st.warning(f"⚠️ File not found: {os.path.basename(target_file)}")
                continue

            lines = []
            with open(target_file, encoding='utf-8') as rf:
                for L in rf:
                    if L.startswith("HEIGHT:"):
                        lines.append(f"HEIGHT: {val}\n")
                    else:
                        lines.append(L)

            with open(target_file, 'w', encoding='utf-8') as wf:
                wf.writelines(lines)

            st.write(f"✅ Updated {os.path.basename(target_file)} → HEIGHT: {val}")

    st.success("✅ All HEIGHT values updated.")





















###Table형식에서 stirrup 추출 

# ✅ RC Column Table 관련 전처리 전체 파이프라인
def extract_rc_column_jsons_from_ocr_results():

    figure_element_folder = os.path.join(BASE_DIR, "Figure_Element")
    rc_column_json_folder = os.path.join(BASE_DIR, "Table_OCR")
    os.makedirs(rc_column_json_folder, exist_ok=True)

    pattern = os.path.join(figure_element_folder, "SCD_*_figure_RC_COLUMN.txt")
    pages = []
    for path in glob.glob(pattern):
        fname = os.path.basename(path)
        match = re.match(r"SCD_(\d+)_figure_RC_COLUMN\.txt", fname)
        if match:
            pages.append(match.group(1))

    json_candidates = [f for f in os.listdir(SURYA_RESULTS_FOLDER) if f.endswith(".json")]
    selected = []

    for page in pages:
        related_files = [f for f in json_candidates if f.startswith(f"SCD_{page}_plain")]
        for file in related_files:
            full_path = os.path.join(SURYA_RESULTS_FOLDER, file)
            try:
                with open(full_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                found = False
                for block in data.get(f"SCD_{page}_plain_{file.split('_')[-1].split('.')[0]}", []):
                    for line in block.get("text_lines", []):
                        if "@" in line.get("text", ""):
                            found = True
                            break
                    if found:
                        break

                if found:
                    dst = os.path.join(rc_column_json_folder, file)
                    shutil.copy(full_path, dst)
                    selected.append(file)

            except Exception as e:
                st.warning(f"⚠️ error: {file} → {e}")

    return selected


def extract_stirrup_rebar_patterns_from_json_surya():
    json_folder = os.path.join(BASE_DIR, "Table_OCR")
    output_folder = os.path.join(BASE_DIR, "RC_COLUMN_REBAR_TXT")
    os.makedirs(output_folder, exist_ok=True)

    pattern1 = re.compile(r'(\d+)\s*-\s*(D|HD|SUHD|UHD|UD)\s*(\d+)\s*@\s*(\d+)', re.IGNORECASE)
    pattern2 = re.compile(r'(D|HD|SUHD|UHD|UD)\s*(\d+)\s*@\s*(\d+)', re.IGNORECASE)

    extracted_count = 0

    for json_path in glob.glob(os.path.join(json_folder, "*.json")):
        found_full = set()
        found_simple = set()

        with open(json_path, 'r', encoding='utf-8') as f:
            try:
                data = json.load(f)
            except Exception as e:
                st.error(f"❌ JSON parsing failed: {json_path} → {e}")
                continue

        def extract_from_text(text):
            if not isinstance(text, str):
                return
            for m in pattern1.findall(text):
                found_full.add(f"{m[0]}-{m[1].upper()}{m[2]}@{m[3]}")
            for m in pattern2.findall(text):
                found_simple.add(f"{m[0].upper()}{m[1]}@{m[2]}")

        if isinstance(data, list):
            for item in data:
                extract_from_text(item.get("text", ""))
        elif isinstance(data, dict):
            if 'results' in data:
                for item in data['results']:
                    extract_from_text(item.get("text", ""))
            elif 'boxes' in data:
                for item in data['boxes']:
                    extract_from_text(item.get("text", ""))
            elif len(data) == 1:
                val = next(iter(data.values()))
                if isinstance(val, list):
                    for entry in val:
                        for line in entry.get("text_lines", []):
                            extract_from_text(line.get("text", ""))
            else:
                for k, v in data.items():
                    if isinstance(v, str):
                        extract_from_text(v)

        # ⚠️ 중복 제거 우선순위 처리: full 패턴 우선
        final_items = found_full.copy()
        for s in found_simple:
            if not any(s in f for f in found_full):
                final_items.add(s)

        if final_items:
            base = os.path.splitext(os.path.basename(json_path))[0]
            out_path = os.path.join(output_folder, f"{base}_stirrup.txt")
            with open(out_path, 'w', encoding='utf-8') as out_f:
                out_f.write("\n".join(sorted(final_items)))
            extracted_count += 1

    st.success(f"✅ extracted and saved Rebar patterns from a total of {extracted_count}files!→ {output_folder})")


def apply_stirrup_to_rc_columns():
    """
    Match stirrup text from RC_COLUMN_REBAR_TXT/*.txt with corresponding
    RC column text files in Figure_Element and append STIRRUPS: <value> at the end.
    """
    stirrup_folder = os.path.join(BASE_DIR, "RC_COLUMN_REBAR_TXT")
    figure_folder = os.path.join(BASE_DIR, "Figure_Element")

    stirrup_files = [f for f in os.listdir(stirrup_folder) if f.endswith('_stirrup.txt')]
    if not stirrup_files:
        st.error("❌ No stirrup files found in RC_COLUMN_REBAR_TXT.")
        return

    updated_count = 0

    for stirrup_file in stirrup_files:
        base = stirrup_file.replace('_stirrup.txt', '')  # e.g., SCD_3_plain_0
        base_id = "_".join(base.split("_")[:2])  # e.g., SCD_3
        figure_path = os.path.join(figure_folder, f"{base_id}_figure_RC_COLUMN.txt")

        stirrup_path = os.path.join(stirrup_folder, stirrup_file)
        if not os.path.exists(figure_path):
            st.warning(f"⚠️ Figure file not found: {figure_path}")
            continue

        with open(stirrup_path, encoding='utf-8') as sf:
            stirrup_lines = [line.strip() for line in sf if line.strip()]
        if not stirrup_lines:
            st.warning(f"⚠️ No stirrup data in: {stirrup_file}")
            continue

        stirrup_value = stirrup_lines[0]

        with open(figure_path, encoding='utf-8') as ff:
            lines = [line.rstrip('\n') for line in ff]

        if any(line.startswith("STIRRUPS:") for line in lines):
            st.info(f"ℹ️ STIRRUPS already exists in {figure_path}, skipping.")
            continue

        lines.append(f"STIRRUPS: {stirrup_value}")

        with open(figure_path, 'w', encoding='utf-8') as out_f:
            for line in lines:
                out_f.write(line + '\n')

        updated_count += 1
        st.write(f"✅ STIRRUPS added to {figure_path} → {stirrup_value}")

    st.success(f"🎯 STIRRUPS information has been added to a total of {updated_count} files")








def extract_member_data():
    """✅ JSON 결과에서 Member·Section·Rebar·Material 정보를 추출 후 TXT로 저장"""
    json_folder = MemberCode_extraction_ALL
    txt_folder  = os.path.join(BASE_DIR, "MEMBER_EXTRACTION")
    os.makedirs(txt_folder, exist_ok=True)

    st.write("DEBUG: Member JSON 폴더 경로:", json_folder)
    st.write("DEBUG: TXT 저장 폴더 경로:", txt_folder)

    member_data = {}
    if not os.path.isdir(json_folder):
        st.warning(f"⚠️ 폴더가 존재하지 않습니다: {json_folder}")
        return member_data

    files = [f for f in os.listdir(json_folder) if f.endswith("_results.json")]
    for fn in files:
        match = re.search(r"SCD_(\d+)_.*_results\.json", fn)
        if not match:
            continue
        page = int(match.group(1))
        data = json.load(open(os.path.join(json_folder, fn), encoding="utf-8"))

        def first(key):
            lst = data.get(key, [])
            return lst[0]["text"] if lst else ""

        member_value        = first("Member Code")
        section_value       = first("Section Property")
        rebar_value         = first("Rebar Pattern")
        material_value      = first("Material")

        member_data[page] = {
            "Member":           member_value,
            "Section Property": section_value,
            "Rebar Pattern":    rebar_value,
            "Material":         material_value
        }

        # ↓ 여기서 페이지별 TXT 파일로 저장
        txt_path = os.path.join(txt_folder, f"SCD_{page}.txt")
        with open(txt_path, 'w', encoding='utf-8') as out:
            out.write(f"Member: {member_value}\n")
            out.write(f"Section Property: {section_value}\n")
            out.write(f"Rebar Pattern: {rebar_value}\n")
            out.write(f"Material: {material_value}\n")
        st.write(f"✅ {os.path.basename(txt_path)} 저장 완료")

    st.success(f"✅ Member 데이터 추출 및 TXT 저장 완료 → {txt_folder}")
    return member_data






# 헬퍼: 섹션 문자열에서 숫자+영문 연속 조합 단어(fullmatch) 확인 (X/x 제외)
def is_simple_token(s: str) -> bool:
    return bool(re.fullmatch(r"(?:\d+(?![Xx])[A-Za-z]+|[A-Za-z]+(?![Xx])\d+)", s))

# 추출 함수 정의
def extract_element_data(output_dir=None):
    """
    INPUT_DIR 폴더에서 Member, Section Property, Rebar Pattern, fck, fy 값을 추출하여
    output_dir에 txt 파일로 저장합니다.

    Args:
        output_dir (str, optional): 결과 txt 파일을 저장할 디렉토리 경로.
            지정하지 않으면 INPUT_DIR/Element_extraction에 저장합니다.

    Returns:
        List[str]: 생성된 파일 경로들의 리스트
    """
    INPUT_DIR = os.path.join(BASE_DIR, "MEMBER_EXTRACTION")

    if output_dir is None:
        output_dir = os.path.join(BASE_DIR, "Element_extraction")
    os.makedirs(output_dir, exist_ok=True)

    output_paths = []
    for filepath in glob.glob(os.path.join(INPUT_DIR, "*.txt")):
        member = ""
        section = ""
        rebar = ""
        material = ""

        # 파일 라인 단위 읽기
        with open(filepath, 'r', encoding='utf-8') as f:
            for line in f:
                ln = line.strip()
                if ln.startswith("Member:"):
                    member = ln[len("Member:"):].strip()
                elif ln.startswith("Section Property:"):
                    section = ln[len("Section Property:"):].strip()
                elif ln.startswith("Rebar Pattern:"):
                    rebar = ln[len("Rebar Pattern:"):].strip()
                elif ln.startswith("Material:"):
                    material = ln[len("Material:"):].strip()

        # 초기화: Member가 잘못 읽힌 경우
        if member.startswith("Section Property:"):
            member = ""

        # 조건 1: member 비어있고 section만 있을 경우
        if not member and section:
            member = section.lstrip(': ').strip()
            section = ""

        # 섹션 정리
        section_clean = section.lstrip(': ').strip()

        # 회피 1: X/x 포함된 경우 단순 무시 (section만)
        if section_clean and re.search(r"[Xx]", section_clean):
            section_clean = ""
            section = ""

        # 조건 2: override 로직 - section이 simple token이거나 괄호 포함
        if member and section_clean and (is_simple_token(section_clean) or '(' in section_clean):
            # 'Number' 접두어만 제거
            if member.lower().startswith('number'):
                member = section_clean
            else:
                # prefix는 원멤버의 접두어 유지
                prefix = member.split(':', 1)[0] + ':'
                member = f"{prefix} {section_clean}".strip()
            section = ""
        # 조건 3: 일반 append 로직
        elif member and section_clean:
            member = f"{member}, {section_clean}"
            section = ""

        # material에서 fck, fy 값 추출
        material_clean = material.lstrip(': ').strip()
        fck_match = re.search(r"\bfck\s*=\s*(\d+)", material_clean)
        fy_match = re.search(r"\bfy\s*=\s*(\d+)", material_clean)
        fck = fck_match.group(1) if fck_match else ""
        fy = fy_match.group(1) if fy_match else ""

        # 결과 작성
        output_content = (
            f"Member: {member}\n"
            f"Section Property: {section}\n"
            f"Rebar Pattern: {rebar}\n"
            f"fck: {fck}\n"
            f"fy: {fy}\n"
        )

        filename = os.path.basename(filepath)
        output_path = os.path.join(output_dir, filename)
        with open(output_path, 'w', encoding='utf-8') as out_f:
            out_f.write(output_content)
        output_paths.append(output_path)

    return output_paths




# CSV 결합 함수
def combine_data_to_csv():
    """
    BASE_DIR 폴더 내 Element_extraction과 Figure_Element 폴더를 결합해
    final_results.csv로 저장합니다.

    - 일반 figure 파일: figure 파일에서 TOP/BOT_REBAR, STIRRUPS, WIDTH, HEIGHT 읽고
      앞의 콜론(:)과 모든 공백을 제거
    - RC_COLUMN figure 파일: Element_extraction의 Rebar Pattern을 Top_Rebar로 사용
      Direction을 빈 문자열로 설정하고, Figure_Element 파일의 'Stirrups :' 값을 읽어 Stirrup 열에 추가

    Returns:
        str: 생성된 CSV 파일 경로
    """
    EXTRACT_DIR = os.path.join(BASE_DIR, "Element_extraction")
    FIGURE_DIR  = os.path.join(BASE_DIR, "Figure_Element")
    OUTPUT_CSV  = os.path.join(BASE_DIR, "final_results.csv")

    fieldnames = [
        "page", "Member", "Direction",
        "Top_Rebar", "Bot_Rebar", "Stirrup",
        "Width", "Height",  # ✅ 추가됨
        "fck", "fy"
    ]
    rows = []

    # Figure 파일 순회
    pattern = os.path.join(FIGURE_DIR, "SCD_*_figure_*.txt")
    for fig_path in glob.glob(pattern):
        name = os.path.basename(fig_path)
        m = re.match(r"SCD_(\d+)_figure_(.+)\.txt", name)
        if not m:
            continue
        page, direction = m.group(1), m.group(2)

        # 1) figure 파일에서 필요한 정보 추출
        top_rebar = bot_rebar = stirrup = width = height = ""
        with open(fig_path, 'r', encoding='utf-8') as f:
            for line in f:
                if line.startswith("TOP_REBAR:"):
                    top_rebar = line.split(":", 1)[1].strip()
                elif line.startswith("BOT_REBAR:"):
                    bot_rebar = line.split(":", 1)[1].strip()
                elif line.lower().startswith("stirrups"):
                    stirrup = line.split(":", 1)[1].strip()
                elif line.startswith("WIDTH:"):
                    width = line.split(":", 1)[1].strip()
                elif line.startswith("HEIGHT:"):
                    height = line.split(":", 1)[1].strip()

        # 2) Rebar 문자열 공백 및 콜론 제거
        top_rebar = top_rebar.replace(":", "").replace(" ", "")
        bot_rebar = bot_rebar.replace(":", "").replace(" ", "")

        # 3) Element_extraction에서 Member, fck, fy, rebar_pattern 읽기
        extract_file = os.path.join(EXTRACT_DIR, f"SCD_{page}.txt")
        member = fck = fy = rebar_pattern = ""
        if os.path.exists(extract_file):
            with open(extract_file, 'r', encoding='utf-8') as ef:
                for line in ef:
                    if line.startswith("Member:"):
                        member = line.split("Member:", 1)[1].strip()
                    elif line.startswith("fck:"):
                        fck = line.split("fck:", 1)[1].strip()
                    elif line.startswith("fy:"):
                        fy = line.split("fy:", 1)[1].strip()
                    elif line.startswith("Rebar Pattern:"):
                        rebar_pattern = line.split(":", 1)[1].strip()

        # 4) RC_COLUMN 처리
        if "COLUMN" in direction.upper():
            top_rebar = rebar_pattern.replace(":", "").replace(" ", "")
            direction = ""  # COLUMN이면 방향 없음

        # 5) CSV 행 추가
        rows.append({
            "page": page,
            "Member": member,
            "Direction": direction,
            "Top_Rebar": top_rebar,
            "Bot_Rebar": bot_rebar,
            "Stirrup": stirrup,
            "Width": width,
            "Height": height,
            "fck": fck,
            "fy": fy,
 
        })

    # 6) CSV 저장 (Excel 호환 위해 UTF-8 BOM)
    with open(OUTPUT_CSV, 'w', encoding='utf-8-sig', newline='') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    return OUTPUT_CSV




def standardize_member_split():
    """
    final_results.csv 파일에서 Member 열의 쉼표(,) 구분 항목을 행으로 분리하고,
    나머지 정보는 동일하게 복제하여 final_results_standardized.csv로 저장
    """
    input_path  = os.path.join(BASE_DIR, "final_results.csv")
    output_path = os.path.join(BASE_DIR, "final_results_standardized.csv")

    if not os.path.exists(input_path):
        print(f"❌ 파일 없음: {input_path}")
        return

    df = pd.read_csv(input_path, encoding='utf-8-sig')
    expanded_rows = []

    for _, row in df.iterrows():
        members = [m.strip() for m in str(row["Member"]).split(',') if m.strip()]
        for m in members:
            new_row = row.copy()
            new_row["Member"] = m
            expanded_rows.append(new_row)

    new_df = pd.DataFrame(expanded_rows)
    new_df.to_csv(output_path, index=False, encoding='utf-8-sig')
    print(f"✅ Member 쉼표 분리 완료! 저장 위치: {output_path}")





def normalize_member_column(max_floor: int = None):
    input_path = os.path.join(BASE_DIR, "final_results_standardized.csv")
    output_path = os.path.join(BASE_DIR, "SCD_Complete.csv")

    if not os.path.exists(input_path):
        print(f"❌ 파일 없음: {input_path}")
        return

    df = pd.read_csv(input_path, encoding='utf-8-sig')
    processed_rows = []

    for _, row in df.iterrows():
        raw = str(row["Member"])

        raw = re.sub(r"\((?!.*Base\s*:).*?\)", "", raw)

        base_match = re.search(r"\(?\s*Base\s*:\s*([BF]?\-?\d+|PIT)\s*\)?", raw, re.IGNORECASE)
        base_floor = None
        if base_match:
            base_raw = base_match.group(1).upper().replace(" ", "")
            if "PIT" in base_raw:
                base_floor = max_floor if max_floor is not None else "PIT"
            elif "B" in base_raw:
                base_floor = -int(re.sub(r"[^\d]", "", base_raw))
            elif "F" in base_raw:
                base_floor = int(re.sub(r"[^\d]", "", base_raw))
            else:
                base_floor = int(base_raw)
            raw = re.sub(r"\(?\s*Base\s*:\s*([BF]?\-?\d+|PIT)\s*\)?", "", raw, flags=re.IGNORECASE)

        raw = re.sub(r":\s*\d+", "", raw)
        raw = raw.strip()

        m = re.match(r"(-?\d+)\s*~\s*(PIT|B?\-?\d+)([A-Za-z0-9]+)", raw, re.IGNORECASE)
        if m:
            start = int(m.group(1))
            end_raw = m.group(2).upper()
            code = m.group(3)

            if end_raw == "PIT":
                if max_floor is not None:
                    end = max_floor
                    for floor in range(start, end + 1):
                        if floor == 0:
                            continue
                        new_row = row.copy()
                        new_row["Member"] = f"{floor} {code}"
                        processed_rows.append(new_row)
                else:
                    for fl in [start, "PIT"]:
                        new_row = row.copy()
                        new_row["Member"] = f"{fl} {code}"
                        processed_rows.append(new_row)
            else:
                try:
                    end = int(re.sub(r"[^\d\-]", "", end_raw))
                    for floor in range(start, end + 1):
                        if floor == 0:
                            continue
                        new_row = row.copy()
                        new_row["Member"] = f"{floor} {code}"
                        processed_rows.append(new_row)
                except ValueError:
                    pass
            continue

        if raw.upper().startswith("PIT") and max_floor is not None:
            code_match = re.match(r"PIT([A-Za-z0-9]+)", raw)
            if code_match:
                code = code_match.group(1)
                new_row = row.copy()
                new_row["Member"] = f"{max_floor} {code}"
                processed_rows.append(new_row)
                continue

        if base_floor is not None:
            m2 = re.match(r"([A-Za-z0-9]+)", raw)
            if m2:
                code = m2.group(1)
                new_row = row.copy()
                new_row["Member"] = f"{base_floor} {code}"
                processed_rows.append(new_row)
                continue

        new_row = row.copy()
        new_row["Member"] = raw
        processed_rows.append(new_row)

    new_df = pd.DataFrame(processed_rows)
    new_df["Member"] = new_df["Member"].str.strip()

    # ✅ 중복 제거
    if "Direction" in new_df.columns:
        key_cols = ["Member", "Direction"]
    else:
        key_cols = ["Member"]
    other_cols = [col for col in new_df.columns if col not in key_cols]
    new_df = new_df.drop_duplicates(subset=key_cols + other_cols, keep="first")

    new_df.to_csv(output_path, index=False, encoding='utf-8-sig')
    print(f"✅ Member 정규화 및 저장 완료! → {output_path}")
################################################################################################################
#SCD COMPLETE###################################################################################################
################################################################################################################



#SDD

def process_uploaded_images_SD(uploaded_pdfs):
    from PIL import Image, ImageEnhance  # 필요한 PIL 모듈 import
    import numpy as np
    import cv2
    
    image_paths = []
    progress_bar = st.progress(0)
    status = st.empty()
    global_page_index = 1

    def enhance_image_for_ocr(pil_image):
        """
        OCR 성능 향상을 위한 이미지 전처리 (크기 변경 절대 금지)
        - 대비 향상: 텍스트와 배경 구분도 증가
        - 선명도 향상: 흐릿한 텍스트 개선
        - 노이즈 제거: 스캔 아티팩트 감소
        - 원본 크기 100% 보존
        """
        try:
            # 🚫 해상도 최적화 코드 완전 삭제 (크기 변경 금지)
            # 원본 크기 확인 (참고용)
            original_size = pil_image.size
            
            # 🚫 PIL ImageEnhance 사용 금지 (크기 변경 위험)
            # 바로 OpenCV로 변환
            img_array = np.array(pil_image.convert('RGB'))
            img_cv = cv2.cvtColor(img_array, cv2.COLOR_RGB2BGR)

            # 그레이스케일 변환
            gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)

            # 대비 향상 (히스토그램 균등화)
            enhanced = cv2.equalizeHist(gray)

            # 선명도 향상 (언샤프 마스킹)
            blurred = cv2.GaussianBlur(enhanced, (3, 3), 0)
            sharpened = cv2.addWeighted(enhanced, 1.5, blurred, -0.5, 0)

            # 적응형 임계값 (지역적 조명 변화 대응)
            adaptive_thresh = cv2.adaptiveThreshold(
                sharpened, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                cv2.THRESH_BINARY, 11, 2
            )

            # 모폴로지 연산 (최소한으로)
            kernel = np.ones((1, 1), np.uint8)
            cleaned = cv2.morphologyEx(adaptive_thresh, cv2.MORPH_CLOSE, kernel)

            # PIL로 변환
            final_img = Image.fromarray(cleaned, mode='L')
            
            # 크기 검증
            if final_img.size != original_size:
                st.error(f"❌ 크기 변경됨: {original_size} → {final_img.size}")
                return pil_image
            else:
                st.success(f"✅ 크기 보존됨: {final_img.size}")
                
            return final_img

        except Exception as e:
            st.warning(f"전처리 실패: {e}, 원본 이미지 사용")
            return pil_image

    for idx, pdf in enumerate(uploaded_pdfs):
        pages = convert_from_bytes(pdf.getvalue(), dpi=400)

        for pg, img in enumerate(pages):
            base_name = f"merged_SD_page_{global_page_index}"
            
            # 🔥 1단계: 원본 이미지 저장 (raw_data_SD)
            original_path = os.path.join(raw_data_folder_SD, f"{base_name}.png")
            
            if not os.path.exists(original_path):
                # PDF에서 변환된 원본 이미지 그대로 저장
                img.save(original_path, "PNG", quality=95, optimize=True)
                st.info(f"💾 원본 저장: {base_name}.png")
            else:
                st.info(f"🖼️ 원본 {base_name}.png 이미 존재 → 생략")
            
            image_paths.append(original_path)
            
            # 🔥 2단계: 전처리된 이미지 저장 (raw_data_SD_processing_folder)
            processed_path = os.path.join(raw_data_folder_SD_processing_folder, f"{base_name}_processed.png")
            
            if not os.path.exists(processed_path):
                # OCR 향상 전처리 적용 후 저장
                enhanced_img = enhance_image_for_ocr(img)
                enhanced_img.save(processed_path, "PNG", quality=95, optimize=True)
                st.info(f"🔧 전처리 저장: {base_name}_processed.png")
            else:
                st.info(f"⚙️ 전처리 {base_name}_processed.png 이미지 존재 → 생략")

            global_page_index += 1

        progress_bar.progress((idx + 1) / len(uploaded_pdfs))
        status.text(f"📄 {pdf.name} 변환 완료 (총 {len(pages)}페이지) - 원본 + 전처리 버전 저장됨")

    st.success(f"✅ 총 {len(image_paths)}개 페이지 완료")
    st.info(f"📁 원본 이미지: {raw_data_folder_SD}")
    st.info(f"📁 전처리 이미지 (OCR용): {raw_data_folder_SD_processing_folder}")
    
    st.markdown("**🔍 처리 결과 요약:**")
    st.markdown("- **원본**: PDF → PNG 변환만")
    st.markdown("- **전처리**: 대비/선명도 향상 + 노이즈 제거 + OCR 최적화 (크기 동일)")
    
    st.code("\n".join(image_paths[:5]), language="text")
    return image_paths


def apply_surya_ocr_SD():
    """Surya OCR을 적용하여 Plain Text에서 텍스트 추출 (결과는 고정 경로에 저장됨)"""
    # Surya OCR 도구의 결과는 고정 경로에 저장됨 (하위에 각 파일명 폴더가 있고 그 안에 results.json 존재)
    st.write("🔍 Surya OCR 실행 중...")
    image_files = [f for f in os.listdir(raw_data_folder_SD_processing_folder) if f.endswith(('.jpg', '.png'))]
    if not image_files:
        st.error("❌ OCR 적용할 파일이 없습니다.")
        return

    progress_bar = st.progress(0)
    status_text = st.empty()
    for idx, image_file in enumerate(tqdm(image_files, desc="Applying Surya OCR", unit="file")):
        input_path = os.path.join(raw_data_folder_SD_processing_folder, image_file)
        # 결과 저장 경로는 도구 내부 고정 경로이므로 --output_dir 옵션 생략
        command = ["surya_ocr", input_path]
        result = subprocess.run(command, capture_output=True, text=True)
        st.write("STDOUT:", result.stdout)
        st.write("STDERR:", result.stderr)
        progress = int((idx + 1) / len(image_files) * 100)
        progress_bar.progress(progress)
        status_text.write(f"🖼️ 처리 중: {image_file} ({progress}%)")
    progress_bar.empty()
    status_text.write("✅ 모든 파일에 대한 OCR 처리가 완료되었습니다!")

    # OCR 도구 결과는 SURYA_RESULTS_FOLDER 내에 각 이미지에 대한 폴더로 저장됨
    # 이를 surya_output_SD_folder (예: D:\TEST_Streamlit\surya_output_SD_folder)로 이동시키면서,
    # 각 폴더 내 results.json 파일은 해당 폴더명으로 이름을 변경하여 저장합니다.
    os.makedirs(surya_output_SD_folder, exist_ok=True)
    for folder_name in os.listdir(SURYA_RESULTS_FOLDER):
        folder_path = os.path.join(SURYA_RESULTS_FOLDER, folder_name)
        if os.path.isdir(folder_path):
            json_file = os.path.join(folder_path, "results.json")
            if os.path.exists(json_file):
                dst_file = os.path.join(surya_output_SD_folder, f"{folder_name}.json")
                try:
                    shutil.move(json_file, dst_file)
                    st.write(f"📁 {folder_name}의 JSON 파일이 {dst_file}로 이동되었습니다.")
                except Exception as e:
                    st.error(f"❌ {folder_name} 이동 중 오류 발생: {e}")
            else:
                st.warning(f"⚠️ {folder_name} 폴더에 results.json 파일이 존재하지 않습니다.")
    st.success(f"✅ Surya OCR 결과가 {surya_output_SD_folder}에 저장되었습니다!")





def draw_and_save_bounding_boxes(canvas_key: str = "box_saver"):
    """
    • raw_data_SD 폴더에서 이미지를 골라
    • MAX 크기에 맞춰 비율 축소 후 캔버스에 띄우고
    • 사각형을 그린 뒤 boxed_coords 폴더에 JSON으로 좌표 저장
    """

    input_folder  = raw_data_folder_SD_processing_folder
    output_folder = os.path.join(BASE_DIR, "boxed_coords")
    os.makedirs(output_folder, exist_ok=True)

    # 1) 이미지 리스트
    imgs = [
        f for f in os.listdir(input_folder)
        if f.lower().endswith((".png", ".jpg", ".jpeg"))
    ]
    if not imgs:
        st.warning(f"이미지 없음: {input_folder}")
        return None, []

    # 2) 메인 페이지에서 이미지 선택
    selected = st.selectbox("이미지 선택", imgs, key=canvas_key + "_sel")
    img_path = os.path.join(input_folder, selected)
    image    = Image.open(img_path)

    # 3) 축소 비율 계산
    MAX_W, MAX_H = 800, 600
    ow, oh       = image.width, image.height
    scale        = min(1.0, MAX_W/ow, MAX_H/oh)
    disp_w       = int(ow * scale)
    disp_h       = int(oh * scale)

    # 4) 축소된 이미지
    img_small = image.resize((disp_w, disp_h), Image.LANCZOS)
    st.image(img_small, caption=selected, width=disp_w)

    # 5) 캔버스 생성
    canvas_result = st_canvas(
        background_image=img_small,
        drawing_mode   ="rect",
        stroke_width   =2,
        stroke_color   ="#FF0000",
        fill_color     ="rgba(255,0,0,0.3)",
        update_streamlit=True,
        height=disp_h,
        width=disp_w,
        key=canvas_key
    )

    # 6) 박스 좌표 수집 & 원본 크기로 역스케일 후 저장
    boxes = []
    if canvas_result and canvas_result.json_data:
        for o in canvas_result.json_data.get("objects", []):
            if o.get("type") == "rect":
                boxes.append({
                    "left":   round(o["left"]   / scale, 2),
                    "top":    round(o["top"]    / scale, 2),
                    "width":  round(o["width"]  / scale, 2),
                    "height": round(o["height"] / scale, 2),
                })
        out_file = os.path.join(
            output_folder,
            f"{os.path.splitext(selected)[0]}_boxes.json"
        )
        with open(out_file, "w", encoding="utf-8") as f:
            json.dump(boxes, f, ensure_ascii=False, indent=2)
        st.success(f"✅ 좌표 저장됨: {out_file}")

    return img_path, boxes



def match_ocr_to_bboxes_all(y_margin: int = 30):
    """
    1) 함수 내에서 입출력 폴더를 정의하고,
    2) 각 페이지별 bounding-box JSON과 OCR JSON을 읽어,
    3) Y축 중심이 y_margin 이내인 텍스트를 매칭한 뒤,
    4) 결과를 페이지별 JSON으로 output_folder에 저장합니다.

    파일명 규칙:
      - bounding box:  merged_SD_page_{i}_boxes.json
      - OCR 결과:      merged_SD_page_{i}.json
      - 결과 저장:     merged_SD_page_{i}_matches.json
    """
    # 1) 입·출력 폴더 정의

    bbox_folder      = os.path.join(BASE_DIR, "boxed_coords")
    ocr_folder       = os.path.join(BASE_DIR, "surya_output_SD_folder")
    output_folder    = os.path.join(BASE_DIR, "matched_results")
    os.makedirs(output_folder, exist_ok=True)

    # 2) 각 bounding box 파일 순회
    for fname in os.listdir(bbox_folder):
        if not fname.endswith("_boxes.json"):
            continue
        page_key = fname.replace("_boxes.json", "")
        bbox_path = os.path.join(bbox_folder, fname)
        ocr_path  = os.path.join(ocr_folder, f"{page_key}.json")
        if not os.path.exists(ocr_path):
            print(f"▶ OCR 파일 누락: {ocr_path}, 스킵")
            continue

        # 3) JSON 로드
        with open(bbox_path, "r", encoding="utf-8") as f:
            bboxes = json.load(f)
        with open(ocr_path, "r", encoding="utf-8") as f:
            ocr = json.load(f)
        # OCR 구조에 맞게 text_lines 추출
        # 예: {"merged_SD_page_1":[{"text_lines":[{...},...]},...]}
        ocr_page = ocr.get(page_key)
        if not ocr_page:
            print(f"▶ OCR 구조 오류: {page_key}")
            continue
        text_boxes = ocr_page[0]["text_lines"]

        # 4) 매칭 로직
        matches = []
        for idx, bb in enumerate(bboxes):
            bb_cy = bb["top"] + bb["height"] / 2
            matched_items = []
            for tb in text_boxes:
                x0, y0, x1, y1 = tb["bbox"]
                tb_cy = y0 + (y1 - y0) / 2
                if abs(tb_cy - bb_cy) <= y_margin:
                    matched_items.append({
                        "text": tb["text"].strip(),
                        "bbox": [x0, y0, x1, y1]
                    })
            matches.append({
                "bbox_index": idx,
                "matched_items": matched_items
        })

        # 5) 결과 저장
        out_path = os.path.join(output_folder, f"{page_key}_matches.json")
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(matches, f, ensure_ascii=False, indent=2)
        print(f"✅ {page_key} 매칭 결과 저장: {out_path}")
#_________________________________________________________________________________________




def combined_line_and_cell_debug():
    """선 검출 + 셀 생성 + 매칭 결과를 한번에 확인하는 통합 디버깅 인터페이스 (개선된 셀 검출)"""

    # ==== 매칭 공용 유틸: 경계 중복 방지 + 단일 후보 선택 + 거리 허용 ====
    def point_inside_rect_strict(px, py, rect, distance_tolerance=0):
        xs = [p[0] for p in rect]; ys = [p[1] for p in rect]
        x0, x1 = min(xs), max(xs); y0, y1 = min(ys), max(ys)
        
        # 기존 내부 체크
        if (x0 <= px < x1) and (y0 <= py < y1):
            return True
        
        # distance_tolerance가 0보다 크면 거리 허용
        if distance_tolerance > 0:
            dx = max(0, max(x0 - px, px - x1))
            dy = max(0, max(y0 - py, py - y1))
            distance = (dx**2 + dy**2)**0.5
            return distance <= distance_tolerance
        
        return False

    def select_best_cell_for_point(px, py, candidates):
        def key_fn(r):
            xs = [p[0] for p in r]; ys = [p[1] for p in r]
            cx = (min(xs) + max(xs)) / 2.0
            cy = (min(ys) + max(ys)) / 2.0
            area = max(1.0, (max(xs) - min(xs)) * (max(ys) - min(ys)))
            return (abs(py - cy), area)
        return min(candidates, key=key_fn) if candidates else None

    # ==== [추가] 매칭 실패 원인 분석 함수 ====
    def analyze_matching_failure(text_box, valid_cells, intersection_tolerance=10):
        """매칭 실패 원인을 분석하여 상세한 정보 반환"""
        cx, cy = text_box["center"]
        x0, y0, x1, y1 = text_box["bbox"]
        
        failure_reasons = []
        nearest_cells = []
        
        # 모든 셀과의 거리 계산
        for i, cell in enumerate(valid_cells):
            xs = [p[0] for p in cell]; ys = [p[1] for p in cell]
            cell_x0, cell_x1 = min(xs), max(xs)
            cell_y0, cell_y1 = min(ys), max(ys)
            cell_cx, cell_cy = (cell_x0 + cell_x1) / 2.0, (cell_y0 + cell_y1) / 2.0
            
            # 중심점 간 거리
            center_distance = ((cx - cell_cx) ** 2 + (cy - cell_cy) ** 2) ** 0.5
            
            # 경계까지의 거리 (음수면 내부, 양수면 외부)
            dx_left = cx - cell_x0
            dx_right = cell_x1 - cx
            dy_top = cy - cell_y0
            dy_bottom = cell_y1 - cy
            
            # 가장 가까운 경계까지의 거리
            if dx_left < 0:  # 왼쪽 밖
                edge_dist_x = -dx_left
            elif dx_right < 0:  # 오른쪽 밖
                edge_dist_x = -dx_right
            else:  # x축으로는 내부
                edge_dist_x = 0
                
            if dy_top < 0:  # 위쪽 밖
                edge_dist_y = -dy_top
            elif dy_bottom < 0:  # 아래쪽 밖
                edge_dist_y = -dy_bottom
            else:  # y축으로는 내부
                edge_dist_y = 0
                
            edge_distance = (edge_dist_x ** 2 + edge_dist_y ** 2) ** 0.5
            
            # 겹침 정도 계산
            overlap_x = max(0, min(x1, cell_x1) - max(x0, cell_x0))
            overlap_y = max(0, min(y1, cell_y1) - max(y0, cell_y0))
            overlap_area = overlap_x * overlap_y
            text_area = (x1 - x0) * (y1 - y0)
            overlap_ratio = overlap_area / text_area if text_area > 0 else 0
            
            nearest_cells.append({
                'cell_idx': i,
                'center_distance': center_distance,
                'edge_distance': edge_distance,
                'overlap_ratio': overlap_ratio,
                'cell_bounds': (cell_x0, cell_y0, cell_x1, cell_y1),
                'is_inside': edge_distance == 0,
                'position_info': {
                    'dx_left': dx_left, 'dx_right': dx_right,
                    'dy_top': dy_top, 'dy_bottom': dy_bottom
                }
            })
        
        # 거리순 정렬
        nearest_cells.sort(key=lambda x: x['center_distance'])
        
        # 실패 원인 분석
        if not valid_cells:
            failure_reasons.append("검출된 셀이 전혀 없음")
        else:
            closest_cell = nearest_cells[0]
            
            if not closest_cell['is_inside']:
                if closest_cell['edge_distance'] <= intersection_tolerance:
                    failure_reasons.append(f"경계 근처에 있음 (거리: {closest_cell['edge_distance']:.1f}px, 허용: {intersection_tolerance}px)")
                else:
                    failure_reasons.append(f"가장 가까운 셀과 거리가 멀음 (거리: {closest_cell['edge_distance']:.1f}px)")
                    
                # 위치 상세 정보
                pos = closest_cell['position_info']
                if pos['dx_left'] < 0:
                    failure_reasons.append(f"  → 셀 왼쪽에 위치 ({-pos['dx_left']:.1f}px 떨어짐)")
                elif pos['dx_right'] < 0:
                    failure_reasons.append(f"  → 셀 오른쪽에 위치 ({-pos['dx_right']:.1f}px 떨어짐)")
                    
                if pos['dy_top'] < 0:
                    failure_reasons.append(f"  → 셀 위쪽에 위치 ({-pos['dy_top']:.1f}px 떨어짐)")
                elif pos['dy_bottom'] < 0:
                    failure_reasons.append(f"  → 셀 아래쪽에 위치 ({-pos['dy_bottom']:.1f}px 떨어짐)")
            
            # 겹침 정보
            if closest_cell['overlap_ratio'] > 0:
                failure_reasons.append(f"가장 가까운 셀과 {closest_cell['overlap_ratio']*100:.1f}% 겹침")
        
        return {
            'reasons': failure_reasons,
            'nearest_cells': nearest_cells[:3],  # 가장 가까운 3개만
            'text_center': (cx, cy),
            'text_bbox': (x0, y0, x1, y1)
        }

    # 이미지 선택
    image_dir = raw_data_folder_SD_processing_folder
    image_files = [f for f in os.listdir(image_dir) if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
    if not image_files:
        st.error("이미지 파일이 없습니다.")
        return
    selected_image = st.selectbox("테스트할 이미지 선택", image_files, key="combined_debug_image_select")
    image_path = os.path.join(image_dir, selected_image)

    # 파라미터 조정 슬라이더들
    col1, col2 = st.columns(2)
    with col1:
        st.write("**전처리 파라미터**")
        adaptive_block_size = st.slider("Adaptive Threshold Block Size", 3, 21, 11, step=2)
        adaptive_c = st.slider("Adaptive Threshold C", 1, 10, 2)
        morph_kernel_size = st.slider("Morphology Kernel Size", 1, 5, 2)
        st.write("**Canny Edge 파라미터**")
        canny_low = st.slider("Canny Low Threshold", 10, 100, 30)
        canny_high = st.slider("Canny High Threshold", 50, 200, 100)
    with col2:
        st.write("**HoughLinesP 파라미터**")
        hough_threshold = st.slider("Hough Threshold", 10, 100, 30)
        min_line_length = st.slider("Min Line Length", 10, 100, 20)
        max_line_gap = st.slider("Max Line Gap", 5, 30, 10)
        st.write("**선 분류 & 셀 생성 파라미터**")
        line_angle_tolerance = st.slider("Line Angle Tolerance", 0.05, 0.3, 0.1, step=0.01)
        intersection_tolerance = st.slider("Intersection Tolerance", 5, 30, 10)
        min_area_threshold = st.slider("Min Cell Area", 500, 5000, 1000)
        # 🔥 NEW: 최소 높이 조건 추가
        min_cell_height = st.slider("Min Cell Height", 10, 100, 30)
        max_cell_width = st.slider("Max Cell Width", 1000, 5000, 2000)
        max_cell_height = st.slider("Max Cell Height", 500, 2000, 1000)
        # 🔥 NEW: 셀 검증 조건 완화를 위한 파라미터
        min_valid_edges = st.slider("Min Valid Edges (out of 4)", 2, 4, 4)
        st.write("**텍스트-셀 매칭 파라미터**")
        distance_tolerance = st.slider("Distance Tolerance (px)", 0, 200, 50)

    # 테스트 실행 버튼
    if st.button("파라미터 적용 및 전체 프로세스 확인"):
        image = cv2.imread(image_path)
        if image is None:
            st.error("이미지를 불러올 수 없습니다.")
            return

        # 코어 함수 호출로 선·교차점·셀 한 번에 생성
        horizontal_lines, vertical_lines, intersections, valid_cells = compute_lines_intersections_cells(
            image,
            params={
                'adaptive_block_size': adaptive_block_size,
                'adaptive_c': adaptive_c,
                'morph_kernel_size': morph_kernel_size,
                'canny_low': canny_low,
                'canny_high': canny_high,
                'hough_threshold': hough_threshold,
                'min_line_length': min_line_length,
                'max_line_gap': max_line_gap,
                'line_angle_tolerance': line_angle_tolerance,
                'intersection_tolerance': intersection_tolerance,
                'min_area_threshold': min_area_threshold,
                'min_cell_height': min_cell_height,  # 🔥 NEW
                'max_cell_width': max_cell_width,
                'max_cell_height': max_cell_height,
                'min_valid_edges': min_valid_edges,  # 🔥 NEW
            }
        )

        # OCR 텍스트 로드
        page_id = selected_image.replace('.png', '').replace('.jpg', '').replace('.jpeg', '')
        matched_json = os.path.join(BASE_DIR, "matched_results", f"{page_id}_matches.json")
        matched_texts = []
        if os.path.exists(matched_json):
            with open(matched_json, 'r', encoding='utf-8') as f:
                match_data = json.load(f)
            for item in match_data:
                for matched in item.get("matched_items", []):
                    text = matched["text"].strip()
                    x0, y0, x1, y1 = matched["bbox"]
                    cx, cy = (x0 + x1) / 2.0, (y0 + y1) / 2.0
                    matched_texts.append({"text": text, "center": (cx, cy), "bbox": [x0, y0, x1, y1]})

        # 텍스트-셀 매칭 → 항상 1개만 유지 + 실패 원인 분석 + 거리 허용치 적용
        text_cell_matches = []
        failed_matches = []  # 실패한 매칭들
        
        for text_box in matched_texts:
            cx, cy = text_box["center"]
            # distance_tolerance 파라미터 적용
            candidates = [(i, cell) for i, cell in enumerate(valid_cells) 
                         if point_inside_rect_strict(cx, cy, cell, distance_tolerance)]
            
            if candidates:
                best_cell = select_best_cell_for_point(cx, cy, [c[1] for c in candidates])
                best_idx = next((i for i, c in enumerate(valid_cells) if c is best_cell or c == best_cell), candidates[0][0])
                text_cell_matches.append({"text": text_box["text"], "center": text_box["center"], "matched_cells": [(best_idx, best_cell)]})
            else:
                # 실패 원인 분석 (distance_tolerance 값도 전달)
                failure_analysis = analyze_matching_failure(text_box, valid_cells, distance_tolerance)
                text_cell_matches.append({"text": text_box["text"], "center": text_box["center"], "matched_cells": []})
                failed_matches.append({
                    "text": text_box["text"],
                    "center": text_box["center"], 
                    "bbox": text_box["bbox"],
                    "analysis": failure_analysis
                })

        # 시각화
        result_image = image.copy()
        for (x1, y1), (x2, y2) in horizontal_lines:
            cv2.line(result_image, (x1, y1), (x2, y2), (0, 0, 255), 2)
        for (x1, y1), (x2, y2) in vertical_lines:
            cv2.line(result_image, (x1, y1), (x2, y2), (255, 0, 0), 2)
        for x, y in intersections:
            cv2.circle(result_image, (int(x), int(y)), 8, (0, 255, 0), -1)
        for i, cell in enumerate(valid_cells):
            xs = [int(p[0]) for p in cell]; ys = [int(p[1]) for p in cell]
            pts = np.array([[xs[0], ys[0]], [xs[1], ys[1]], [xs[2], ys[2]], [xs[3], ys[3]]], np.int32)
            cv2.polylines(result_image, [pts], True, (255, 255, 0), 2)
        
        # 성공한 매칭 표시
        for match in text_cell_matches:
            cx, cy = match["center"]
            if match["matched_cells"]:
                cv2.circle(result_image, (int(cx), int(cy)), 6, (255, 0, 255), -1)  # 성공: 자홍색
            else:
                cv2.circle(result_image, (int(cx), int(cy)), 6, (0, 0, 255), -1)  # 실패: 빨간색

        # ==== [추가] 실패한 텍스트의 가장 가까운 셀들 표시 ====
        for failed in failed_matches:
            cx, cy = failed["center"]
            # 실패한 텍스트 주변에 큰 빨간 원
            cv2.circle(result_image, (int(cx), int(cy)), 12, (0, 0, 255), 3)
            
            # 가장 가까운 셀들과 연결선 그리기
            for nearest in failed["analysis"]["nearest_cells"][:2]:  # 가장 가까운 2개만
                cell_idx = nearest["cell_idx"]
                cell = valid_cells[cell_idx]
                xs = [p[0] for p in cell]; ys = [p[1] for p in cell]
                cell_cx = int((min(xs) + max(xs)) / 2.0)
                cell_cy = int((min(ys) + max(ys)) / 2.0)
                
                # 점선으로 연결
                cv2.line(result_image, (int(cx), int(cy)), (cell_cx, cell_cy), (128, 128, 128), 1)
                # 거리 표시
                dist = nearest["center_distance"]
                mid_x, mid_y = int((cx + cell_cx) / 2), int((cy + cell_cy) / 2)
                cv2.putText(result_image, f"{dist:.0f}", (mid_x, mid_y), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)

        # ==== [추가] 매칭된 셀 강조 및 크롭 미리보기 ====
        matched_crops = []
        for match in text_cell_matches:
            if not match["matched_cells"]:
                continue
            cell_idx, cell = match["matched_cells"][0]
            # 원본 결과 이미지에 매칭 셀 테두리 강조
            x0, y0, x1, y1 = _cell_to_xyxy(cell)
            cv2.rectangle(result_image, (x0, y0), (x1, y1), (0, 255, 255), 3)
            # 실제 크롭 추출
            crop = _crop_cell(image, cell, pad=2)
            if crop is not None:
                matched_crops.append((match["text"], cell_idx, crop))

        # 결과 표시
        st.write("**검출 및 매칭 결과**")
        st.image(cv2.cvtColor(result_image, cv2.COLOR_BGR2RGB))

        # 통계
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("수평선", len(horizontal_lines))
            st.metric("수직선", len(vertical_lines))
        with col2:
            st.metric("교차점", len(intersections))
            st.metric("유효 셀", len(valid_cells))
        with col3:
            st.metric("텍스트 박스", len(matched_texts))
            st.metric("매칭 성공", sum(1 for m in text_cell_matches if m["matched_cells"]))

        # 매칭 상세
        st.write("**텍스트-셀 매칭 상세**")
        for match in text_cell_matches:
            if match["matched_cells"]:
                st.write(f"✅ '{match['text']}' → {len(match['matched_cells'])}개 셀 매칭")
            else:
                st.write(f"❌ '{match['text']}' → 매칭 실패")

        # ==== [추가] 매칭 실패 원인 분석 결과 표시 ====
        if failed_matches:
            st.write("**🔍 매칭 실패 원인 분석**")
            st.write("---")  # 구분선
            
            for i, failed in enumerate(failed_matches):
                st.write(f"**❌ {i+1}. '{failed['text']}' 실패 원인**")
                analysis = failed["analysis"]
                
                # expander 없이 직접 내용 표시
                st.write("**실패 원인:**")
                for reason in analysis["reasons"]:
                    st.write(f"• {reason}")
                
                st.write(f"**텍스트 위치:** 중심 ({analysis['text_center'][0]:.1f}, {analysis['text_center'][1]:.1f})")
                x0, y0, x1, y1 = analysis["text_bbox"]
                st.write(f"**텍스트 경계:** ({x0:.1f}, {y0:.1f}) ~ ({x1:.1f}, {y1:.1f})")
                
                if analysis["nearest_cells"]:
                    st.write("**가장 가까운 셀들:**")
                    for j, nearest in enumerate(analysis["nearest_cells"]):
                        cell_bounds = nearest["cell_bounds"]
                        st.write(f"  {j+1}. 셀 #{nearest['cell_idx']}: "
                            f"중심거리 {nearest['center_distance']:.1f}px, "
                            f"경계거리 {nearest['edge_distance']:.1f}px, "
                            f"겹침 {nearest['overlap_ratio']*100:.1f}%")
                        st.write(f"     셀 경계: ({cell_bounds[0]:.1f}, {cell_bounds[1]:.1f}) ~ ({cell_bounds[2]:.1f}, {cell_bounds[3]:.1f})")
                
                st.write("---")  # 각 항목 구분
        else:
            st.success("🎉 모든 텍스트가 셀과 성공적으로 매칭되었습니다!")

        # ==== [추가] 크롭 미리보기 + ZIP 다운로드 ====
        if matched_crops:
            st.write("**매칭된 셀 크롭 미리보기**")
            cols = st.columns(4)
            for i, (txt, idx, crop_img) in enumerate(matched_crops):
                with cols[i % 4]:
                    h, w = crop_img.shape[:2]  # 크롭 이미지 크기 추출
                    st.image(cv2.cvtColor(crop_img, cv2.COLOR_BGR2RGB), 
                            caption=f"[{idx}] {txt}\n{w}×{h}px", use_column_width=True)

            import io, zipfile
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                for k, (txt, idx, crop_img) in enumerate(matched_crops):
                    safe_txt = "".join(ch if ch.isalnum() or ch in "-_@" else "_" for ch in txt)[:40]
                    fname = f"{page_id}_cell{idx:04d}_{safe_txt or 'text'}.png"
                    _, png_bytes = cv2.imencode(".png", crop_img)
                    zf.writestr(fname, png_bytes.tobytes())
            st.download_button("크롭 ZIP 다운로드", buf.getvalue(), file_name=f"{page_id}_matched_cell_crops.zip", mime="application/zip")
        else:
            st.info("매칭된 셀 크롭이 없습니다")

    # ==== 파라미터 저장 + 크롭 실행 버튼 ====
    st.markdown("---")  # 구분선
    st.markdown("**파라미터 설정 완료 후 크롭 실행:**")
    
    # 크롭 옵션
    col1, col2 = st.columns(2)
    with col1:
        crop_pad = st.slider("크롭 패딩 (px)", 0, 10, 2, key="final_crop_pad")
    with col2:
        min_save_width = st.slider("최소 저장 너비 (px)", 50, 1000, 500, key="final_min_save_width")

    st.info("최소 저장 너비: 이 너비보다 작은 크롭 이미지는 저장하지 않습니다")
    
    # 2개 버튼: 현재 이미지만 vs 모든 이미지
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("해당 이미지에만 적용", key="apply_current_image"):
            # 파라미터 저장
            _save_params_to_session(
                adaptive_block_size, adaptive_c, morph_kernel_size,
                canny_low, canny_high, hough_threshold, min_line_length, max_line_gap,
                line_angle_tolerance, intersection_tolerance, min_area_threshold,
                min_cell_height, max_cell_width, max_cell_height, min_valid_edges, distance_tolerance
            )
            
            # 현재 이미지 크롭 실행
            image = cv2.imread(image_path)
            if image is not None:
                # OCR 텍스트 로드
                page_id = selected_image.replace('.png', '').replace('.jpg', '').replace('.jpeg', '')
                matched_json = os.path.join(BASE_DIR, "matched_results", f"{page_id}_matches.json")
                matched_texts = []
                if os.path.exists(matched_json):
                    with open(matched_json, 'r', encoding='utf-8') as f:
                        match_data = json.load(f)
                    for item in match_data:
                        for matched in item.get("matched_items", []):
                            text = matched["text"].strip()
                            x0, y0, x1, y1 = matched["bbox"]
                            cx, cy = (x0 + x1) / 2.0, (y0 + y1) / 2.0
                            matched_texts.append({"text": text, "center": (cx, cy), "bbox": [x0, y0, x1, y1]})
                
                # 검출 실행
                horizontal_lines, vertical_lines, intersections, valid_cells = compute_lines_intersections_cells(image, st.session_state.line_detection_params)
                
                # 매칭 실행
                text_cell_matches = _execute_text_cell_matching(matched_texts, valid_cells, st.session_state.line_detection_params['distance_tolerance'])
                
                # 크롭 실행 (min_save_width 파라미터 추가)
                _execute_extended_crop_single(
                    image, page_id, text_cell_matches, valid_cells,
                    horizontal_lines, vertical_lines, intersections,
                    crop_pad, min_save_width  # min_crop_height → min_save_width로 변경
                )
            else:
                st.error("이미지를 불러올 수 없습니다.")

    with col2:
        if st.button("모든 이미지에 적용", key="apply_all_images"):
            # 파라미터 저장
            _save_params_to_session(
                adaptive_block_size, adaptive_c, morph_kernel_size,
                canny_low, canny_high, hough_threshold, min_line_length, max_line_gap,
                line_angle_tolerance, intersection_tolerance, min_area_threshold,
                min_cell_height, max_cell_width, max_cell_height, min_valid_edges, distance_tolerance
            )
            
            # 모든 이미지 크롭 실행 (min_save_width 파라미터 추가)
            _execute_extended_crop_all(
                crop_pad, min_save_width,  # min_crop_height → min_save_width로 변경
                adaptive_block_size, adaptive_c, morph_kernel_size,
                canny_low, canny_high, hough_threshold, min_line_length, max_line_gap,
                line_angle_tolerance, intersection_tolerance, min_area_threshold,
                min_cell_height, max_cell_width, max_cell_height, min_valid_edges, distance_tolerance
            )

# 🔥 주요 변경: compute_lines_intersections_cells 함수 개선
def compute_lines_intersections_cells(image, params):
    """
    Step 3.75와 완전히 동일한 로직으로 선, 교차점, 셀을 생성
    모든 파라미터를 정확히 적용 + 개선된 셀 검출 로직
    """
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    binary = cv2.adaptiveThreshold(
        gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV, 
        params['adaptive_block_size'], 
        params['adaptive_c']
    )
    
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (params['morph_kernel_size'], params['morph_kernel_size']))
    morph = cv2.morphologyEx(binary, cv2.MORPH_CLOSE, kernel)
    edges = cv2.Canny(morph, params['canny_low'], params['canny_high'])

    lines = cv2.HoughLinesP(
        edges, 1, np.pi / 180, params['hough_threshold'],
        minLineLength=params['min_line_length'], 
        maxLineGap=params['max_line_gap']
    )
    
    horizontal_lines, vertical_lines = [], []
    lat = params['line_angle_tolerance']
    
    if lines is not None:
        for line in lines:
            x1, y1, x2, y2 = line[0]
            dx, dy = x2 - x1, y2 - y1
            if abs(dx) > abs(dy) and abs(dx) > 0:
                if abs(dy / dx) < lat:
                    horizontal_lines.append(((x1, y1), (x2, y2)))
            elif abs(dy) > 0:
                if abs(dx / dy) < lat:
                    vertical_lines.append(((x1, y1), (x2, y2)))

    # 교차점 계산
    intersections = []
    tol = params['intersection_tolerance']
    for (hx1, hy1), (hx2, hy2) in horizontal_lines:
        for (vx1, vy1), (vx2, vy2) in vertical_lines:
            vx = (vx1 + vx2) / 2.0
            hy = (hy1 + hy2) / 2.0
            h_x_min, h_x_max = min(hx1, hx2), max(hx1, hx2)
            v_y_min, v_y_max = min(vy1, vy2), max(vy1, vy2)
            if (h_x_min - tol <= vx <= h_x_max + tol) and (v_y_min - tol <= hy <= v_y_max + tol):
                intersections.append((int(vx), int(hy)))

    # 개선된 셀 생성 로직
    valid_cells = []
    row_groups = {}
    for x, y in intersections:
        row_groups.setdefault(y, []).append((x, y))

    for y in sorted(row_groups.keys()):
        row = sorted(row_groups[y])
        for i in range(len(row) - 1):
            for j in range(i + 1, len(row)):
                p1 = row[i]
                p3 = row[j]
                x1, x2 = p1[0], p3[0]
                for next_y in sorted([k for k in row_groups.keys() if k > y]):
                    next_row = row_groups[next_y]
                    p2_candidates = [p for p in next_row if abs(p[0] - x1) <= tol]
                    p4_candidates = [p for p in next_row if abs(p[0] - x2) <= tol]
                    if p2_candidates and p4_candidates:
                        p2 = p2_candidates[0]
                        p4 = p4_candidates[0]
                        width = abs(x2 - x1)
                        height = abs(p2[1] - p1[1])
                        area = width * height
                        
                        # 개선된 조건: 기존 조건 + 최소 높이 조건
                        if (area >= params.get('min_area_threshold', 1000) and
                            width <= params.get('max_cell_width', 2000) and
                            height <= params.get('max_cell_height', 1000) and
                            height >= params.get('min_cell_height', 30)):  # NEW: 최소 높이 조건
                            
                            # 개선된 선 검증: 4변 중 최소 N개만 실제 선 위에 있으면 OK
                            min_valid_edges = params.get('min_valid_edges', 3)
                            valid_edge_count = 0
                            
                            # 각 변이 실제 선 위에 있는지 확인
                            edges_to_check = [
                                (p1, p3, horizontal_lines),  # 상단 가로선
                                (p2, p4, horizontal_lines),  # 하단 가로선  
                                (p1, p2, vertical_lines),    # 좌측 세로선
                                (p3, p4, vertical_lines)     # 우측 세로선
                            ]
                            
                            for edge_p1, edge_p2, line_list in edges_to_check:
                                if line_on_detected_lines(edge_p1, edge_p2, line_list, tolerance=10):
                                    valid_edge_count += 1
                            
                            # 조건 완화: min_valid_edges 개 이상의 변만 실제 선 위에 있으면 셀로 인정
                            if valid_edge_count >= min_valid_edges:
                                valid_cells.append([p1, p2, p4, p3])
                        break
    
    return horizontal_lines, vertical_lines, intersections, valid_cells


def line_on_detected_lines(p1, p2, detected_lines, tolerance=20):
    """선분이 검출된 선들 위에 있는지 확인하는 함수"""
    x1, y1 = p1
    x2, y2 = p2
    for (lx1, ly1), (lx2, ly2) in detected_lines:
        line_min_x, line_max_x = sorted([lx1, lx2])
        line_min_y, line_max_y = sorted([ly1, ly2])
        seg_min_x, seg_max_x = sorted([x1, x2])
        seg_min_y, seg_max_y = sorted([y1, y2])
        is_horizontal = abs(ly1 - ly2) < tolerance and abs(y1 - y2) < tolerance
        is_vertical   = abs(lx1 - lx2) < tolerance and abs(x1 - x2) < tolerance
        if is_horizontal and (line_min_x - tolerance) <= seg_min_x <= (line_max_x + tolerance) \
            and (line_min_x - tolerance) <= seg_max_x <= (line_max_x + tolerance) \
            and abs(ly1 - y1) <= tolerance:
            return True
        if is_vertical and (line_min_y - tolerance) <= seg_min_y <= (line_max_y + tolerance) \
            and (line_min_y - tolerance) <= seg_max_y <= (line_max_y + tolerance) \
            and abs(lx1 - x1) <= tolerance:
            return True
    return False


def _cell_to_xyxy(cell):
    """셀 좌표를 xyxy 형식으로 변환"""
    xs = [int(p[0]) for p in cell]; ys = [int(p[1]) for p in cell]
    return min(xs), min(ys), max(xs), max(ys)


def _crop_cell(image, cell, pad=2):
    """셀 영역을 크롭하는 함수"""
    h, w = image.shape[:2]
    x0, y0, x1, y1 = _cell_to_xyxy(cell)
    x0 = max(0, x0 - pad); y0 = max(0, y0 - pad)
    x1 = min(w, x1 + pad); y1 = min(h, y1 + pad)
    if x1 <= x0 or y1 <= y0:
        return None
    return image[y0:y1, x0:x1].copy()


# ===== NEW: 확장 크롭 관련 함수들 =====
def _find_next_matched_row_top(current_cell, text_cell_matches, tolerance=150):
    """매칭된 셀들 중에서 다음 행의 상단 Y좌표를 찾는 함수"""
    curr_y0, curr_y1 = _cell_to_xyxy(current_cell)[1], _cell_to_xyxy(current_cell)[3]
    curr_row_center = (curr_y0 + curr_y1) / 2.0
    
    # 매칭된 셀들만 고려
    next_row_tops = []
    for match in text_cell_matches:
        if not match["matched_cells"]:
            continue
        
        _, matched_cell = match["matched_cells"][0]
        cell_y0, cell_y1 = _cell_to_xyxy(matched_cell)[1], _cell_to_xyxy(matched_cell)[3]
        cell_row_center = (cell_y0 + cell_y1) / 2.0
        
        # 현재 행보다 아래에 있는 매칭된 셀들만
        if cell_row_center > curr_row_center + tolerance:
            next_row_tops.append(cell_y0)
    
    return min(next_row_tops) if next_row_tops else None




def _find_next_matched_row_top(current_cell, text_cell_matches, tolerance=150):
    """매칭된 셀들 중에서 다음 행의 상단 Y좌표를 찾는 함수"""
    curr_y0, curr_y1 = _cell_to_xyxy(current_cell)[1], _cell_to_xyxy(current_cell)[3]
    curr_row_center = (curr_y0 + curr_y1) / 2.0
    
    # 매칭된 셀들만 고려
    next_row_tops = []
    for match in text_cell_matches:
        if not match["matched_cells"]:
            continue
        
        _, matched_cell = match["matched_cells"][0]
        cell_y0, cell_y1 = _cell_to_xyxy(matched_cell)[1], _cell_to_xyxy(matched_cell)[3]
        cell_row_center = (cell_y0 + cell_y1) / 2.0
        
        # 현재 행보다 아래에 있는 매칭된 셀들만
        if cell_row_center > curr_row_center + tolerance:
            next_row_tops.append(cell_y0)
    
    return min(next_row_tops) if next_row_tops else None


def _find_next_matched_row_top(current_cell, text_cell_matches, tolerance=150):
    """매칭된 셀들 중에서 다음 행의 상단 Y좌표를 찾는 함수"""
    curr_y0, curr_y1 = _cell_to_xyxy(current_cell)[1], _cell_to_xyxy(current_cell)[3]
    curr_row_center = (curr_y0 + curr_y1) / 2.0
    
    # 매칭된 셀들만 고려
    next_row_tops = []
    for match in text_cell_matches:
        if not match["matched_cells"]:
            continue
        
        _, matched_cell = match["matched_cells"][0]
        cell_y0, cell_y1 = _cell_to_xyxy(matched_cell)[1], _cell_to_xyxy(matched_cell)[3]
        cell_row_center = (cell_y0 + cell_y1) / 2.0
        
        # 현재 행보다 아래에 있는 매칭된 셀들만
        if cell_row_center > curr_row_center + tolerance:
            next_row_tops.append(cell_y0)
    
    return min(next_row_tops) if next_row_tops else None


def _clear_page_data(page_id, basic_save_dir, extended_save_dir, intersections_dir, metadata_path):
   """특정 페이지의 모든 데이터 삭제"""
   deleted_count = 0
   
   # 1. 기본 크롭 삭제
   if os.path.exists(basic_save_dir):
       basic_files = [f for f in os.listdir(basic_save_dir) if f.startswith(f"{page_id}_basic_")]
       for f in basic_files:
           os.remove(os.path.join(basic_save_dir, f))
           deleted_count += 1
   
   # 2. 확장 크롭 삭제  
   if os.path.exists(extended_save_dir):
       extended_files = [f for f in os.listdir(extended_save_dir) if f.startswith(f"{page_id}_extended_")]
       for f in extended_files:
           os.remove(os.path.join(extended_save_dir, f))
           deleted_count += 1
   
   # 3. 교차점 파일 삭제
   intersection_file = os.path.join(intersections_dir, f"{page_id}_intersections.json")
   if os.path.exists(intersection_file):
       os.remove(intersection_file)
       deleted_count += 1
   
   # 4. 메타데이터에서 해당 페이지 제거
   if os.path.exists(metadata_path):
       with open(metadata_path, 'r', encoding='utf-8') as f:
           metadata = json.load(f)
       original_count = len(metadata)
       metadata = [m for m in metadata if m["page_id"] != page_id]
       with open(metadata_path, 'w', encoding='utf-8') as f:
           json.dump(metadata, f, indent=2, ensure_ascii=False)
       deleted_count += (original_count - len(metadata))
   
   return deleted_count


def _execute_extended_crop_single(image, page_id, text_cell_matches, valid_cells, 
                               h_lines, v_lines, intersections, crop_pad, min_save_width):
   """현재 이미지에 기본 크롭 + 확장 크롭 적용 (페이지 단위 덮어쓰기)"""
   
   # 저장 디렉토리 - 2개 생성
   basic_save_dir = os.path.join(BASE_DIR, "SD_cell_extraction","SD_basic_crops")
   extended_save_dir = os.path.join(BASE_DIR, "SD_cell_extraction","SD_extended_crops")
   intersections_dir = os.path.join(BASE_DIR, "SD_cell_extraction","SD_intersections")
   metadata_path = os.path.join(BASE_DIR, "SD_cell_extraction","metadata","SD_crop_metadata.json")
   os.makedirs(basic_save_dir, exist_ok=True)
   os.makedirs(extended_save_dir, exist_ok=True)
   os.makedirs(intersections_dir, exist_ok=True)
   
   # 해당 페이지의 기존 데이터 모두 삭제
   deleted_count = _clear_page_data(page_id, basic_save_dir, extended_save_dir, intersections_dir, metadata_path)
   if deleted_count > 0:
       st.write(f"🗑️ {page_id} 관련 기존 데이터 {deleted_count}개 삭제")
   
   basic_cropped_count = 0
   extended_cropped_count = 0
   skipped_count = 0
   h, w = image.shape[:2]
   
   # 메타데이터 수집용 리스트
   metadata_list = []
   
   # 교차점 정보 저장 (후행 단계용)
   intersections_py = [[int(x), int(y)] for x, y in intersections]
   intersection_path = os.path.join(intersections_dir, f"{page_id}_intersections.json")
   with open(intersection_path, 'w', encoding='utf-8') as f:
       json.dump(intersections_py, f, indent=2)
   
   st.write(f"**교차점 정보 저장**: {len(intersections_py)}개 → `{intersection_path}`")
   
   for match in text_cell_matches:
       if not match["matched_cells"]:
           continue
           
       cell_idx, current_cell = match["matched_cells"][0]
       text = match["text"]
       safe_text = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in text)[:20]
       
       # 1️⃣ 기본 셀 크롭 (미리보기와 동일)
       basic_crop = _crop_cell(image, current_cell, pad=crop_pad)
       if basic_crop is not None:
           basic_width = basic_crop.shape[1]
           if basic_width >= min_save_width:
               basic_filename = f"{page_id}_basic_{basic_cropped_count:03d}_{safe_text}.png"
               basic_save_path = os.path.join(basic_save_dir, basic_filename)
               cv2.imwrite(basic_save_path, basic_crop)
               
               # 기본 크롭 메타데이터 추가
               basic_x0, basic_y0, basic_x1, basic_y1 = _cell_to_xyxy(current_cell)
               basic_x0 = max(0, basic_x0 - crop_pad)
               basic_y0 = max(0, basic_y0 - crop_pad) 
               basic_x1 = min(w, basic_x1 + crop_pad)
               basic_y1 = min(h, basic_y1 + crop_pad)
               
               metadata_list.append({
                   "page_id": page_id,
                   "filename": basic_filename,
                   "bbox": [int(basic_x0), int(basic_y0), int(basic_x1), int(basic_y1)],
                   "text": text,
                   "crop_type": "basic",
                   "cell_idx": cell_idx
               })
               
               basic_cropped_count += 1
               st.write(f"✅ 기본 '{text}' → {basic_crop.shape[1]}×{basic_crop.shape[0]}px → `{basic_filename}`")
           else:
               st.write(f"⚠️ 기본 '{text}' → 너비 {basic_width}px < {min_save_width}px (스킵)")
               skipped_count += 1
       
       # 2️⃣ 확장 크롭 (매칭된 셀들 간 거리 기반)
       curr_x0, curr_y0, curr_x1, curr_y1 = _cell_to_xyxy(current_cell)
       
       # 매칭된 셀들 중에서 다음 행의 상단 찾기
       next_row_top = _find_next_matched_row_top(current_cell, text_cell_matches, tolerance=150)
       
       if next_row_top:
           # 다음 매칭된 행의 상단까지 확장
           extended_y1 = next_row_top
       else:
           # 매칭된 다음 행이 없으면 이미지 끝까지
           extended_y1 = h
       
       # 패딩 적용해서 최종 크롭 영역 계산
       crop_x0 = max(0, curr_x0 - crop_pad)
       crop_y0 = max(0, curr_y0 - crop_pad)
       crop_x1 = min(w, curr_x1 + crop_pad)
       crop_y1 = min(h, extended_y1 + crop_pad)
       
       # 크롭 너비 계산
       crop_width = crop_x1 - crop_x0
       crop_height = crop_y1 - crop_y0
       
       # 확장 크롭 저장
       if crop_width >= min_save_width and crop_x1 > crop_x0 and crop_y1 > crop_y0:
           extended_cropped = image[crop_y0:crop_y1, crop_x0:crop_x1]
           extended_filename = f"{page_id}_extended_{extended_cropped_count:03d}_{safe_text}.png"
           extended_save_path = os.path.join(extended_save_dir, extended_filename)
           cv2.imwrite(extended_save_path, extended_cropped)
           
           # 확장 크롭 메타데이터 추가
           metadata_list.append({
               "page_id": page_id,
               "filename": extended_filename,
               "bbox": [int(crop_x0), int(crop_y0), int(crop_x1), int(crop_y1)],
               "text": text,
               "crop_type": "extended",
               "cell_idx": cell_idx
           })
           
           extended_cropped_count += 1
           st.write(f"✅ 확장 '{text}' → {crop_width}×{crop_height}px → `{extended_filename}`")
       else:
           if crop_width < min_save_width:
               st.write(f"⚠️ 확장 '{text}' → 너비 {crop_width}px < {min_save_width}px (스킵)")
           else:
               st.warning(f"❌ 확장 '{text}' → 잘못된 크롭 영역")
           skipped_count += 1
   
   # 메타데이터 저장
   if metadata_list:
       # 기존 메타데이터가 있으면 읽어서 추가
       existing_metadata = []
       if os.path.exists(metadata_path):
           with open(metadata_path, 'r', encoding='utf-8') as f:
               existing_metadata = json.load(f)
       
       existing_metadata.extend(metadata_list)
       
       with open(metadata_path, 'w', encoding='utf-8') as f:
           json.dump(existing_metadata, f, indent=2, ensure_ascii=False)
       
       st.write(f"**메타데이터 저장**: {len(metadata_list)}개 → `{metadata_path}`")
   
   if skipped_count > 0:
       st.info(f"너비 부족으로 스킵: {skipped_count}개")
   st.success(f"크롭 완료 - 기본: {basic_cropped_count}개 → `{basic_save_dir}`, 확장: {extended_cropped_count}개 → `{extended_save_dir}`")


def _execute_extended_crop_all(crop_pad, min_save_width, *params):
   """모든 이미지에 기본 크롭 + 확장 크롭 적용 (전체 데이터 새로 생성)"""
   
   # 파라미터 재구성
   params_dict = {
       'adaptive_block_size': params[0], 'adaptive_c': params[1], 'morph_kernel_size': params[2],
       'canny_low': params[3], 'canny_high': params[4], 'hough_threshold': params[5],
       'min_line_length': params[6], 'max_line_gap': params[7], 'line_angle_tolerance': params[8],
       'intersection_tolerance': params[9], 'min_area_threshold': params[10],
       'min_cell_height': params[11], 'max_cell_width': params[12], 
       'max_cell_height': params[13], 'min_valid_edges': params[14], 'distance_tolerance': params[15]
   }
   
   # 디렉토리 설정 - 2개 폴더
   image_dir = raw_data_folder_SD_processing_folder
   matched_dir = os.path.join(BASE_DIR, "matched_results")
   basic_save_dir = os.path.join(BASE_DIR, "SD_cell_extraction","SD_basic_crops")
   extended_save_dir = os.path.join(BASE_DIR,"SD_cell_extraction", "SD_extended_crops")
   intersections_dir = os.path.join(BASE_DIR,"SD_cell_extraction", "SD_intersections")
   metadata_path = os.path.join(BASE_DIR,"SD_cell_extraction","metadata" "SD_crop_metadata.json")
   os.makedirs(basic_save_dir, exist_ok=True)
   os.makedirs(extended_save_dir, exist_ok=True)
   os.makedirs(intersections_dir, exist_ok=True)
   
   image_files = [f for f in os.listdir(image_dir) if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
   if not image_files:
       st.error("이미지 파일이 없습니다.")
       return
   
   progress_bar = st.progress(0)
   status_text = st.empty()
   total_basic_cropped = 0
   total_extended_cropped = 0
   total_skipped = 0
   
   # 전체 메타데이터 수집용 (새로 생성)
   all_metadata = []
   
   for i, image_file in enumerate(image_files):
       progress_bar.progress((i + 1) / len(image_files))
       status_text.text(f"처리 중: {image_file} ({i+1}/{len(image_files)})")
       
       page_id = os.path.splitext(image_file)[0]
       image_path = os.path.join(image_dir, image_file)
       matched_path = os.path.join(matched_dir, f"{page_id}_matches.json")
       
       if not os.path.exists(matched_path):
           continue
       
       # 이미지 로드
       image = cv2.imread(image_path)
       if image is None:
           continue
       
       # 매칭 데이터 로드
       with open(matched_path, 'r', encoding='utf-8') as f:
           match_data = json.load(f)
       
       matched_texts = []
       for item in match_data:
           for matched in item.get("matched_items", []):
               text = matched["text"].strip()
               x0, y0, x1, y1 = matched["bbox"]
               cx, cy = (x0 + x1) / 2.0, (y0 + y1) / 2.0
               matched_texts.append({"text": text, "center": (cx, cy), "bbox": [x0, y0, x1, y1]})
       
       # 셀 검출
       horizontal_lines, vertical_lines, intersections, valid_cells = compute_lines_intersections_cells(image, params_dict)
       
       # 텍스트-셀 매칭
       text_cell_matches = _execute_text_cell_matching(matched_texts, valid_cells, params_dict['distance_tolerance'])
       
       # 교차점 정보 저장
       intersections_py = [[int(x), int(y)] for x, y in intersections]
       intersection_path = os.path.join(intersections_dir, f"{page_id}_intersections.json")
       with open(intersection_path, 'w', encoding='utf-8') as f:
           json.dump(intersections_py, f, indent=2)
       
       # 크롭 실행
       h, w = image.shape[:2]
       basic_cropped_count = 0
       extended_cropped_count = 0
       skipped_count = 0
       
       for match in text_cell_matches:
           if not match["matched_cells"]:
               continue
               
           cell_idx, current_cell = match["matched_cells"][0]
           text = match["text"]
           safe_text = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in text)[:20]
           
           # 1️⃣ 기본 셀 크롭
           basic_crop = _crop_cell(image, current_cell, pad=crop_pad)
           if basic_crop is not None:
               basic_width = basic_crop.shape[1]
               if basic_width >= min_save_width:
                   basic_filename = f"{page_id}_basic_{basic_cropped_count:03d}_{safe_text}.png"
                   basic_save_path = os.path.join(basic_save_dir, basic_filename)
                   cv2.imwrite(basic_save_path, basic_crop)
                   
                   # 기본 크롭 메타데이터 추가
                   basic_x0, basic_y0, basic_x1, basic_y1 = _cell_to_xyxy(current_cell)
                   basic_x0 = max(0, basic_x0 - crop_pad)
                   basic_y0 = max(0, basic_y0 - crop_pad)
                   basic_x1 = min(w, basic_x1 + crop_pad)
                   basic_y1 = min(h, basic_y1 + crop_pad)
                   
                   all_metadata.append({
                       "page_id": page_id,
                       "filename": basic_filename,
                       "bbox": [int(basic_x0), int(basic_y0), int(basic_x1), int(basic_y1)],
                       "text": text,
                       "crop_type": "basic",
                       "cell_idx": cell_idx
                   })
                   
                   basic_cropped_count += 1
               else:
                   skipped_count += 1
           
           # 2️⃣ 확장 크롭 (매칭된 셀들 간 거리 기반)
           curr_x0, curr_y0, curr_x1, curr_y1 = _cell_to_xyxy(current_cell)
           
           # 매칭된 셀들 중에서 다음 행의 상단 찾기
           next_row_top = _find_next_matched_row_top(current_cell, text_cell_matches, tolerance=150)
           
           if next_row_top:
               # 다음 매칭된 행의 상단까지 확장
               extended_y1 = next_row_top
           else:
               # 매칭된 다음 행이 없으면 이미지 끝까지
               extended_y1 = h
           
           # 패딩 적용해서 최종 크롭 영역 계산
           crop_x0 = max(0, curr_x0 - crop_pad)
           crop_y0 = max(0, curr_y0 - crop_pad)
           crop_x1 = min(w, curr_x1 + crop_pad)
           crop_y1 = min(h, extended_y1 + crop_pad)
           
           # 크롭 너비 계산
           crop_width = crop_x1 - crop_x0
           
           # 확장 크롭 저장
           if crop_width >= min_save_width and crop_x1 > crop_x0 and crop_y1 > crop_y0:
               extended_cropped = image[crop_y0:crop_y1, crop_x0:crop_x1]
               extended_filename = f"{page_id}_extended_{extended_cropped_count:03d}_{safe_text}.png"
               extended_save_path = os.path.join(extended_save_dir, extended_filename)
               cv2.imwrite(extended_save_path, extended_cropped)
               
               # 확장 크롭 메타데이터 추가
               all_metadata.append({
                   "page_id": page_id,
                   "filename": extended_filename,
                   "bbox": [int(crop_x0), int(crop_y0), int(crop_x1), int(crop_y1)],
                   "text": text,
                   "crop_type": "extended",
                   "cell_idx": cell_idx
               })
               
               extended_cropped_count += 1
           else:
               skipped_count += 1
       
       total_basic_cropped += basic_cropped_count
       total_extended_cropped += extended_cropped_count
       total_skipped += skipped_count
   
   progress_bar.empty()
   status_text.empty()
   
   # 전체 메타데이터 저장 (새로 생성)
   if all_metadata:
       with open(metadata_path, 'w', encoding='utf-8') as f:
           json.dump(all_metadata, f, indent=2, ensure_ascii=False)
       
       st.write(f"**전체 메타데이터 저장**: {len(all_metadata)}개 → `{metadata_path}`")
   
   if total_skipped > 0:
       st.info(f"너비 부족으로 전체 스킵: {total_skipped}개")
   st.success(f"모든 이미지 크롭 완료! 기본: {total_basic_cropped}개 → `{basic_save_dir}`, 확장: {total_extended_cropped}개 → `{extended_save_dir}`")



def _point_inside_rect_strict(px, py, rect, distance_tolerance=0):
    """점이 사각형 내부에 있는지 확인 (거리 허용치 포함)"""
    xs = [p[0] for p in rect]; ys = [p[1] for p in rect]
    x0, x1 = min(xs), max(xs); y0, y1 = min(ys), max(ys)
    
    # 기존 내부 체크
    if (x0 <= px < x1) and (y0 <= py < y1):
        return True
    
    # distance_tolerance가 0보다 크면 거리 허용
    if distance_tolerance > 0:
        dx = max(0, max(x0 - px, px - x1))
        dy = max(0, max(y0 - py, py - y1))
        distance = (dx**2 + dy**2)**0.5
        return distance <= distance_tolerance
    
    return False


def _select_best_cell_for_point(px, py, candidates):
    """점에 가장 적합한 셀 선택"""
    def key_fn(r):
        xs = [p[0] for p in r]; ys = [p[1] for p in r]
        cx = (min(xs) + max(xs)) / 2.0
        cy = (min(ys) + max(ys)) / 2.0
        area = max(1.0, (max(xs) - min(xs)) * (max(ys) - min(ys)))
        return (abs(py - cy), area)
    return min(candidates, key=key_fn) if candidates else None


# ===== NEW: 헬퍼 함수들 =====

def _save_params_to_session(*params):
    """파라미터를 세션에 저장"""
    st.session_state.line_detection_params = {
        'adaptive_block_size': params[0],
        'adaptive_c': params[1], 
        'morph_kernel_size': params[2],
        'canny_low': params[3],
        'canny_high': params[4],
        'hough_threshold': params[5],
        'min_line_length': params[6],
        'max_line_gap': params[7],
        'line_angle_tolerance': params[8],
        'intersection_tolerance': params[9],
        'min_area_threshold': params[10],
        'min_cell_height': params[11],
        'max_cell_width': params[12],
        'max_cell_height': params[13],
        'min_valid_edges': params[14],
        'distance_tolerance': params[15],
    }
    
    # 캐시 초기화
    if 'stage3_cache' in st.session_state:
        st.session_state.stage3_cache.clear()
    st.session_state.force_recompute_stage3 = True
    
    st.success("파라미터가 저장되었습니다!")


def _execute_text_cell_matching(matched_texts, valid_cells, distance_tolerance):
    """텍스트-셀 매칭 실행"""
    text_cell_matches = []
    
    for text_box in matched_texts:
        cx, cy = text_box["center"]
        candidates = [(i, cell) for i, cell in enumerate(valid_cells) 
                     if _point_inside_rect_strict(cx, cy, cell, distance_tolerance)]
        
        if candidates:
            best_cell = _select_best_cell_for_point(cx, cy, [c[1] for c in candidates])
            best_idx = next((i for i, c in enumerate(valid_cells) if c is best_cell or c == best_cell), candidates[0][0])
            text_cell_matches.append({"text": text_box["text"], "center": text_box["center"], "matched_cells": [(best_idx, best_cell)]})
        else:
            text_cell_matches.append({"text": text_box["text"], "center": text_box["center"], "matched_cells": []})
    
    return text_cell_matches


































def apply_surya_ocr_on_combined():
    """
    Run Surya OCR on all images in 'combined_crops' folder,
    collect result JSONs, and store them in the specified output folder.
    Progress is displayed; internal errors are handled silently.
    """
    combined_folder      = os.path.join(BASE_DIR, "SD_cell_extraction", "SD_extended_crops")
    surya_output_folder  = os.path.join(BASE_DIR, "SD_cell_extraction", "SD_img_Surya_output")
    os.makedirs(surya_output_folder, exist_ok=True)

    image_files = [f for f in os.listdir(combined_folder) if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
    if not image_files:
        st.error("❌ No target images found for OCR.")
        return

    st.info("🔍 Running Surya OCR on extracted figure images...")
    progress_bar = st.progress(0)
    status_text = st.empty()

    for idx, img_name in enumerate(image_files):
        img_path = os.path.join(combined_folder, img_name)
        command = ["surya_ocr", img_path]

        try:
            subprocess.run(command, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception as e:
            print(f"⚠️ Failed to run OCR on: {img_name} → {e}")
            continue

        progress = (idx + 1) / len(image_files)
        progress_bar.progress(progress)
        status_text.write(f"🖼️ Processing: {img_name} ({int(progress * 100)}%)")

    progress_bar.empty()
    status_text.write("✅ OCR processing completed for all images.")

    # Collect OCR result JSONs
    st.info("📦 Collecting OCR result files...")
    moved = 0
    for folder in os.listdir(SURYA_RESULTS_FOLDER):
        folder_path = os.path.join(SURYA_RESULTS_FOLDER, folder)
        if not os.path.isdir(folder_path):
            continue
        src = os.path.join(folder_path, "results.json")
        if not os.path.exists(src):
            st.warning(f"⚠️ results.json not found in: {folder}")
            continue
        dst = os.path.join(surya_output_folder, f"{folder}.json")
        try:
            shutil.move(src, dst)
            moved += 1
            st.markdown(f"📄 Moved: `{folder}.json`")
        except Exception as e:
            st.warning(f"❌ Failed to move result from {folder}: {e}")

    st.success(f"✅ Successfully saved {moved} OCR result file(s) to `{surya_output_folder}`")


#----------------------------------------------------------------------------------------------
def annotate_rebar_info():
    ocr_json_dir = os.path.join(BASE_DIR, "SD_cell_extraction", "SD_img_Surya_output")
    img_dir      = os.path.join(BASE_DIR, "SD_cell_extraction", "combined_crops")
    output_dir   = os.path.join(BASE_DIR, "SD_cell_extraction", "annotated_crops")
    os.makedirs(output_dir, exist_ok=True)

    rebar_pat = re.compile(r'(?:\d+-)?(?:SUHD|UHD|HD|SHD|D)[.\-]?\d+', re.IGNORECASE)
    floor_pat = re.compile(r'\b(?:B?\d+~)?(?:B?\d+F)\b', re.IGNORECASE)
    dim_pat   = re.compile(r'^\(?(\d+)\s*X\s*(\d+)\)?$', re.IGNORECASE)
    num_pat   = re.compile(r'^\d{1,4}(?:,\d{1,3})*$', re.IGNORECASE)
    dir_terms = {'both', 'all', 'center', 'cen', 'int', 'ext',"전단면","양단부","중앙부"}
    exclude_floor_keywords = {'TOP', 'BOTTOM', 'BOT', 'STIRRUP', 'BAR', 'FACE', 'APPROVAL'}

    tol_y = 50
    ratio_thresh = 1.05
    member_extra_gap = 50
    x_thresh = 100
    y_thresh = 30
    MAX_DIM = 3000

    all_results = []
    for fn in tqdm(os.listdir(ocr_json_dir), desc="Annotating"):
        if not fn.lower().endswith('.json'):
            continue
        base = os.path.splitext(fn)[0]
        json_path = os.path.join(ocr_json_dir, fn)
        img_path = os.path.join(img_dir, f"{base}.png")
        out_path = os.path.join(output_dir, f"{base}_annot.png")

        img = cv2.imread(img_path)
        if img is None:
            continue

        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        pages = next(iter(data.values()), [])

        boxes, labels, used_boxes = [], [], set()
        for page in pages:
            for itm in page.get('text_lines', []):
                raw_txt = itm.get('text', '').strip()
                norm_txt = unicodedata.normalize("NFKC", raw_txt).replace(" ", "")
                x0, y0, x1, y1 = itm.get('bbox', [0, 0, 0, 0])
                cx, cy = (x0 + x1) / 2, (y0 + y1) / 2
                boxes.append({'text': norm_txt, 'bbox': (x0, y0, x1, y1), 'cx': cx, 'cy': cy})

        if boxes:
            top_box = min(boxes, key=lambda b: b['bbox'][1])
            top_y = top_box['bbox'][1]
            if top_box['text'].lower() not in dir_terms and 'x' not in top_box['text'].lower():
                labels.append((top_box['bbox'], 'Member_Code', (255, 0, 0), top_box['text']))
                used_boxes.add(id(top_box))
            for b in boxes:
                if b is top_box or id(b) in used_boxes:
                    continue
                if b['text'].lower() in dir_terms or 'x' in b['text'].lower():
                    continue
                if 0 < b['bbox'][1] - top_y <= member_extra_gap:
                    labels.append((b['bbox'], 'Member_Code', (255, 0, 0), b['text']))
                    used_boxes.add(id(b))

        for b in boxes:
            txt_raw = b['text']
            txt_norm = unicodedata.normalize("NFKC", txt_raw).strip().lower().replace('.', '').replace('．', '')
            if id(b) in used_boxes:
                continue
            if any(s in txt_raw for s in [':', 'x', '×']) and txt_norm not in dir_terms:
                continue
            if txt_norm in dir_terms:
                labels.append((b['bbox'], 'Direction', (0, 255, 0), b['text']))
                used_boxes.add(id(b))

        stirrup_set = set()
        for b in boxes:
            txt = unicodedata.normalize("NFKC", b['text']).strip().lower()
            if 'x' in txt or '×' in txt:
                continue
            if id(b) in used_boxes:
                continue
            if '@' in txt or '＠' in txt:
                labels.append((b['bbox'], 'Stirrup', (255, 0, 255), b['text']))
                used_boxes.add(id(b))
                stirrup_set.add(id(b))

        rebar = []
        for b in boxes:
            if id(b) in used_boxes or 'x' in b['text'].lower():
                continue
            txt = b['text'].lower()
            if '@' in txt or '＠' in txt or ':' in txt:
                continue
            if rebar_pat.search(txt):
                rebar.append(b)

        if len(rebar) == 1:
            labels.append((rebar[0]['bbox'], 'Top_Rebar', (0, 0, 255), rebar[0]['text']))
            used_boxes.add(id(rebar[0]))
        elif len(rebar) == 2:
            r1, r2 = rebar
            x_diff = abs(r1['cx'] - r2['cx'])
            y_diff = abs(r1['cy'] - r2['cy'])
            if x_diff < x_thresh:
                sorted_rebar = sorted(rebar, key=lambda b: b['cy'])
                labels.append((sorted_rebar[0]['bbox'], 'Top_Rebar', (0, 0, 255), sorted_rebar[0]['text']))
                labels.append((sorted_rebar[1]['bbox'], 'Bot_Rebar', (0, 0, 255), sorted_rebar[1]['text']))
            elif y_diff < y_thresh:
                labels.append((r1['bbox'], 'Top_Rebar', (0, 0, 255), r1['text']))
                labels.append((r2['bbox'], 'Top_Rebar', (0, 0, 255), r2['text']))
            else:
                labels.append((r1['bbox'], 'Top_Rebar', (0, 0, 255), r1['text']))
                labels.append((r2['bbox'], 'Top_Rebar', (0, 0, 255), r2['text']))
            used_boxes.update([id(r1), id(r2)])
        elif len(rebar) >= 3:
            sorted_rebar = sorted(rebar, key=lambda b: b['cy'])
            top_y = sorted_rebar[0]['cy']
            top_rebars = [b for b in sorted_rebar if abs(b['cy'] - top_y) <= tol_y]
            bot_rebars = [b for b in sorted_rebar if abs(b['cy'] - top_y) > tol_y]
            for b in top_rebars:
                labels.append((b['bbox'], 'Top_Rebar', (0, 0, 255), b['text']))
                used_boxes.add(id(b))
            for b in bot_rebars:
                labels.append((b['bbox'], 'Bot_Rebar', (0, 0, 255), b['text']))
                used_boxes.add(id(b))

        height_candidates = []
        has_direction = any(lbl[1] == "Direction" for lbl in labels)
        for b in boxes:
            if id(b) in used_boxes or 'x' in b['text'].lower():
                continue
            txt = b['text'].replace("（", "(").replace("）", ")")
            if ':' in txt:
                continue
            x0, y0, x1, y1 = b['bbox']
            w, h = x1 - x0, y1 - y0
            m = dim_pat.match(txt)
            if m:
                wv, hv = int(m.group(1)), int(m.group(2))
                if wv <= MAX_DIM and hv <= MAX_DIM:
                    labels.append((b['bbox'], 'Width', (255, 165, 0), b['text']))
                    labels.append((b['bbox'], 'Height', (255, 165, 0), b['text']))
                    used_boxes.add(id(b))
                continue
            if num_pat.match(txt):
                val = int(txt.replace(',', ''))
                if val <= MAX_DIM:
                    ratio = w / h if h > 0 else float('inf')
                    if ratio > ratio_thresh:
                        labels.append((b['bbox'], 'Width', (255, 165, 0), b['text']))
                        used_boxes.add(id(b))
                    elif ratio < 1 / ratio_thresh:
                        height_candidates.append((b, val))
            else:
                num_txt = txt.replace(',', '')
                if num_txt.isdigit() and h > w * ratio_thresh:
                    height_candidates.append((b, int(num_txt)))

        if not has_direction and height_candidates:
            b, _ = max(height_candidates, key=lambda x: x[1])
            labels.append((b['bbox'], 'Height', (255, 165, 0), b['text']))
            used_boxes.add(id(b))
        elif has_direction:
            for b, _ in height_candidates:
                labels.append((b['bbox'], 'Height', (255, 165, 0), b['text']))
                used_boxes.add(id(b))

        for b in boxes:
            if id(b) in used_boxes or 'x' in b['text'].lower():
                continue
            txt = b['text']
            if any(ex in txt.upper() for ex in exclude_floor_keywords):
                continue
            if floor_pat.fullmatch(txt):
                labels.append((b['bbox'], 'Floor', (0, 128, 255), b['text']))
                used_boxes.add(id(b))

        for (x0, y0, x1, y1), label, color, _ in labels:
            cv2.rectangle(img, (int(x0), int(y0)), (int(x1), int(y1)), color, 2)
            cv2.putText(img, label, (int(x0), int(y0) - 6),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 1)

        cv2.imwrite(out_path, img)
        print("✅ Annotated image saved:", out_path)
        all_results.append((labels, base))  # ✅ 이걸로 변경

    return all_results  # ✅ 여러 이미지 처리 결과 반환


#----------------------------------------------------------------------
def export_merged_excel(results):
    """
    Parameters:
        results: list of (labels, base) from annotate_rebar_info()
    Output:
        하나의 병합된 Excel 파일 (BASE_DIR/SD_cell_extraction/merged_elements.xlsx)
    """

    if not results:
        print("❌ results가 None이거나 비어 있음. export_merged_excel 중단.")
        return

    all_rows = []

    for labels, base in results:
        # Step 1: Member 추출
        member_texts = [t for (_, label, _, t) in labels if label == "Member_Code"]
        member = member_texts[0] if member_texts else base

        # Step 2: Direction
        directions = []
        for bbox, label, _, text in labels:
            if label == 'Direction':
                x0, y0, x1, y1 = bbox
                cx = (x0 + x1) / 2
                directions.append({'text': text, 'cx': cx})

        if not directions:
            directions = [{'text': '', 'cx': 0}]

        # Step 3: 클러스터링
        data = {d['text']: {
            'Top_Rebar': '',
            'Bot_Rebar': '',
            'Stirrup': '',
            'Width': '',
            'Height': ''
        } for d in directions}

        TARGET_LABELS = {'Top_Rebar', 'Bot_Rebar', 'Stirrup', 'Width', 'Height'}
        for bbox, label, _, text in labels:
            if label not in TARGET_LABELS:
                continue
            x0, _, x1, _ = bbox
            cx = (x0 + x1) / 2
            closest = min(directions, key=lambda d: abs(d['cx'] - cx))
            dir_key = closest['text']
            if label == 'Stirrup':
                existing = data[dir_key][label]
                if text not in existing.split(', '):  # ✅ 중복 방지
                    if existing:
                        data[dir_key][label] += ', ' + text
                    else:
                        data[dir_key][label] = text
            else:
                data[dir_key][label] = text

        for d in directions:
            d_text = d['text']
            values = data[d_text]
            row = {
                'Image': base,
                'Member': member,
                'Direction': d_text,
                'Top_Rebar': values['Top_Rebar'],
                'Bot_Rebar': values['Bot_Rebar'],
                'Stirrup': values['Stirrup'],
                'Width': values['Width'],
                'Height': values['Height'],
            }
            all_rows.append(row)

    df = pd.DataFrame(all_rows, columns=["Image", "Member", "Direction", "Top_Rebar", "Bot_Rebar", "Stirrup", "Width", "Height"])
    save_path = os.path.join(BASE_DIR, "SD_cell_extraction", "merged_elements.xlsx")
    df.to_excel(save_path, index=False)
    
    print(f"✅ 병합된 Excel 저장 완료: {save_path}")



#######################################################


def annotate_rebar_info2():
    """
    1. combined_crops 폴더 내 이미지들
    2. 반시계 90도 회전
    3. 회전 후, 중앙 기준 좌/우 25% 이내 영역만 유지 (중앙 50% crop)
    4. combined_crops_rotated에 저장
    """

    img_dir = os.path.join(BASE_DIR, "SD_cell_extraction", "combined_crops")
    out_dir = os.path.join(BASE_DIR, "SD_cell_extraction", "combined_crops_rotated")
    os.makedirs(out_dir, exist_ok=True)

    image_ext = ['.png', '.jpg', '.jpeg']

    for filename in os.listdir(img_dir):
        if any(filename.lower().endswith(ext) for ext in image_ext):
            img_path = os.path.join(img_dir, filename)
            out_path = os.path.join(out_dir, filename)

            try:
                # 1. 이미지 열기
                img = Image.open(img_path)

                # 2. 반시계 90도 회전
                rotated = img.rotate(-90, expand=True)

                # 3. 중앙 50% 영역만 자르기 (좌우 25% 제외)
                width, height = rotated.size
                left = int(width * 0.25)
                right = int(width * 0.75)
                cropped = rotated.crop((left, 0, right, height))  # (left, upper, right, lower)

                # 4. 저장
                cropped.save(out_path)

            except Exception as e:
                print(f"❌ {filename} 처리 실패: {e}")





def apply_surya_ocr_SD_Height():
    os.makedirs(surya_output_folder, exist_ok=True)

    # ✅ 결과 JSON 파일이 이미 존재하면 재실행 생략
    existing_jsons = [f for f in os.listdir(surya_output_folder) if f.endswith(".json")]
    if existing_jsons:
        st.info(f"ℹ️ 이미 {len(existing_jsons)}개의 OCR 결과가 존재합니다. Surya OCR 생략.")
        return

    image_files = [f for f in os.listdir(plain_text_folder) if f.endswith(('.jpg', '.png'))]
    if not image_files:
        st.error("❌ No files to apply OCR")
        return

    progress_bar = st.progress(0)
    status_text = st.empty()
    st.write("🔍Running OCR...")

    for idx, image_file in enumerate(image_files):
        input_path = os.path.join(plain_text_folder, image_file)
        command = ["surya_ocr", input_path]
        result = subprocess.run(command, capture_output=True, text=True)

        stderr = result.stderr.strip()
        if stderr and all(x not in stderr for x in ["Detecting bboxes", "Recognizing Text"]):
            st.warning(f"⚠️ 오류: {image_file} → {stderr}")

        progress = int((idx + 1) / len(image_files) * 100)
        progress_bar.progress(progress)
        status_text.write(f"🖼️ running ocr: {image_file} ({progress}%)")

    progress_bar.empty()
    status_text.write("✅ OCR complete! Result moving...")

    moved, skipped = 0, 0
    for folder_name in os.listdir(SURYA_RESULTS_FOLDER):
        folder_path = os.path.join(SURYA_RESULTS_FOLDER, folder_name)
        if os.path.isdir(folder_path):
            json_file = os.path.join(folder_path, "results.json")
            if os.path.exists(json_file):
                dst_file = os.path.join(surya_output_folder, f"{folder_name}.json")
                try:
                    shutil.move(json_file, dst_file)
                    moved += 1
                except Exception as e:
                    st.error(f"❌ Move Error: {folder_name} → {e}")
            else:
                skipped += 1

    st.success(f"📁 OCR 결과 {moved}개 이동 완료 ({skipped}개는 누락됨)")












def apply_surya_ocr_on_rotated_images():
    Height_SD_rocog = r"D:\codeforpaper\SD_cell_extraction\combined_crops_rotated"
    surya_output_folder = r"D:\codeforpaper\SD_cell_extraction\ocr_results_rotated"

    os.makedirs(surya_output_folder, exist_ok=True)

    image_files = [f for f in os.listdir(Height_SD_rocog) if f.endswith(('.jpg', '.png'))]
    if not image_files:
        st.error("❌ No files to apply OCR")
        return

    progress_bar = st.progress(0)
    status_text = st.empty()
    st.write("🔍 Running Surya OCR...")

    for idx, image_file in enumerate(image_files):
        input_path = os.path.join(Height_SD_rocog, image_file)
        command = ["surya_ocr", input_path]
        result = subprocess.run(command, capture_output=True, text=True)

        stderr = result.stderr.strip()
        if stderr and all(x not in stderr for x in ["Detecting bboxes", "Recognizing Text"]):
            st.warning(f"⚠️ 오류: {image_file} → {stderr}")

        progress = int((idx + 1) / len(image_files) * 100)
        progress_bar.progress(progress)
        status_text.write(f"🖼️ running OCR: {image_file} ({progress}%)")

    progress_bar.empty()
    status_text.write("📂 OCR complete! Collecting results...")

    moved, skipped = 0, 0
    for folder_name in os.listdir(SURYA_RESULTS_FOLDER):
        folder_path = os.path.join(SURYA_RESULTS_FOLDER, folder_name)
        if os.path.isdir(folder_path):
            json_file = os.path.join(folder_path, "results.json")
            if os.path.exists(json_file):
                dst_file = os.path.join(surya_output_folder, f"{folder_name}.json")
                try:
                    shutil.move(json_file, dst_file)
                    moved += 1
                except Exception as e:
                    st.error(f"❌ Move Error: {folder_name} → {e}")
            else:
                skipped += 1

    st.success(f"📁 OCR 결과 {moved}개 이동 완료 ({skipped}개는 누락됨)")













def visualize_and_extract_ranked_heights(save_csv=True):
    st.subheader("🔍 OCR 숫자 시각화 + Height1~3 추출 (이미지 3장만 시각화)")

    # 📂 폴더 경로 입력
    ocr_folder = st.text_input(
        "📂 OCR 결과 JSON 폴더 경로",
        value= os.path.join(BASE_DIR,"SD_cell_extraction","ocr_results_rotated")
    )
    image_folder = st.text_input(
        "🖼️ 원본 이미지 폴더 경로",
        value= os.path.join(BASE_DIR,"SD_cell_extraction","combined_crops_rotated")
    )

    output_csv_path = st.text_input(
        "💾 저장할 CSV 파일 경로", 
        value= os.path.join(BASE_DIR,"SD_cell_extraction","height_results.csv")
    )

    if not os.path.isdir(ocr_folder) or not os.path.isdir(image_folder):
        st.warning("❗ 유효한 폴더 경로를 입력해주세요.")
        return

    if st.button("시각화 및 Height 추출 실행"):
        json_files = [f for f in os.listdir(ocr_folder) if f.endswith(".json")]
        if not json_files:
            st.error("❌ OCR 결과 JSON 파일이 없습니다.")
            return

        all_results = {}
        vis_count = 0

        for json_file in json_files:
            json_path = os.path.join(ocr_folder, json_file)
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            for image_name, content_list in data.items():
                image_path = os.path.join(image_folder, image_name + ".png")
                if not os.path.exists(image_path):
                    st.warning(f"❌ 이미지 없음: {image_path}")
                    continue

                image = Image.open(image_path).convert("RGB")
                draw = ImageDraw.Draw(image)

                numbers = []

                for entry in content_list:
                    for line in entry.get("text_lines", []):
                        text = line.get("text", "")
                        bbox = line.get("bbox", [])
                        cleaned_text = text.replace(",", "")

                        if cleaned_text.isdigit():
                            value = int(cleaned_text)
                            if value >= 100 and len(bbox) == 4:
                                y_center = (bbox[1] + bbox[3]) / 2
                                numbers.append((value, y_center, bbox, text))

                numbers_sorted = sorted(numbers, key=lambda x: x[1])

                height_dict = {}
                for i, (value, _, bbox, raw_text) in enumerate(numbers_sorted[:3]):
                    draw.rectangle(bbox, outline="red", width=2)
                    draw.text((bbox[0], bbox[1] - 10), f"{raw_text} (H{i+1})", fill="red")
                    height_dict[f"Height{i+1}"] = value

                all_results[image_name] = height_dict

                if vis_count < 3:
                    st.image(image, caption=image_name, use_column_width=True)
                    st.json(height_dict)
                    vis_count += 1

        # ✅ CSV 저장
        if save_csv:
            df_heights = pd.DataFrame.from_dict(all_results, orient='index').reset_index()
            df_heights = df_heights.rename(columns={'index': 'Image'})
            df_heights.to_csv(output_csv_path, index=False, encoding='utf-8-sig')
            st.success(f"✅ Height 정보 CSV 저장 완료: {output_csv_path}")









def replace_height_in_merged_elements():
    st.subheader("📐 Height 치환 병합 (Direction 기준, CSV 또는 XLSX 저장 지원)")

    # 🔹 입력 경로
    elements_path = st.text_input("📥 merged_elements 경로 (.csv 또는 .xlsx)", 
                                  value= os.path.join(BASE_DIR,"SD_cell_extraction","merged_elements.xlsx"))
    height_path = st.text_input("📥 height_results.csv 경로", 
                                value=os.path.join(BASE_DIR,"SD_cell_extraction","height_results.csv"))
    save_path = st.text_input("💾 저장 경로 (.csv 또는 .xlsx)", 
                              value=os.path.join(BASE_DIR,"SD_cell_extraction","merged_with_height.xlsx"))

    if st.button("병합 및 Height 치환 실행"):
        try:
            # 1. 파일 로딩
            if elements_path.endswith(".xlsx"):
                df_elements = pd.read_excel(elements_path)
            else:
                df_elements = pd.read_csv(elements_path)

            df_heights = pd.read_csv(height_path)

            # 2. 열 이름 정리
            df_elements.rename(columns={df_elements.columns[0]: "Key"}, inplace=True)
            df_heights.rename(columns={df_heights.columns[0]: "Key"}, inplace=True)

            # 3. Dictionary 형태로 height 추출
            height_dict = df_heights.set_index("Key").to_dict(orient="index")

            # 4. Height 치환 로직
            def replace_height(row):
                key = row["Key"]
                direction = str(row.get("Direction", "")).strip().upper()

                heights = height_dict.get(key, None)
                if not heights:
                    return row.get("Height")  # height 없으면 원래값 유지

                h1 = heights.get("Height1")
                h2 = heights.get("Height2")
                h3 = heights.get("Height3")

                # 3개 존재
                if pd.notna(h1) and pd.notna(h2) and pd.notna(h3):
                    if "INT" in direction:
                        return h1
                    elif "CENTER" in direction:
                        return h2
                    elif "EXT" in direction:
                        return h3

                # 2개 존재
                elif pd.notna(h1) and pd.notna(h2):
                    if "BOTH" in direction:
                        return h1
                    elif "CENTER" in direction:
                        return h2

                # 1개만 존재
                elif pd.notna(h1):
                    return h1

                return row.get("Height")  # fallback

            # 5. 적용
            df_elements["Height"] = df_elements.apply(replace_height, axis=1)

            # 6. 저장
            if save_path.endswith(".xlsx"):
                df_elements.to_excel(save_path, index=False)
            else:
                df_elements.to_csv(save_path, index=False, encoding="utf-8-sig")

            st.success(f"✅ 저장 완료: {save_path}")
            st.dataframe(df_elements.head())

        except Exception as e:
            st.error(f"❌ 오류 발생: {e}")













#######################################################



















def expand_floor_range(start, end, pitf_max=None):
    def parse(val):
        val = val.strip().upper()
        if val.startswith("PIT"):
            return pitf_max if pitf_max is not None else "PIT"
        val = val.replace("F", "").replace("B", "-")
        if re.match(r"-?\d+", val):
            return int(val)
        return val

    s, e = parse(start), parse(end)

    if isinstance(s, str) or isinstance(e, str):
        return [str(s), str(e)] if str(s) != str(e) else [str(s)]

    step = 1 if s < e else -1
    return [i for i in range(s, e + step, step) if i != 0]

def standardize_member_column(pitf_max=None):
    input_path = os.path.join(BASE_DIR, "SD_cell_extraction", "merged_with_height.xlsx")
    output_path = os.path.join(BASE_DIR, "SD_member_split.csv")

    if not os.path.exists(input_path):
        print(f"❌ 파일 없음: {input_path}")
        return

    df = pd.read_excel(input_path)
    split_rows = []

    for _, row in df.iterrows():
        member_raw = str(row["Member"])
        parts = [p.strip() for p in member_raw.split(",") if p.strip()]
        new_members = []
        i = 0

        while i < len(parts):
            part = parts[i]
            part = re.sub(r'(\-?\d+~\-?\d+)([A-Za-z]+\d+[A-Za-z0-9]*)', r'\1 \2', part)

            # ✅ (2~4)B5 형태 처리
            match = re.match(r'^\((\-?\d+)~(\-?\d+)\)([A-Za-z]+\d+[A-Za-z0-9]*)$', part)
            if match:
                start, end, code = match.groups()
                floors = expand_floor_range(start, end, pitf_max)
                for fl in floors:
                    new_members.append(f"{fl} {code}")
                i += 1
                continue

            # ✅ B3F~PITF C1, C2, C3 → 확장
            match = re.match(r'([BF]?\-?\d+F?|PITF?)\s*~\s*([BF]?\-?\d+F?|PITF?)\s+([A-Za-z]+\d+[A-Za-z0-9]*)$', part, re.IGNORECASE)
            if match:
                f1, f2, code1 = match.groups()
                floors = expand_floor_range(f1, f2, pitf_max)
                codes = [code1]

                j = i + 1
                while j < len(parts) and re.match(r'^[A-Za-z]+\d+[A-Za-z0-9]*$', parts[j]):
                    codes.append(parts[j])
                    j += 1

                for fl in floors:
                    for c in codes:
                        new_members.append(f"{fl} {c}")
                i = j
                continue

            # ✅ -2~3 RMG2 → 층수별 확장
            match = re.match(r'([\-]?\d+~\-?\d+)\s+([A-Za-z]+\d+[A-Za-z0-9]*)$', part)
            if match:
                floor_range_str, code = match.groups()
                start, end = floor_range_str.split("~")
                floors = expand_floor_range(start, end, pitf_max)
                for fl in floors:
                    new_members.append(f"{fl} {code}")
                i += 1
                continue

            # ✅ 공백 있는 층수+코드 (ex: B3F TC10A)
            if re.match(r'^([BF]?\-?\d+F?)\s+([A-Za-z]+\d+[A-Za-z0-9]*)$', part, re.IGNORECASE):
                f, c = re.match(r'^([BF]?\-?\d+F?)\s+([A-Za-z]+\d+[A-Za-z0-9]*)$', part, re.IGNORECASE).groups()
                fl = expand_floor_range(f, f, pitf_max)[0]
                new_members.append(f"{fl} {c}")
                i += 1
                continue

            # ✅ 붙어 있는 층수+코드 (ex: 1B21A → 1 B21A)
            if re.match(r'^([BF]?\-?\d+)([A-Za-z]+\d*[A-Za-z0-9]*)$', part, re.IGNORECASE):
                f, c = re.match(r'^([BF]?\-?\d+)([A-Za-z]+\d*[A-Za-z0-9]*)$', part, re.IGNORECASE).groups()
                fl = expand_floor_range(f, f, pitf_max)[0]
                new_members.append(f"{fl} {c}")
                i += 1
                continue

            # ✅ 기타 그대로 추가
            new_members.append(part.strip())
            i += 1

        for member in new_members:
            new_row = row.copy()
            new_row["Member"] = member
            split_rows.append(new_row)

    result_df = pd.DataFrame(split_rows)
    result_df["Member"] = result_df["Member"].str.strip()
    result_df.to_csv(output_path, index=False, encoding='utf-8-sig')
    print(f"✅ Member 정규화 + 층수/코드 확장 완료! → {output_path}")






def expand_direction_column():
    input_path = os.path.join(BASE_DIR, "SD_member_split.csv")
    output_path = os.path.join(BASE_DIR, "SD_member_final.csv")

    if not os.path.exists(input_path):
        print(f"❌ 파일 없음: {input_path}")
        return

    df = pd.read_csv(input_path)

    if "Direction" not in df.columns:
        print("❌ 'Direction' 컬럼 없음 (대소문자 확인 필요)")
        return

    expanded_rows = []

    for _, row in df.iterrows():
        raw_val = row["Direction"]
        if pd.isna(raw_val):
            mapped = [""]
        else:
            dir_val = str(raw_val).strip().lower().replace(".", "")
            if dir_val in ["center", "cen"]:
                mapped = ["MID"]
            elif dir_val == "ext":
                mapped = ["END-J"]
            elif dir_val == "int":
                mapped = ["END-I"]
            elif dir_val == "both":
                mapped = ["END-I", "END-J"]
            elif dir_val == "all":
                mapped = ["MID", "END-J", "END-I"]
            else:
                mapped = [dir_val.upper()]

        for m in mapped:
            new_row = row.copy()
            new_row["Direction"] = m
            expanded_rows.append(new_row)

    result_df = pd.DataFrame(expanded_rows)
    result_df.to_csv(output_path, index=False, encoding='utf-8-sig')
    print(f"✅ Direction 변환 + NaN 처리 완료! → {output_path}")

##############################################################################################
#SDD complete
#############################################################################################

# def detect_duplication_error(scd_df: pd.DataFrame, sd_df: pd.DataFrame) -> pd.DataFrame:
#     """
#     Duplication 오류를 탐지하는 함수
#     - Member + Direction 조합이 중복되며, 다른 하위 정보가 존재할 경우 'Duplication in SCD'로 간주
#     - 결과는 [Member, Direction, Error, Detail] 구성의 DataFrame으로 반환
#     """

#     # 1. Direction 있는 경우만 추출
#     scd = scd_df[scd_df['Direction'].notna()].copy()
#     sd = sd_df[sd_df['Direction'].notna()].copy()

#     # 2. 비교 키 생성: Member + Direction
#     scd['Key'] = scd['Member'].astype(str).str.strip() + '|' + scd['Direction'].astype(str).str.strip()

#     # 3. Key별로 중복 여부와 하위 정보(Top_Rebar, Stirrup 등) 차이 확인
#     duplicated_errors = []

#     for key, group in scd.groupby('Key'):
#         if len(group) > 1:
#             # 하위 정보 중복 여부 판단
#             sub_cols = ['Top_Rebar', 'Bot_Rebar', 'Stirrup', 'Width', 'Height', 'fck', 'fy']
#             unique_rows = group[sub_cols].drop_duplicates()
#             if len(unique_rows) > 1:
#                 # 오류로 판정
#                 member, direction = key.split('|')
#                 duplicated_errors.append({
#                     'Member': member,
#                     'Direction': direction,
#                     'Error': 'Duplication',
#                     'Detail': 'Duplication in SCD'
#                 })

#     return pd.DataFrame(duplicated_errors)



##########################################################################################################
#정보추출정확도#########################################
##########################################################################################################




def evaluate_extraction_accuracy_SCD():
    st.subheader("✅ 정보 추출 정확도 평가 (SCD)")

    gt_file = st.file_uploader("📥 SCD Ground Truth 파일 업로드", type=["csv", "xlsx"], key="gt_scd")
    pred_file = st.file_uploader("📥 SCD 추출 결과 파일 업로드", type=["csv", "xlsx"], key="pred_scd")

    save_path = st.text_input("💾 저장할 결과 파일 경로 (.xlsx)", value=os.path.join(BASE_DIR,"SCD_extraction_evaluation_result.xlsx"))

    def load_any_file(file):
        if file.name.endswith(".csv"):
            return pd.read_csv(file)
        elif file.name.endswith(".xlsx"):
            return pd.read_excel(file)

    if st.button("SCD 정확도 평가 실행"):
        if not gt_file or not pred_file:
            st.warning("⚠️ 두 파일을 모두 업로드해주세요.")
            return

        df_gt = load_any_file(gt_file)
        df_pred = load_any_file(pred_file)

        # ✅ 여기서 세션 상태에 저장
        st.session_state.df_gt_scd = df_gt

        if df_gt is None or df_pred is None:
            st.error("❌ 파일 형식이 잘못되었습니다. csv 또는 xlsx만 허용됩니다.")
            return

        compare_cols = ["Top_Rebar", "Bot_Rebar", "Stirrup", "Width", "Height"]

        # 1. Key 구성
        df_gt["__key__"] = df_gt["Member"].astype(str).str.strip() + "|" + df_gt["Direction"].astype(str).str.strip()
        df_pred["__key__"] = df_pred["Member"].astype(str).str.strip() + "|" + df_pred["Direction"].astype(str).str.strip()

        df_merge = pd.merge(df_gt, df_pred, on="__key__", how="outer", suffixes=("_gt", "_pred"))

        # 2. 정확도 계산 및 오류 추적
        total = 0
        correct = 0
        error_cells = []
        error_details = []

        for col in compare_cols:
            col_gt = col + "_gt"
            col_pred = col + "_pred"
            if col_gt in df_merge.columns and col_pred in df_merge.columns:
                for i in df_merge.index:
                    val_gt = str(df_merge.at[i, col_gt]).strip()
                    val_pred = str(df_merge.at[i, col_pred]).strip()
                    if val_gt == val_pred:
                        correct += 1
                    else:
                        error_cells.append((i, col_pred))
                        member = str(df_merge.at[i, "Member_gt"]) if "Member_gt" in df_merge.columns else ""
                        direction = str(df_merge.at[i, "Direction_gt"]) if "Direction_gt" in df_merge.columns else ""
                        error_details.append(f"{member} {direction} {col}")
                    total += 1

        accuracy = (correct / total) * 100 if total > 0 else 0

        # 3. 결과 시트 구성
        result_df = df_merge[["Member_gt", "Direction_gt"] + [c + "_pred" for c in compare_cols]].copy()
        result_df.rename(columns={"Member_gt": "Member", "Direction_gt": "Direction"}, inplace=True)

        # 4. 저장 (엑셀 + 배경 강조)
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            result_df.to_excel(tmp.name, sheet_name="Evaluation", index=False)
            wb = load_workbook(tmp.name)
            ws = wb["Evaluation"]

            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for i, col in error_cells:
                col_idx = compare_cols.index(col.replace("_pred", "")) + 3  # Member, Direction → +2 → +1 more (1-indexed)
                ws.cell(row=i + 2, column=col_idx).fill = yellow_fill

            # 요약 시트 추가
            summary_ws = wb.create_sheet("Summary")
            summary_ws.append(["총 비교 셀 수", total])
            summary_ws.append(["일치한 셀 수", correct])
            summary_ws.append(["정확도 (%)", round(accuracy, 2)])
            summary_ws.append([])
            summary_ws.append(["❌ 불일치 내역"])
            for err in error_details:
                summary_ws.append([err])

            wb.save(save_path)

        st.success(f"✅ 정확도 평가 완료! 결과 파일 저장됨: {save_path}")
        st.write(f"🎯 정확도: {round(accuracy, 2)}%")





def evaluate_extraction_accuracy_SD():
    st.subheader("✅ 정보 추출 정확도 평가 (SD)")

    gt_file = st.file_uploader("📥 SD Ground Truth 파일 업로드", type=["csv", "xlsx"], key="gt_sd")
    pred_file = st.file_uploader("📥 SD 추출 결과 파일 업로드", type=["csv", "xlsx"], key="pred_sd")

    save_path = st.text_input("💾 저장할 결과 파일 경로 (.xlsx)", value=os.path.join(BASE_DIR, "SD_extraction_evaluation_result.xlsx"))

    def load_any_file(file):
        if file.name.endswith(".csv"):
            return pd.read_csv(file)
        elif file.name.endswith(".xlsx"):
            return pd.read_excel(file)

    if st.button("SD 정확도 평가 실행"):
        if not gt_file or not pred_file:
            st.warning("⚠️ 두 파일을 모두 업로드해주세요.")
            return

        df_gt = load_any_file(gt_file)
        df_pred = load_any_file(pred_file)

        # ✅ 세션 상태에 저장
        st.session_state.df_gt_sd = df_gt

        if df_gt is None or df_pred is None:
            st.error("❌ 파일 형식이 잘못되었습니다. csv 또는 xlsx만 허용됩니다.")
            return

        compare_cols = ["Top_Rebar", "Bot_Rebar", "Stirrup", "Width", "Height"]

        # 1. Key 구성
        df_gt["__key__"] = df_gt["Member"].astype(str).str.strip() + "|" + df_gt["Direction"].astype(str).str.strip()
        df_pred["__key__"] = df_pred["Member"].astype(str).str.strip() + "|" + df_pred["Direction"].astype(str).str.strip()

        df_merge = pd.merge(df_gt, df_pred, on="__key__", how="outer", suffixes=("_gt", "_pred"))

        # 2. 정확도 계산 및 오류 추적
        total = 0
        correct = 0
        error_cells = []
        error_details = []

        for col in compare_cols:
            col_gt = col + "_gt"
            col_pred = col + "_pred"
            if col_gt in df_merge.columns and col_pred in df_merge.columns:
                for i in df_merge.index:
                    val_gt = str(df_merge.at[i, col_gt]).strip()
                    val_pred = str(df_merge.at[i, col_pred]).strip()
                    if val_gt == val_pred:
                        correct += 1
                    else:
                        error_cells.append((i, col_pred))
                        member = str(df_merge.at[i, "Member_gt"]) if "Member_gt" in df_merge.columns else ""
                        direction = str(df_merge.at[i, "Direction_gt"]) if "Direction_gt" in df_merge.columns else ""
                        error_details.append(f"{member} {direction} {col}")
                    total += 1

        accuracy = (correct / total) * 100 if total > 0 else 0

        # 3. 결과 시트 구성
        result_df = df_merge[["Member_gt", "Direction_gt"] + [c + "_pred" for c in compare_cols]].copy()
        result_df.rename(columns={"Member_gt": "Member", "Direction_gt": "Direction"}, inplace=True)

        # 4. 저장 (엑셀 + 배경 강조)
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            result_df.to_excel(tmp.name, sheet_name="Evaluation", index=False)
            wb = load_workbook(tmp.name)
            ws = wb["Evaluation"]

            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for i, col in error_cells:
                col_idx = compare_cols.index(col.replace("_pred", "")) + 3  # Member, Direction → +2 → +1 more (1-indexed)
                ws.cell(row=i + 2, column=col_idx).fill = yellow_fill

            # 요약 시트 추가
            summary_ws = wb.create_sheet("Summary")
            summary_ws.append(["총 비교 셀 수", total])
            summary_ws.append(["일치한 셀 수", correct])
            summary_ws.append(["정확도 (%)", round(accuracy, 2)])
            summary_ws.append([])
            summary_ws.append(["❌ 불일치 내역"])
            for err in error_details:
                summary_ws.append([err])

            wb.save(save_path)

        st.success(f"✅ 정확도 평가 완료! 결과 파일 저장됨: {save_path}")
        st.write(f"🎯 정확도: {round(accuracy, 2)}%")





def generate_ground_truth_error_report(df_scd, df_sd, save_path):
    errors = detect_structural_errors(df_scd, df_sd)

    summary = (
        errors.groupby(["Error Type", "Detail"])
        .size()
        .reset_index(name="Count")
        .sort_values("Count", ascending=False)
    )
    summary.loc[len(summary)] = ["Total", "", summary["Count"].sum()]

    with pd.ExcelWriter(save_path, engine="openpyxl", mode="w") as writer:
        errors.to_excel(writer, index=False, sheet_name="Detailed Errors")
        summary.to_excel(writer, index=False, sheet_name="Error Summary")

    wb = load_workbook(save_path)
    ws = wb["Error Summary"]

    merge_cells_in_column(ws, 1)

    align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = align
            cell.border = border

    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(save_path)













def parse_rebar(value):
    try:
        val = str(value).strip()
        numbers = re.findall(r"\d+", val)
        if len(numbers) >= 2:
            count = int(numbers[0])
            dia = int(numbers[-1])
            return (count, dia)
        match = re.match(r"(\d+)[-~]?[A-Z]*[^\d]*(\d+)", val)
        if match:
            return (int(match.group(1)), int(match.group(2)))
    except:
        pass
    return None

def extract_stirrup_spacing(val):
    try:
        match = re.search(r"@(\d+)", str(val))
        if match:
            return int(match.group(1))
    except:
        pass
    return None

def normalize_dimension(val):
    try:
        return float(str(val).replace(",", "").strip())
    except:
        return None

def dimension_match(v1, v2):
    if v1 == v2:
        return True
    elif v1 is None or v2 is None:
        return False
    elif abs(v1 * 1000 - v2) < 1e-3 or abs(v1 - v2 * 1000) < 1e-3:
        return True
    return False

def detect_structural_errors(df1, df2, label1="SCD", label2="SD"):
    errors = []
    df1["__key__"] = df1["Member"].astype(str).str.strip() + "|" + df1["Direction"].astype(str).str.strip()
    df2["__key__"] = df2["Member"].astype(str).str.strip() + "|" + df2["Direction"].astype(str).str.strip()
    df1_map = df1.drop_duplicates(subset="__key__").set_index("__key__").to_dict(orient="index")
    df2_map = df2.drop_duplicates(subset="__key__").set_index("__key__").to_dict(orient="index")
    keys_all = set(df1_map.keys()).union(set(df2_map.keys()))

    for key in keys_all:
        row1 = df1_map.get(key)
        row2 = df2_map.get(key)
        member, direction = key.split("|")

        if row1 is None:
            errors.append([member, direction, "Missing entry", f"Missing entry in {label1}"])
            continue
        elif row2 is None:
            errors.append([member, direction, "Missing entry", f"Missing entry in {label2}"])
            continue

        for col in ["Top_Rebar", "Bot_Rebar"]:
            r1 = parse_rebar(row1.get(col))
            r2 = parse_rebar(row2.get(col))
            if r1 and r2:
                count1, dia1 = r1
                count2, dia2 = r2
                if count1 != count2:
                    errors.append([member, direction, "Main Rebar mismatch", f"{col} count error"])
                if dia1 != dia2:
                    errors.append([member, direction, "Main Rebar mismatch", f"{col} diameter error"])
            elif r1 != r2:
                errors.append([member, direction, "Main Rebar Error", f"{col} format error"])

        s1 = extract_stirrup_spacing(row1.get("Stirrup"))
        s2 = extract_stirrup_spacing(row2.get("Stirrup"))
        if s1 != s2:
            errors.append([member, direction, "Stirrup mismatch", f"Stirrup spacing error"])

        for col in ["Width", "Height"]:
            v1 = normalize_dimension(row1.get(col))
            v2 = normalize_dimension(row2.get(col))
            if not dimension_match(v1, v2):
                errors.append([member, direction, "Dimension mismatch", f"{col} error"])

    for label, df in zip([label1, label2], [df1, df2]):
        duplicated_keys = df["__key__"].value_counts()
        duplicated_keys = duplicated_keys[duplicated_keys > 1]
        for dup_key in duplicated_keys.index:
            member, direction = dup_key.split("|")
            errors.append([member, direction, "Duplication", f"Duplication in {label}"])

    return pd.DataFrame(errors, columns=["Member", "Direction", "Error Type", "Detail"])

# 셀 병합 함수
def merge_cells_in_column(ws, col_idx):
    current_val = None
    start_row = 2
    for row in range(2, ws.max_row + 2):
        cell_val = ws.cell(row=row, column=col_idx).value
        if cell_val != current_val:
            if start_row < row - 1:
                ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=row - 1, end_column=col_idx)
            current_val = cell_val
            start_row = row
    if start_row < ws.max_row:
        ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=ws.max_row, end_column=col_idx)









def save_error_split_txt(error_df, save_dir=".", prefix="error"):
    calc_lines = []
    draw_lines = []

    for _, row in error_df.iterrows():
        member = str(row["Member"])
        direction = str(row["Direction"])
        detail = str(row["Detail"])

        line = f"{member} | {direction} | {detail}"

        if "in SCD" in detail or "in SCD" in row["Error Type"]:
            calc_lines.append(line)
        elif "in SD" in detail or "in SD" in row["Error Type"]:
            draw_lines.append(line)
        else:
            # 공통 오류는 둘 다 넣기
            calc_lines.append(line)
            draw_lines.append(line)

    with open(f"{save_dir}/{prefix}_calc_errors.txt", "w", encoding="utf-8") as f:
        f.write("\n".join(calc_lines))

    with open(f"{save_dir}/{prefix}_drawing_errors.txt", "w", encoding="utf-8") as f:
        f.write("\n".join(draw_lines))



# def copy_scd_error_images():
#     st.markdown("🔍 오류 기반 이미지 복사")


#     error_txt_path = st.text_input("📄 오류 텍스트 파일 경로", value="D:\codeforpaper\ground_truth_calc_errors.txt")
#     gt_excel_path = st.text_input("📊 구조계산서 GT 엑셀 경로", value="D:\SD_SCD_TRUE\SCD_ground_truth.csv")
#     figures_dir = st.text_input("🖼 Figures 폴더 경로", value="D:\codeforpaper/Figures")
#     plain_dir = st.text_input("📝 Plain_Texts 폴더 경로", value="D:\codeforpaper/Plain_Texts")
#     ocr_plain_dir = st.text_input("🔤 Plain OCR 결과 폴더", value="D:/codeforpaper/Surya_output")
#     ocr_figure_dir = st.text_input("🔤 Figure OCR 결과 폴더", value="D:/codeforpaper/Figure_OCR")
#     output_dir = st.text_input("📂 저장 폴더 경로", value="D:\codeforpaper\image_output_SCD")

#     if st.button("📤 오류 이미지 + OCR 복사 실행"):
#             try:
#                 def load_ground_truth_file(path):
#                     if path.lower().endswith(".csv"):
#                         return pd.read_csv(path)
#                     else:
#                         return pd.read_excel(path)

#                 with open(error_txt_path, "r", encoding="utf-8") as f:
#                     lines = f.readlines()

#                 error_entries = []
#                 for line in lines:
#                     parts = [x.strip() for x in line.strip().split("|")]
#                     if len(parts) == 3:
#                         error_entries.append(tuple(parts))

#                 gt_df = load_ground_truth_file(gt_excel_path)

#                 copied = 0
#                 unmatched = []

#                 for member, direction, detail in error_entries:
#                     match = gt_df[
#                         (gt_df["Member"].astype(str).str.strip() == member) &
#                         (gt_df["Direction"].astype(str).str.strip() == direction)
#                     ]

#                     if match.empty:
#                         st.warning(f"⚠️ 매칭 실패: {member} | {direction}")
#                         unmatched.append({
#                             "Member": member,
#                             "Direction": direction,
#                             "Detail": detail
#                         })
#                         continue

#                     id_val = str(match.iloc[0, 0])  # 첫 번째 열이 ID
#                     figure_files = glob.glob(os.path.join(figures_dir, f"SCD_{id_val}_figure.*"))
#                     plain_files = glob.glob(os.path.join(plain_dir, f"SCD_{id_val}_plain_*.png"))
#                     all_files = figure_files + plain_files

#                     safe_detail = detail.replace("/", "_").replace("\\", "_").strip()
#                     save_path = os.path.join(output_dir, safe_detail)
#                     os.makedirs(save_path, exist_ok=True)

#                     for file in all_files:
#                         filename = os.path.basename(file)
#                         dest = os.path.join(save_path, filename)
#                         shutil.copy(file, dest)
#                         copied += 1

#                         # OCR 파일도 함께 복사
#                         name_wo_ext = os.path.splitext(filename)[0]
#                         if "plain" in filename:
#                             ocr_src = os.path.join(ocr_plain_dir, f"{name_wo_ext}.json")
#                         elif "figure" in filename:
#                             ocr_src = os.path.join(ocr_figure_dir, f"{name_wo_ext}.json")
#                         else:
#                             ocr_src = None

#                         if ocr_src and os.path.exists(ocr_src):
#                             shutil.copy(ocr_src, os.path.join(save_path, os.path.basename(ocr_src)))

#                 st.success(f"✅ 총 {copied}개 이미지 및 OCR 파일 복사 완료!")

#                 if unmatched:
#                     unmatched_df = pd.DataFrame(unmatched)
#                     unmatched_path = os.path.join(output_dir, "unmatched_entries.csv")
#                     unmatched_df.to_csv(unmatched_path, index=False, encoding="utf-8-sig")
#                     st.info(f"⚠️ 매칭 실패 {len(unmatched)}건 → {unmatched_path} 저장됨")

#             except Exception as e:
#                 st.error(f"❌ 처리 중 오류 발생: {e}")


def copy_scd_error_images():
    st.markdown("🔍 오류 기반 이미지 복사")

    error_txt_path = st.text_input("📄 오류 텍스트 파일 경로", value=os.path.join(BASE_DIR,"ground_truth_calc_errors.txt"))
    gt_excel_path = st.text_input("📊 구조계산서 GT 엑셀 경로", value="D:/SD_SCD_TRUE/SCD_ground_truth.csv")
    figures_dir = st.text_input("🖼 Figures 폴더 경로", value=os.path.join(BASE_DIR,"Figures"))
    plain_dir = st.text_input("📝 Plain_Texts 폴더 경로", value=os.path.join(BASE_DIR, "Plain_Texts"))
    ocr_plain_dir = st.text_input("🔤 Plain OCR 결과 폴더", value=os.path.join(BASE_DIR,"Surya_output"))
    ocr_figure_dir = st.text_input("🔤 Figure OCR 결과 폴더", value=os.path.join(BASE_DIR,"Figure_OCR"))
    output_dir = st.text_input("📂 저장 폴더 경로", value=os.path.join(BASE_DIR,"image_output_SCD"))

    if st.button("📤 오류 이미지 + OCR 복사 실행"):
        try:
            def load_ground_truth_file(path):
                if path.lower().endswith(".csv"):
                    return pd.read_csv(path)
                else:
                    return pd.read_excel(path)

            with open(error_txt_path, "r", encoding="utf-8") as f:
                lines = f.readlines()

            error_entries = []
            for line in lines:
                parts = [x.strip() for x in line.strip().split("|")]
                if len(parts) == 3:
                    error_entries.append(tuple(parts))

            gt_df = load_ground_truth_file(gt_excel_path)
            copied = 0
            unmatched = []

            for member, direction, detail in error_entries:
                match = gt_df[
                    (gt_df["Member"].astype(str).str.strip() == member) &
                    (gt_df["Direction"].astype(str).str.strip() == direction)
                ]

                if match.empty:
                    st.warning(f"⚠️ 매칭 실패: {member} | {direction}")
                    unmatched.append({
                        "Member": member,
                        "Direction": direction,
                        "Detail": detail
                    })
                    continue

                id_val = str(match.iloc[0, 0])  # 첫 번째 열이 ID
                figure_files = glob.glob(os.path.join(figures_dir, f"SCD_{id_val}_figure.*"))
                plain_files = glob.glob(os.path.join(plain_dir, f"SCD_{id_val}_plain_*.png"))

                all_files = figure_files + plain_files

                safe_detail = detail.replace("/", "_").replace("\\", "_").strip()
                save_path = os.path.join(output_dir, safe_detail)
                os.makedirs(save_path, exist_ok=True)

                # member와 direction을 파일명으로 변환
                base_name = f"{member.replace(' ', '_')}_{direction.strip()}"

                for idx, file in enumerate(all_files):
                    ext = os.path.splitext(file)[1]
                    new_filename = f"{base_name}_{idx}{ext}"
                    dest = os.path.join(save_path, new_filename)
                    shutil.copy(file, dest)
                    copied += 1

                    # OCR 파일도 함께 복사
                    name_wo_ext = os.path.splitext(os.path.basename(file))[0]
                    if "plain" in name_wo_ext:
                        ocr_src = os.path.join(ocr_plain_dir, f"{name_wo_ext}.json")
                        ocr_dest = os.path.join(save_path, f"{base_name}_{idx}.json")
                    elif "figure" in name_wo_ext:
                        ocr_src = os.path.join(ocr_figure_dir, f"{name_wo_ext}.json")
                        ocr_dest = os.path.join(save_path, f"{base_name}_{idx}.json")
                    else:
                        ocr_src = None
                        ocr_dest = None

                    if ocr_src and os.path.exists(ocr_src):
                        shutil.copy(ocr_src, ocr_dest)

            st.success(f"✅ 총 {copied}개 이미지 및 OCR 파일 복사 완료!")

            if unmatched:
                unmatched_df = pd.DataFrame(unmatched)
                unmatched_path = os.path.join(output_dir, "unmatched_entries.csv")
                unmatched_df.to_csv(unmatched_path, index=False, encoding="utf-8-sig")
                st.info(f"⚠️ 매칭 실패 {len(unmatched)}건 → {unmatched_path} 저장됨")

        except Exception as e:
            st.error(f"❌ 처리 중 오류 발생: {e}")

# def copy_sd_error_images():
#     # 경로 설정
#     error_txt_path = "D:/codeforpaper/ground_truth_drawing_errors.txt"
#     gt_csv_path = "D:/SD_SCD_TRUE/SD_ground_truth.csv"  # ← 수정: 구조도면 GT CSV
#     sd_image_folder = "D:/codeforpaper/SD_cell_extraction/combined_crops"
#     sd_ocr_folder = "D:/codeforpaper/SD_cell_extraction/SD_img_Surya_output"
#     output_folder = "D:/codeforpaper/image_output_SD"

#     if st.button("📥 구조도면 오류 이미지 + OCR 복사 실행"):
#         try:
#             with open(error_txt_path, "r", encoding="utf-8") as f:
#                 lines = f.readlines()

#             errors = []
#             for line in lines:
#                 parts = [x.strip() for x in line.strip().split("|")]
#                 if len(parts) == 3:
#                     errors.append(tuple(parts))  # (member, direction, detail)

#             gt_df = pd.read_csv(gt_csv_path)

#             copied = 0
#             missed = []

#             for member, direction, detail in errors:
#                 match = gt_df[
#                     (gt_df["Member"].astype(str).str.strip() == member) &
#                     (gt_df["Direction"].astype(str).str.strip() == direction)
#                 ]

#                 if match.empty:
#                     missed.append((member, direction, detail))
#                     continue

#                 key = str(match.iloc[0, 0])  # 첫 번째 열: ID
#                 safe_detail = detail.replace("/", "_").replace("\\", "_").strip()
#                 error_dir = os.path.join(output_folder, safe_detail)
#                 os.makedirs(error_dir, exist_ok=True)

#                 # 이미지 복사
#                 img_src = os.path.join(sd_image_folder, key + ".png")
#                 img_dst = os.path.join(error_dir, key + ".png")
#                 if os.path.exists(img_src):
#                     shutil.copy(img_src, img_dst)

#                 # OCR 복사
#                 ocr_src = os.path.join(sd_ocr_folder, key + ".json")
#                 ocr_dst = os.path.join(error_dir, key + ".json")
#                 if os.path.exists(ocr_src):
#                     shutil.copy(ocr_src, ocr_dst)

#                 copied += 1

#             st.success(f"✅ 구조도면 이미지 및 OCR 총 {copied}건 복사 완료!")

#             if missed:
#                 missed_df = pd.DataFrame(missed, columns=["Member", "Direction", "Detail"])
#                 missed_path = os.path.join(output_folder, "unmatched_entries.csv")
#                 missed_df.to_csv(missed_path, index=False, encoding="utf-8-sig")
#                 st.warning(f"⚠️ 매칭 실패 {len(missed)}건 → {missed_path} 저장됨")

#         except Exception as e:
#             st.error(f"❌ 오류 발생: {e}")
def copy_sd_error_images():
    # 경로 설정
    error_txt_path = os.path.join(BASE_DIR,"ground_truth_drawing_errors.txt")
    gt_csv_path = "D:/SD_SCD_TRUE/SD_ground_truth.csv"
    sd_image_folder = os.path.join(BASE_DIR,"SD_cell_extraction/combined_crops")
    sd_ocr_folder = os.path.join(BASE_DIR,"SD_cell_extraction/SD_img_Surya_output")
    output_folder = os.path.join(BASE_DIR,"image_output_SD")

    if st.button("📥 구조도면 오류 이미지 + OCR 복사 실행"):
        try:
            # 오류 텍스트 파일 로딩
            with open(error_txt_path, "r", encoding="utf-8") as f:
                lines = f.readlines()

            errors = []
            for line in lines:
                parts = [x.strip() for x in line.strip().split("|")]
                if len(parts) == 3:
                    errors.append(tuple(parts))  # (member, direction, detail)

            # GT CSV 로딩
            gt_df = pd.read_csv(gt_csv_path)

            copied = 0
            missed = []

            for member, direction, detail in errors:
                match = gt_df[
                    (gt_df["Member"].astype(str).str.strip() == member) &
                    (gt_df["Direction"].astype(str).str.strip() == direction)
                ]

                if match.empty:
                    missed.append((member, direction, detail))
                    continue

                key = str(match.iloc[0, 0])  # 첫 번째 열(ID 열)
                safe_detail = detail.replace("/", "_").replace("\\", "_").strip()
                error_dir = os.path.join(output_folder, safe_detail)
                os.makedirs(error_dir, exist_ok=True)

                # 파일명: 예) 1_B2A_MID.png
                base_filename = f"{member.strip().replace(' ', '_')}_{direction.strip()}"

                # 이미지 복사
                img_src = os.path.join(sd_image_folder, key + ".png")
                img_dst = os.path.join(error_dir, base_filename + ".png")
                if os.path.exists(img_src):
                    shutil.copy(img_src, img_dst)

                # OCR 복사
                ocr_src = os.path.join(sd_ocr_folder, key + ".json")
                ocr_dst = os.path.join(error_dir, base_filename + ".json")
                if os.path.exists(ocr_src):
                    shutil.copy(ocr_src, ocr_dst)

                copied += 1

            st.success(f"✅ 구조도면 이미지 및 OCR 총 {copied}건 복사 완료!")

            if missed:
                missed_df = pd.DataFrame(missed, columns=["Member", "Direction", "Detail"])
                missed_path = os.path.join(output_folder, "unmatched_entries.csv")
                missed_df.to_csv(missed_path, index=False, encoding="utf-8-sig")
                st.warning(f"⚠️ 매칭 실패 {len(missed)}건 → {missed_path} 저장됨")

        except Exception as e:
            st.error(f"❌ 오류 발생: {e}")




def merge_error_files_without_missing():
    # 경로는 내부에서 지정
    calc_path = os.path.join(BASE_DIR,"ground_truth_calc_errors.txt")
    drawing_path = os.path.join(BASE_DIR,"ground_truth_drawing_errors.txt")
    output_path = os.path.join(BASE_DIR,"merged_errors_without_missing.txt")

    try:
        with open(calc_path, 'r', encoding='utf-8') as f:
            calc_lines = f.readlines()

        with open(drawing_path, 'r', encoding='utf-8') as f:
            drawing_lines = f.readlines()

        # 줄바꿈 제거 후 병합
        all_errors = [line.strip() for line in calc_lines + drawing_lines]

        # "Missing entry" 제거
        filtered_errors = [e for e in all_errors if "Missing entry" not in e]

        # 중복 제거
        unique_errors = sorted(set(filtered_errors))

        # 저장
        with open(output_path, 'w', encoding='utf-8') as f:
            for line in unique_errors:
                f.write(line + '\n')

        return f"✅ 총 {len(unique_errors)}개의 오류가 저장됨:\n{output_path}"
    
    except Exception as e:
        return f"❌ 오류 발생: {e}"




# def generate_final_comparison_report():
#      # === 경로 설정 ===
    # error_txt_path = "D:/codeforpaper/ground_truth_calc_errors.txt"
    # scd_gt_path = "D:/SD_SCD_TRUE/SCD_ground_truth.csv"
    # sd_gt_path = "D:/SD_SCD_TRUE/SD_ground_truth.csv"
    # scd_img_base = "D:/codeforpaper/image_output_SCD"
    # sd_img_base = "D:/codeforpaper/image_output_SD"
    # output_excel = "D:/codeforpaper/structured_error_report.xlsx"

    # def extract_plain_index(fname):
    #     match = re.search(r'_plain_(\d+)', fname)
    #     return int(match.group(1)) if match else -1

    # def get_scaled_dimensions(img_path, target_width):
    #     with Image.open(img_path) as im:
    #         w, h = im.size
    #         scale = target_width / w
    #         return scale, int(h * scale)

    # def get_total_scaled_height(imgs, target_width):
    #     total = 0
    #     for img in imgs:
    #         scale, h = get_scaled_dimensions(img, target_width)
    #         total += h + 5
    #     return total

    # # === 오류 로딩 ===
    # with open(error_txt_path, "r", encoding="utf-8") as f:
    #     lines = f.readlines()

    # errors = []
    # for line in lines:
    #     parts = [x.strip() for x in line.strip().split("|")]
    #     if len(parts) == 3:
    #         errors.append(tuple(parts))

    # scd_gt = pd.read_csv(scd_gt_path)
    # sd_gt = pd.read_csv(sd_gt_path)

    # workbook = xlsxwriter.Workbook(output_excel)
    # ws = workbook.add_worksheet("Error Report")

    # # 셀 크기 고정: B~G는 내용, A는 설명
    # ws.set_column("A:A", 25)
    # ws.set_column("B:D", 23)
    # ws.set_column("E:G", 23)
    # for r in range(1000):
    #     ws.set_row(r, 30)

    # # 서식 정의
    # title_format = workbook.add_format({
    #     'bold': True, 'align': 'center', 'valign': 'vcenter',
    #     'font_size': 16, 'border': 1
    # })
    # label_format = workbook.add_format({
    #     'bold': True, 'align': 'center', 'valign': 'vcenter',
    #     'font_size': 12, 'border': 1
    # })
    # error_format = workbook.add_format({
    #     'align': 'left', 'valign': 'vcenter',
    #     'font_size': 11, 'border': 1
    # })
    # border_format = workbook.add_format({'border': 1})

    # # 제목
    # ws.merge_range("B1:G1", "Error Report", title_format)
    # ws.merge_range("B2:D2", "Structural Calculation Document", label_format)
    # ws.merge_range("E2:G2", "Structural Drawing Document", label_format)
    # ws.write("A2", "Drawing with errors", label_format)

    # row = 3
    # valid_exts = (".png", ".jpg", ".jpeg")

    # for idx, (member, direction, detail) in enumerate(errors):
    #     safe_detail = detail.replace("/", "_").replace("\\", "_").strip()

    #     scd_match = scd_gt[
    #         (scd_gt["Member"].astype(str).str.strip() == member) &
    #         (scd_gt["Direction"].astype(str).str.strip() == direction)
    #     ]
    #     sd_match = sd_gt[
    #         (sd_gt["Member"].astype(str).str.strip() == member) &
    #         (sd_gt["Direction"].astype(str).str.strip() == direction)
    #     ]

    #     if scd_match.empty or sd_match.empty:
    #         continue

    #     scd_id = str(scd_match.iloc[0, 0])
    #     sd_key = str(sd_match.iloc[0, 0])

    #     scd_folder = os.path.join(scd_img_base, safe_detail)
    #     sd_folder = os.path.join(sd_img_base, safe_detail)

    #     scd_plain = sorted(
    #         glob.glob(os.path.join(scd_folder, f"SCD_{scd_id}_plain_*.png")),
    #         key=extract_plain_index, reverse=True
    #     )
    #     scd_figure = [
    #         f for f in glob.glob(os.path.join(scd_folder, f"SCD_{scd_id}_figure.*"))
    #         if f.lower().endswith(valid_exts)
    #     ]
    #     scd_images = scd_plain + scd_figure
    #     sd_image = os.path.join(sd_folder, f"{sd_key}.png")

    #     if not scd_images or not os.path.exists(sd_image):
    #         continue

    #     # 이미지 높이 계산
    #     scd_height = get_total_scaled_height(scd_images, 300)
    #     sd_scale, sd_height = get_scaled_dimensions(sd_image, 300)
    #     total_px = max(scd_height, sd_height)
    #     rows_needed = int(total_px / 30) + 1
    #     start_row = row

    #     # 이미지 삽입 - 구조계산서
    #     offset_y = 0
    #     for img in scd_images:
    #         scale, h = get_scaled_dimensions(img, 300)
    #         ws.insert_image(start_row, 1, img, {
    #             'x_scale': scale, 'y_scale': scale,
    #             'x_offset': 0, 'y_offset': offset_y
    #         })
    #         offset_y += h + 5

    #     # 이미지 삽입 - 구조도면
    #     ws.insert_image(start_row, 4, sd_image, {
    #         'x_scale': sd_scale, 'y_scale': sd_scale,
    #         'x_offset': 0, 'y_offset': 0
    #     })

    #     # 오류 설명 (A열)
    #     ws.write(start_row, 0, "Error Type", border_format)
    #     ws.write(start_row + 1, 0, detail.split()[0], error_format)
    #     ws.write(start_row + 2, 0, "Member", border_format)
    #     ws.write(start_row + 3, 0, member, error_format)
    #     ws.write(start_row + 4, 0, "Direction", border_format)
    #     ws.write(start_row + 5, 0, direction, error_format)
    #     ws.write(start_row + 6, 0, "Detail", border_format)
    #     ws.write(start_row + 7, 0, detail, error_format)

    #     row = start_row + rows_needed + 3

    # workbook.close()
    # print(f"✅ 보고서 저장 완료: {output_excel}")




def resize_image_keep_aspect(img_path, target_width_px, target_height_px):
    try:
        with PILImage.open(img_path) as pil_img:
            w, h = pil_img.size
            ratio = min(target_width_px / w, target_height_px / h)
            return int(w * ratio), int(h * ratio)
    except Exception as e:
        print(f"❌ [resize_image_keep_aspect] 에러: {img_path}, {e}")
        return None, None

def add_padding_top_left(img_path, pad_left=30, pad_top=20):
    try:
        img = PILImage.open(img_path)
        w, h = img.size
        new_w, new_h = w + pad_left, h + pad_top
        padded = PILImage.new("RGB", (new_w, new_h), (255, 255, 255))
        padded.paste(img, (pad_left, pad_top))
        padded_path = img_path.replace(".png", "_padded.png")
        padded.save(padded_path)
        return padded_path
    except Exception as e:
        print(f"❌ [add_padding_top_left] 에러: {img_path}, {e}")
        return None

def add_black_border(img_path, border_width=3):
    try:
        img = PILImage.open(img_path)
        w, h = img.size
        draw = ImageDraw.Draw(img)
        for i in range(border_width):
            draw.rectangle([i, i, w - 1 - i, h - 1 - i], outline="black")
        bordered_path = img_path.replace(".png", "_bordered.png")
        img.save(bordered_path)
        return bordered_path
    except Exception as e:
        print(f"❌ [add_black_border] 에러: {img_path}, {e}")
        return None

def filter_calc_images_by_group(img_list, member_key, direction_key):
    direction_key = direction_key.strip()
    prefix = f"{member_key}_{direction_key}"
    matched = [p for p in img_list if os.path.basename(p).startswith(prefix)]

    group_map = defaultdict(list)
    for img_path in matched:
        fname = os.path.basename(img_path)
        match = re.search(r'_(\d+)\.png$', fname)
        if match:
            group_idx = int(match.group(1))
            group_map[group_idx].append(img_path)

    sorted_imgs = [group_map[idx][0] for idx in sorted(group_map, reverse=True)]
    return sorted_imgs

def merge_images_vertically(image_paths, save_path, padding=20, scale_factor=1.2):
    try:
        if not image_paths:
            return None
        images = [PILImage.open(p) for p in image_paths]
        widths = [img.width for img in images]
        heights = [img.height for img in images]

        total_height = sum(heights) + padding * (len(images) - 1)
        max_width = max(widths)
        new_img = PILImage.new("RGB", (max_width, total_height), (255, 255, 255))

        y_offset = 0
        for img in images:
            new_img.paste(img, (0, y_offset))
            y_offset += img.height + padding

        if scale_factor != 1.0:
            new_size = (int(new_img.width * scale_factor), int(new_img.height * scale_factor))
            new_img = new_img.resize(new_size, PILImage.LANCZOS)

        new_img.save(save_path)
        return save_path
    except Exception as e:
        print(f"❌ [merge_images_vertically] 에러: {save_path}, {e}")
        return None

def filter_drawing_images_by_member(img_list, member_key):
    return [p for p in sorted(img_list) if os.path.basename(p).startswith(member_key)]

def apply_outer_border(ws, start_row, end_row, start_col, end_col):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = thin_border

def generate_error_reporting_excel(
    error_txt_path=os.path.join(BASE_DIR,"merged_errors_without_missing.txt"),
    scd_img_root=os.path.join(BASE_DIR,"image_output_SCD"),
    sd_img_root=os.path.join(BASE_DIR,"image_output_SD"),
    output_excel_path=os.path.join(BASE_DIR,"error_reporting_structured_final.xlsx")
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Error Report"

    title_font = Font(size=15, bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for col, width in zip(["A", "B", "C", "D", "E", "F"], [5, 15, 40, 40, 40, 40]):
        ws.column_dimensions[col].width = width

    row = 5
    excel_col_to_px = 7.5
    excel_row_to_px = 1.3

    grouped_errors = defaultdict(lambda: defaultdict(list))
    with open(error_txt_path, "r", encoding="utf-8") as f:
        for line in f:
            parts = [x.strip() for x in line.strip().split("|")]
            if len(parts) == 3:
                member, direction, detail = parts
                grouped_errors[member][direction].append(detail)

    for member, dir_dict in grouped_errors.items():
        member_key = member.replace(" ", "_")

        for direction, details in dir_dict.items():
            direction_key = direction.strip()
            error_text = f"{member} | {direction} | " + ", ".join(details)
            error_folder = details[0].replace("/", "_").replace("\\", "_")

            # 구분 라벨 삽입
            b_labels = ["구분", "도면 이미지", "오류 내역"]
            for i, label in enumerate(b_labels):
                ws.cell(row=row + i, column=2, value=label).font = title_font
                ws.cell(row=row + i, column=2).alignment = center_align
            apply_outer_border(ws, row, row + 2, 2, 2)

            scd_width_px = (ws.column_dimensions["C"].width + ws.column_dimensions["D"].width) * excel_col_to_px
            sdd_width_px = (ws.column_dimensions["E"].width + ws.column_dimensions["F"].width) * excel_col_to_px
            img_row_height_px = 300 * excel_row_to_px

            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
            ws.cell(row=row, column=3, value="SCD").font = title_font
            ws.cell(row=row, column=3).alignment = center_align

            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
            ws.cell(row=row, column=5, value="SDD").font = title_font
            ws.cell(row=row, column=5).alignment = center_align

            try:
                all_scd_imgs = glob.glob(os.path.join(scd_img_root, error_folder, "*.png"))
                filtered_scd_imgs = [p for p in all_scd_imgs if os.path.basename(p).startswith(f"{member_key}_{direction_key}")]
                selected_calc_imgs = filter_calc_images_by_group(filtered_scd_imgs, member_key, direction_key)
                merged_img_temp = f"D:/codeforpaper/temp_merged_calc_{member_key}_{direction_key}.png"
                merged_calc_img = merge_images_vertically(selected_calc_imgs, merged_img_temp)

                if merged_calc_img and os.path.exists(merged_calc_img):
                    padded_img_path = add_padding_top_left(merged_calc_img)
                    final_img_path = add_black_border(padded_img_path)
                    if final_img_path and os.path.exists(final_img_path):
                        img = XLImage(final_img_path)
                        resized_w, resized_h = resize_image_keep_aspect(final_img_path, scd_width_px, img_row_height_px)
                        if resized_w and resized_h:
                            img.width = resized_w
                            img.height = resized_h
                            ws.merge_cells(start_row=row + 1, start_column=3, end_row=row + 1, end_column=4)
                            ws.add_image(img, f"C{row + 1}")
            except Exception as e:
                print(f"⚠️ SCD 이미지 오류: {e}")

            try:
                all_sd_imgs = glob.glob(os.path.join(sd_img_root, error_folder, f"{member_key}_*.png"))
                valid_sd_imgs = filter_drawing_images_by_member(all_sd_imgs, member_key)
                if valid_sd_imgs:
                    padded_sd_path = add_padding_top_left(valid_sd_imgs[0])
                    final_sd_img_path = add_black_border(padded_sd_path)
                    if final_sd_img_path and os.path.exists(final_sd_img_path):
                        img = XLImage(final_sd_img_path)
                        resized_w, resized_h = resize_image_keep_aspect(final_sd_img_path, sdd_width_px, img_row_height_px)
                        if resized_w and resized_h:
                            img.width = resized_w
                            img.height = resized_h
                            ws.merge_cells(start_row=row + 1, start_column=5, end_row=row + 1, end_column=6)
                            ws.add_image(img, f"E{row + 1}")
            except Exception as e:
                print(f"⚠️ SDD 이미지 오류: {e}")

            ws.merge_cells(start_row=row + 2, start_column=3, end_row=row + 2, end_column=6)
            ws.cell(row=row + 2, column=3, value=error_text).font = title_font
            ws.cell(row=row + 2, column=3).alignment = center_align

            apply_outer_border(ws, row, row + 2, 3, 6)

            ws.row_breaks.append(Break(id=row + 2))

            ws.row_dimensions[row + 1].height = 300
            ws.row_dimensions[row + 2].height = 60
            row += 5

    wb.save(output_excel_path)
    print(f"COMPLETE: {output_excel_path}")























# 📌 Streamlit 메인 탭 구성
# ==========================
st.title("Human Error Detection")

tab1, tab2, tab3 = st.tabs(["SCD info extraction", "SDD info extraction", "Human Error detection"])

# ----------------------------


with tab1:
    st.header("Structural Calculation Documents Info Extraction Pipeline")
    
    # 순서 변경: 먼저 PDF 업로드
    with st.expander("Step 1: Upload PDF"):
        uploaded_pdfs = st.file_uploader("select your pdf", type=["pdf"], accept_multiple_files=True)
        if uploaded_pdfs:
            st.success(f"{len(uploaded_pdfs)} uploaded")
            
            # PDF 업로드 후 바로 이미지 변환 실행
            if st.button("Convert PDF to Images"):
                with st.spinner("Converting PDF to images..."):
                    image_paths = process_uploaded_images(uploaded_pdfs)
                    st.session_state.image_paths = image_paths  # 세션에 저장
                st.success("PDF conversion completed")
                
                # 첫 번째 이미지를 예시로 표시
                if image_paths:
                    first_image_path = image_paths[0]
                    st.markdown("**Example image for keyword reference:**")
                    
                    # 이미지 파일이 존재하는지 확인
                    if os.path.exists(first_image_path):
                        try:
                            # PIL을 사용하여 이미지 로드 후 표시
                            from PIL import Image
                            img = Image.open(first_image_path)
                            st.image(img, caption="Sample document image", width=600)
                        except Exception as e:
                            st.error(f"이미지 로딩 실패: {e}")
                            st.write(f"이미지 경로: {first_image_path}")
                    else:
                        st.error(f"이미지 파일을 찾을 수 없습니다: {first_image_path}")
                        
                    # 디버깅을 위한 추가 정보
                    st.write(f"변환된 이미지 개수: {len(image_paths)}")
                    st.write("변환된 이미지 경로 (처음 5개):")
                    for i, path in enumerate(image_paths[:5]):
                        exists = "✅" if os.path.exists(path) else "❌"
                        st.write(f"{exists} {i+1}: {path}")

    # 변환된 이미지가 있으면 항상 표시 (expander 밖에서)
    if hasattr(st.session_state, 'image_paths') and st.session_state.image_paths:
        st.markdown("---")
        st.markdown("### Sample Image for Keyword Reference")
        first_image_path = st.session_state.image_paths[0]
        if os.path.exists(first_image_path):
            try:
                from PIL import Image
                img = Image.open(first_image_path)
                st.image(img, caption="Sample document image", width=600)
            except Exception as e:
                st.error(f"이미지 로딩 실패: {e}")
        
        st.markdown("---")
        
        with st.expander("Step 2: Enter Keywords"):
            st.markdown("Please enter keywords based on the sample image shown above:")
            
            # 기존 키워드 값이 있으면 기본값으로 사용
            existing_keywords = getattr(st.session_state, 'keywords', {})
            
            code_kw = st.text_input(
                "Member Code Keyword", 
                value=existing_keywords.get('code_kw', ''),
                help="Enter the keyword that appears before member codes in your document"
            )
            scope_kw = st.text_input(
                "Section Property Keyword", 
                value=existing_keywords.get('scope_kw', ''),
                help="Enter the keyword that appears before section properties"
            )
            rebar_kw = st.text_input(
                "Rebar Pattern Keyword", 
                value=existing_keywords.get('rebar_kw', ''),
                help="Enter the keyword that appears before rebar patterns"
            )
            material_kw = st.text_input(
                "Material Data Keyword", 
                value=existing_keywords.get('material_kw', ''),
                help="Enter the keyword that appears before material data"
            )
            
            # 키워드 저장 버튼
            if st.button("Save Keywords"):
                if all([code_kw, scope_kw, rebar_kw, material_kw]):
                    st.session_state.keywords = {
                        'code_kw': code_kw,
                        'scope_kw': scope_kw, 
                        'rebar_kw': rebar_kw,
                        'material_kw': material_kw
                    }
                    st.success("All keywords saved!")
                else:
                    st.error("Please fill in all keyword fields")
            
            # 현재 저장된 키워드 표시
            if hasattr(st.session_state, 'keywords'):
                st.write("**Currently saved keywords:**")
                for key, value in st.session_state.keywords.items():
                    st.write(f"• {key}: {value}")


    else:
        st.info("Please upload and convert PDF files first to see the keyword input section.")





    with st.expander("Layout analysis"):
        if st.button("Run Layout Analysis"):
            if not uploaded_pdfs or not all([code_kw, scope_kw, rebar_kw, material_kw]):
                st.error("Check your input")
            else:
                with st.spinner("During analyzing"):
                    image_paths = process_uploaded_images(uploaded_pdfs)
                    outputs = apply_yolo_on_images(image_paths)
                st.success("✅ Complete")
                st.code(f"{len(image_paths)} complete")

    with st.expander("Plain Text Step1 : OCR"):
        if st.button("Run OCR"):
            with st.spinner("During OCR"):
                apply_surya_ocr()

    with st.expander("Plain Text Step 2 : OCR Filtering"):
        if st.button("Run filtering"):
            with st.spinner("During filtering"):
                apply_filtering()

    with st.expander("Plain Text Step 3 : Info extraction based keyword"):
        if st.button("Run keyword extraction"):
            if not all([code_kw, scope_kw, rebar_kw]):
                st.error("enter yout all ketword")
            else:
                with st.spinner("During extartion"):
                    apply_keyword_extraction()


    with st.expander("Figure Step 1 : OCR"):
        st.markdown("**🖼️ Applying OCR to figure.**")
        if st.button("📌 Run Figure OCR"):
            with st.spinner("applying ocr"):
                apply_surya_ocr_to_figures()

    with st.expander("Figure Step 2 :  OCR filtering and RC extraction"):
        st.markdown("Figure OCR filtering and RC extraction**🧹 Filter the OCR result file and start extracting RC members**")
        if st.button("🧩 Run Figure Filtering & RC extraction"):
            with st.spinner("Cleaning up OCR results...."):
                rename_and_move_ocr_results()
                filter_and_merge_figure_ocr()
                extract_rc_elements()
            st.success("✅RC element extraction has been completed.")

    with st.expander("Figure Step 3 : Rotation and crop for height extraction"):
        if st.button("Rotation and crop", key="rc_figure_1"):
            with st.spinner("Extracting Table images for RC_COLUMN pages..."):
                rotate_and_crop_images()

            st.success("✅ Rotated and truncated! saved in rotated_crop_folder)")



    with st.expander("Figure Step 4: OCR for height"):
        if st.button("Run OCR on Vertical Cropped Images", key="rc_figure_2"):
            with st.spinner("Executing OCR and cleaning JSON..."):
                OCR_figure_rotated_cropped_figure()
            st.success("✅ Vertical image OCR completed and results cleaned up")


    with st.expander("Figure Step 5: height substitution "):
        if st.button("height substitution", key="rc_figure_3"):
            with st.spinner("During substitution"):
                extract_vertical_texts()
                update_height_from_direction_file()




    with st.expander("Plain text-stirrup Step 1 : RC Column Table Png extraction"):
        if st.button("Run RC Table Image Extract", key="rc_step1"):
            with st.spinner("Extracting Table images for RC_COLUMN pages..."):
                extract_rc_column_jsons_from_ocr_results()
            st.success("✅ Table Image Copy Completed! saved in {Table_OCR}")

    with st.expander("Plain text-stirrup Step 2 : RC Column Stirrup Info extraction"):
        if st.button("Extraction Info", key="rc_step2"):
            with st.spinner("Extracting Table images for RC_COLUMN pages..."):
                extract_stirrup_rebar_patterns_from_json_surya()
            st.success("✅ Stirrup information extraction complete! saved in {RC_COLUMN_REBAR_TXT}")

    with st.expander("Plain text-stirrup Step 3 : Substitution"):
        if st.button("Substitute stirrup Info", key="rc_step3"):
            with st.spinner("running substitution"):
                apply_stirrup_to_rc_columns()
            st.success("✅ Done")



    with st.expander("Make CSV 1 : All informatiion extraction"):
        if st.button("Information extracion", key="extracion_step1"):
            with st.spinner("running extraion"):
                extract_member_data()
            st.success("✅ Done")

    with st.expander("Make CSV Step 2 : combine information"):
        if st.button("Information combaine", key="extracion_step2"):
            with st.spinner("running extraion"):
                extract_element_data()
            st.success("✅ Done")



    with st.expander("Standardization Step 1 : csv file"):
        if st.button("Make a csv files", key="standardization_step1"):
            with st.spinner("running extraion"):
                combine_data_to_csv()
            st.success("✅ Done")            

    with st.expander("Standardization Step 2 : csv file"):
        st.markdown("Enter below if the PIT layer is considered a number (optional)")
        user_input = st.text_input("How many floors should we consider the PIT layer? (e.g. 3)", key="pit_input_step2")

        if st.button("standardization", key="standardization_step2"):
            with st.spinner("running standardization"):
                standardize_member_split()

                try:
                    max_floor = int(user_input.strip()) if user_input.strip() else None
                except ValueError:
                    max_floor = None  # 잘못된 값은 무시

                normalize_member_column(max_floor=max_floor)

            st.success("✅ Done")

#################################################################################################
#################################################################################################





















with tab2:
    st.header("Structural Drawing Documents Info Extraction Pipeline")

    st.session_state.setdefault("pdf_converted", False)
    st.session_state.setdefault("image_paths", [])
    st.session_state.setdefault("show_canvas", False)
    st.session_state.setdefault("bbox_complete", False)

    with st.expander("Step 1 : PDF → PNG"):
        uploaded_pdfs = st.file_uploader("Automatic conversion when PDF file is selected", type=["pdf"], accept_multiple_files=True)

        if uploaded_pdfs and not st.session_state.pdf_converted:
            with st.spinner("Processing..."):
                image_paths = process_uploaded_images_SD(uploaded_pdfs)
                st.session_state.image_paths = image_paths
                st.session_state.pdf_converted = True  # ✅ 변환 완료 저장
            st.success(f"✅ total {len(image_paths)} image conversion completed!")
            st.markdown("**Conversion Results (Top 5 Path Preview)**")
            st.code("\n".join(image_paths[:5]), language="text")

        elif st.session_state.pdf_converted:
            st.success(f"✅ 변환 완료된 이미지 수: {len(st.session_state.image_paths)}")
            st.code("\n".join(st.session_state.image_paths[:5]), language="text")

    with st.expander("📐 Step 2 : Draw Bounding Box and Check Results"):
        if not st.session_state.show_canvas:
            if st.button("📌 바운딩 박스 입력 시작"):
                st.session_state.show_canvas = True
                st.experimental_rerun()

        if st.session_state.show_canvas:
            img_path, boxes = draw_and_save_bounding_boxes()

            if img_path and boxes:
                st.success("✅ 바운딩 박스 저장 완료!")
                st.markdown(f"**📄 파일:** `{os.path.basename(img_path)}`")
                st.json(boxes)

                if st.button("✅ 바운딩 박스 입력 완료", key="bbox_done"):
                    st.session_state["bbox_complete"] = True
                    st.success("🎉 모든 작업 완료! 다음 단계로 이동 가능합니다.")


    with st.expander("Step 3 : Member code Recognition"):
        if st.button("Member code Recognition", key="Member code Recognition"):
            with st.spinner("running..."):
                apply_surya_ocr_SD()
                match_ocr_to_bboxes_all()
            st.success("✅ Done")


    # with st.expander("Step 3.5 : Line Detection Parameter Tuning (Optional)"):
    #     st.markdown("이미지별로 최적의 선 검출 파라미터를 찾아 저장하는 단계입니다.")
    #     interactive_line_detection_interface()


    with st.expander("Step 4: 통합 선검출 & 셀매칭 디버깅"):
        st.markdown("선 검출부터 텍스트 매칭까지 전체 과정을 실시간으로 확인")
        combined_line_and_cell_debug()



    with st.expander("Step 5 : 크롭된 이미지 대상 OCR 적용"):
        if st.button("apply ocr to cropped image", key="cropped image ocr"):
            with st.spinner("running..."):
                apply_surya_ocr_on_combined()
            st.success("✅ Done")


    with st.expander("Step 7 : Elements allocation"):
        if st.button("elements allocation", key="allocation"):
            with st.spinner("running..."):
                annotate_rebar_info()
            st.success("✅ Done")

    with st.expander("Step 8 : clustering"):
        if st.button("elements clustering", key="clustering"):
            with st.spinner("running..."):
                results = annotate_rebar_info()
                export_merged_excel(results)
            st.success("✅ 병합 Excel 저장 완료!")




    with st.expander("Step 9 : rotation"):
        if st.button("run rotation", key="rotatata"):
            with st.spinner("running..."):
                annotate_rebar_info2()
            st.success("✅ Done! 모든 이미지 회전 완료")


    with st.expander("Step 10 : Apply Surya OCR to Rotated Images"):
        if st.button("Run Suryasda OCR", key="ocr_roadatated"):
            with st.spinner("Running OCR on rotated images..."):
                apply_surya_ocr_on_rotated_images()



    with st.expander("Step 11 : Visualization"):
                visualize_and_extract_ranked_heights()


    with st.expander("Step 12 :memerge height"):
        replace_height_in_merged_elements()



    with st.expander("Step 13 : Standardization"):
        # ✅ PIT 층수 입력창 추가
        pit_input = st.text_input("PIT 층수를 입력하세요 (비우면 'PIT'로 유지됨)", key="pit_input")

        # 입력값 처리 → 정수면 변환, 아니면 None
        pitf_max = int(pit_input) if pit_input.strip().isdigit() else None

        if st.button("ellocation", key="allotion"):
            with st.spinner("running..."):
                standardize_member_column(pitf_max=pitf_max)
            st.success("✅ Done")





    with st.expander("Step 14 : Standardization 2"):
        if st.button("standardization2", key="standardization2"):
            with st.spinner("running..."):
                expand_direction_column()
            st.success("✅ Done")
 ###############################################################################


 
with tab3:
    st.header("Human Error Detection Pipeline")


    with st.expander("Step 1: Structural calculation document_Information extraction accuracy evaluation"):
        evaluate_extraction_accuracy_SCD()

    with st.expander("Step 2: Structural drawing document_Information extraction accuracy evaluation"):
        evaluate_extraction_accuracy_SD()


    with st.expander("Step 3 : Human error detection"):
        save_path = st.text_input("💾 result storage path (.xlsx)", 
                                value=r"D:\codeforpaper\SD_cell_extraction\ground_truth_error_result.xlsx")

        if st.button("Start"):
            if "df_gt_scd" not in st.session_state or "df_gt_sd" not in st.session_state:
                st.warning("⚠️ 먼저 SCD 및 SD Ground Truth 파일을 업로드하고 정확도 평가를 실행해주세요.")
            else:
                try:
                    generate_ground_truth_error_report(
                        st.session_state.df_gt_scd.copy(),
                        st.session_state.df_gt_sd.copy(),
                        save_path
                    )
                    st.success("DONE!")

                    # ✅ 엑셀 시트 읽기 및 화면 표시
                    detailed_errors_df = pd.read_excel(save_path, sheet_name="Detailed Errors")
                    summary_df = pd.read_excel(save_path, sheet_name="Error Summary")

                    st.subheader("Error detail")
                    st.dataframe(detailed_errors_df, use_container_width=True)

                    st.subheader("Error summary")
                    st.dataframe(summary_df, use_container_width=True)

                except Exception as e:
                    st.error(f"❌ 보고서 생성 실패: {e}")




    with st.expander("Step 4 : Error reporting_1_error intergration"):
        uploaded_error_file = st.file_uploader("📥 upload error detection excel file", type=["xlsx"])
        save_dir = st.text_input("📂 storage path", value="D:\codeforpaper")
        prefix = st.text_input("📄 file name", value="ground_truth")

        if st.button("📤 Save in text file"):
            if not uploaded_error_file:
                st.warning("❗ 먼저 오류 결과 엑셀 파일을 업로드해주세요")
            else:
                try:
                    error_df = pd.read_excel(uploaded_error_file, sheet_name="Detailed Errors")
                    save_error_split_txt(error_df, save_dir, prefix)
                    st.success(f"✅ 저장 완료: {prefix}_calc_errors.txt, {prefix}_drawing_errors.txt")
                except Exception as e:
                    st.error(f"❌ 처리 중 오류 발생: {e}")

    with st.expander("Step 5 : Error reporting_2_scd_error image copy"):
        copy_scd_error_images()  # 호출만 하면 내부에서 UI + 로직 다 처리

    with st.expander("Step 6 : Error reporting_3_sdd_error image copy"):
        copy_sd_error_images()



    with st.expander("Step 7 : Generation Error reports"):
        if st.button("generate reports"):
            generate_error_reporting_excel()
            st.success("Done!")
